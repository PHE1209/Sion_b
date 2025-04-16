from datetime import date, datetime, timedelta
from decimal import Decimal
import base64
import csv
import io
import json
import logging
import os
from io import BytesIO

# Bibliotecas para gráficos y datos
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib.ticker import FuncFormatter, LogLocator, MaxNLocator

# Django - Autenticación y Decoradores
from django.apps import apps
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import User

# Django - Funcionalidades adicionales
from django.core.files.storage import FileSystemStorage
from django.core.paginator import Paginator
from django.db.models import (
    CharField, DateField, DateTimeField, DecimalField, FloatField,
    ForeignKey, Q, Sum, TextField, F
)
from django.forms.models import model_to_dict
from django.http import HttpResponse, JsonResponse, StreamingHttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import TemplateView

# ReportLab
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfgen import canvas
from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle

from django.db import models

# OpenPyXL
from openpyxl import Workbook

# Formularios y Modelos específicos del proyecto
from .forms import (
    GranulometriaForm, HumedadForm, LoginForm, MuestreoForm, NominaForm,
    ProgramaForm, ProyectoEditForm, ProspeccionesForm, GravedadEspecificaForm, UscsForm
)
from .models import (
    Area, Granulometria, Humedad, JornadaTeorica, Muestreo, Nomina,
    Programa, Prospecciones, Proyectos, Roster, gravedad_especifica, uscs
)


import logging
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from .models import Proyectos, Muestreo  # Asegúrate de importar tus modelos
from .forms import HumedadForm  # Asegúrate de tener este formulario definido
import logging
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from .models import Proyectos, Muestreo, Humedad, Prospecciones
import logging
import json
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from io import BytesIO
import base64
from django.shortcuts import render
from django.http import JsonResponse
from .models import uscs, Proyectos, Muestreo, Prospecciones
import logging
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.http import JsonResponse
from .models import Proyectos, Muestreo, gravedad_especifica
import logging
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.http import JsonResponse
from .models import Proyectos, Muestreo, uscs
import logging
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.db.models import Q
from django.apps import apps
from django.core.paginator import Paginator
from django.contrib.auth.models import User
from administrador.models import Limites_atterberg, Proyectos, Prospecciones, Muestreo
from administrador.forms import LimitesAtterbergForm
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from io import BytesIO
import base64
import json
import random
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
from io import BytesIO
import base64
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.http import JsonResponse
from administrador.models import gravedad_especifica  # Nombre correcto del modelo
import json
import random
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404
from .models import Prospecciones
from pyproj import Proj, transform
import logging
import logging
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from .models import ProspeccionImage, Prospecciones, Historial
from django.utils import timezone
import logging
import random
from io import BytesIO
import base64
import json
import pandas as pd
import matplotlib.pyplot as plt
from django.apps import apps
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import models
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from .models import Proyectos, Prospecciones
from .forms import ProspeccionesForm  # Asumo que tienes un formulario para Prospecciones
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.http import JsonResponse
from .models import Prospecciones, Proyectos, ProspeccionImage
import logging



# Create your views here.
TEMPLATE_DIR = (
    'os.path.join(BASE_DIR, "templates),'
)

#### INICIO DE SESION  #############################################################################

@login_required
def index_view(request):
    return render(request, 'index.html')

# Inicio de sesión
def home_view(request):
    return render(request, 'home.html')

# Vista de inicio de sesión
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login
from django.contrib import messages
from django.contrib.auth.models import User
from .forms import LoginForm
import logging

# Configurar logger
logger = logging.getLogger(__name__)

def login_view(request):
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            password = form.cleaned_data['password']
            try:
                logger.info(f"Intentando autenticar usuario con email: {email}")
                user = User.objects.get(email=email)
                logger.info(f"Usuario encontrado: {user.username}")
                user = authenticate(request, username=user.username, password=password)
                if user is not None:
                    login(request, user)
                    logger.info(f"Autenticación exitosa para {user.username}")
                    return redirect('index')
                else:
                    logger.warning(f"Autenticación fallida para email: {email}")
                    form.add_error(None, 'Contraseña incorrecta.')
            except User.DoesNotExist:
                logger.error(f"Email no registrado: {email}")
                form.add_error(None, 'Email no registrado.')
            except Exception as e:
                logger.error(f"Error en login_view: {str(e)}", exc_info=True)
                form.add_error(None, f'Error del servidor: {str(e)}')
        else:
            logger.warning("Formulario inválido")
            form.add_error(None, 'Formulario inválido.')
    else:
        form = LoginForm()
    return render(request, 'login.html', {'form': form})



# Vista protegida usando decorador
@login_required
def mi_vista_protegida(request):
    return render(request, 'index_master.html')

# Vista protegida usando LoginRequiredMixin
class MiVistaProtegida(LoginRequiredMixin, TemplateView):
    template_name = 'index.html'

# Vista index
@login_required
def index_view(request):
    return render(request, 'index.html')






#### PROYECTO ##############################################

# Agregar proyecto
@login_required
def agregar_proyectos(request):
    if request.method == 'POST':
        form = ProyectoEditForm(request.POST)
        if form.is_valid():
            proyecto = form.save(commit=False)
            proyecto.user = request.user
            proyecto.save()
            return redirect('listar_proyectos')
    else:
        form = ProyectoEditForm()
    return render(request, 'proyectos/agregar_proyectos.html', {'form': form})

# Listar proyecto
@login_required
def listar_proyectos(request):
    proyectos = Proyectos.objects.all().order_by('id')
    query = request.GET.get('q', '')
    if query:
        Proyectos_model = apps.get_model('administrador', 'Proyectos')
        fields = [field.name for field in Proyectos_model._meta.get_fields() if isinstance(field, (CharField, TextField, DateField, DateTimeField))]
        query_filter = Q()
        for field in fields:
            if isinstance(Proyectos_model._meta.get_field(field), (DateField, DateTimeField)):
                query_filter |= Q(**{f"{field}__year__icontains": query})
                query_filter |= Q(**{f"{field}__month__icontains": query})
                query_filter |= Q(**{f"{field}__day__icontains": query})
            else:
                query_filter |= Q(**{f"{field}__icontains": query})
        proyectos = proyectos.filter(query_filter)

    paginator = Paginator(proyectos, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'proyectos/proyectos_table.html', {'page_obj': page_obj})
    return render(request, 'proyectos/listar_proyectos.html', {'page_obj': page_obj, 'query': query})

# Ver proyecto
@login_required
def ver_proyectos(request, id):
    proyecto = get_object_or_404(Proyectos, id=id)
    return render(request, 'proyectos/ver_proyectos.html', {'proyecto': proyecto})

# Eliminar proyecto
@login_required
def eliminar_proyectos(request, id):
    proyecto = get_object_or_404(Proyectos, id=id)
    if request.method == 'POST':
        proyecto.delete()
        return redirect('listar_proyectos')
    return render(request, 'proyectos/eliminar_proyectos.html', {'proyecto': proyecto})

# Editar proyecto
@login_required
def editar_proyectos(request, id):
    proyecto = get_object_or_404(Proyectos, id=id)
    if request.method == 'POST':
        proyecto.pm = request.POST.get('pm')
        proyecto.empresa = request.POST.get('empresa')
        proyecto.nombre = request.POST.get('nombre')
        proyecto.alcance = request.POST.get('alcance')
        fecha_inicio = request.POST.get('fecha_inicio')
        fecha_termino = request.POST.get('fecha_termino')
        if fecha_inicio:
            proyecto.fecha_inicio = pd.to_datetime(fecha_inicio, format='%Y-%m-%d').date()
        if fecha_termino:
            proyecto.fecha_termino = pd.to_datetime(fecha_termino, format='%Y-%m-%d').date()
        proyecto.save()
        return redirect('listar_proyectos')
    return render(request, 'proyectos/editar_proyectos.html', {'proyecto': proyecto})

# Exportar a Excel (adaptado de granulometría)
@login_required
def export_to_excel_proyectos(request):
    query = request.GET.get('q', '')
    headers_str = request.GET.get('headers', '')
    headers = headers_str.split(',') if headers_str else [
        'ID', 'Estatus Proyecto', 'PM', 'Empresa', 'Nombre', 'Fecha Inicio', 'Fecha Término', 'Alcance', 'Usuario'
    ]

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="proyectos_filtrados.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Proyectos Filtrados"
    ws.append(headers)

    proyectos = Proyectos.objects.all().order_by('id')
    if query:
        Proyectos_model = apps.get_model('administrador', 'Proyectos')
        fields = [field.name for field in Proyectos_model._meta.get_fields() if isinstance(field, (CharField, TextField, DateField, DateTimeField))]
        query_filter = Q()
        for field in fields:
            if isinstance(Proyectos_model._meta.get_field(field), (DateField, DateTimeField)):
                query_filter |= Q(**{f"{field}__year__icontains": query})
                query_filter |= Q(**{f"{field}__month__icontains": query})
                query_filter |= Q(**{f"{field}__day__icontains": query})
            else:
                query_filter |= Q(**{f"{field}__icontains": query})
        proyectos = proyectos.filter(query_filter)

    for proyecto in proyectos:
        row = []
        for header in headers:
            if header == 'ID':
                value = str(proyecto.id)
            elif header == 'Estatus Proyecto':
                value = proyecto.estatus_proyecto or 'N/A'
            elif header == 'PM':
                value = proyecto.pm or 'N/A'
            elif header == 'Empresa':
                value = proyecto.empresa or 'N/A'
            elif header == 'Nombre':
                value = proyecto.nombre or 'N/A'
            elif header == 'Fecha Inicio':
                value = proyecto.fecha_inicio.strftime('%d/%m/%Y') if proyecto.fecha_inicio else 'N/A'
            elif header == 'Fecha Término':
                value = proyecto.fecha_termino.strftime('%d/%m/%Y') if proyecto.fecha_termino else 'N/A'
            elif header == 'Alcance':
                value = proyecto.alcance or 'N/A'
            elif header == 'Usuario':
                value = proyecto.user.username if proyecto.user else 'Sin usuario'
            else:
                value = '-'
            row.append(value)
        ws.append(row)

    wb.save(response)
    return response

# Exportar a PDF
@login_required
def export_to_pdf_proyectos(request):
    query = request.GET.get('q', '')
    headers_str = request.GET.get('headers', '')
    headers = headers_str.split(',') if headers_str else [
        'ID', 'Estatus Proyecto', 'PM', 'Empresa', 'Nombre', 'Fecha Inicio', 'Fecha Término', 'Alcance', 'Usuario'
    ]

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="proyectos_filtrados.pdf"'

    p = canvas.Canvas(response)
    p.drawString(100, 750, "Reporte de Proyectos Filtrados")
    p.drawString(100, 730, " | ".join(headers))

    proyectos = Proyectos.objects.all().order_by('id')
    if query:
        Proyectos_model = apps.get_model('administrador', 'Proyectos')
        fields = [field.name for field in Proyectos_model._meta.get_fields() if isinstance(field, (CharField, TextField, DateField, DateTimeField))]
        query_filter = Q()
        for field in fields:
            if isinstance(Proyectos_model._meta.get_field(field), (DateField, DateTimeField)):
                query_filter |= Q(**{f"{field}__year__icontains": query})
                query_filter |= Q(**{f"{field}__month__icontains": query})
                query_filter |= Q(**{f"{field}__day__icontains": query})
            else:
                query_filter |= Q(**{f"{field}__icontains": query})
        proyectos = proyectos.filter(query_filter)

    y = 710
    for proyecto in proyectos:
        row = []
        for header in headers:
            if header == 'ID':
                value = str(proyecto.id)
            elif header == 'Estatus Proyecto':
                value = proyecto.estatus_proyecto or 'N/A'
            elif header == 'PM':
                value = proyecto.pm or 'N/A'
            elif header == 'Empresa':
                value = proyecto.empresa or 'N/A'
            elif header == 'Nombre':
                value = proyecto.nombre or 'N/A'
            elif header == 'Fecha Inicio':
                value = proyecto.fecha_inicio.strftime('%d/%m/%Y') if proyecto.fecha_inicio else 'N/A'
            elif header == 'Fecha Término':
                value = proyecto.fecha_termino.strftime('%d/%m/%Y') if proyecto.fecha_termino else 'N/A'
            elif header == 'Alcance':
                value = proyecto.alcance or 'N/A'
            elif header == 'Usuario':
                value = proyecto.user.username if proyecto.user else 'Sin usuario'
            else:
                value = '-'
            row.append(value)
        text = " | ".join(row)
        p.drawString(100, y, text[:500])  # Limitar longitud para evitar desbordamiento
        y -= 20
        if y < 50:
            p.showPage()
            y = 750

    p.showPage()
    p.save()
    return response

#### PROSPECCIONES ##############################################


# 1. Agregar Prospecciones

logger = logging.getLogger(__name__)

@login_required
def agregar_prospecciones(request):
    proyectos = Proyectos.objects.all()
    tipo_prospeccion_choices = Prospecciones._meta.get_field('tipo_prospeccion').choices
    tipo_sondaje_choices = Prospecciones._meta.get_field('tipo_sondaje').choices
    metodologia_sondaje_choices = Prospecciones._meta.get_field('metodologia_sondaje').choices
    metodologia_geofisica_choices = Prospecciones._meta.get_field('metodologia_geofisica').choices
    diametro_sondaje_choices = Prospecciones._meta.get_field('diametro_sondaje').choices
    habilitacion_choices = Prospecciones._meta.get_field('habilitacion').choices
    monolito_choices = Prospecciones._meta.get_field('monolito').choices
    tapado_choices = Prospecciones._meta.get_field('tapado').choices

    if request.method == 'POST':
        logger.info(f"Datos POST recibidos: {request.POST}")
        logger.info(f"Archivos recibidos: {request.FILES}")
        proyecto_id = request.POST.get('id_proyecto')
        tipo_prospeccion = request.POST.get('tipo_prospeccion')
        id_prospecciones = request.POST.getlist('id_prospeccion[]')
        areas = request.POST.getlist('area[]')
        coordenadas_este = request.POST.getlist('coordenada_este[]')
        coordenadas_norte = request.POST.getlist('coordenada_norte[]')
        elevaciones = request.POST.getlist('elevacion[]')
        profundidades = request.POST.getlist('profundidad[]')
        inclinaciones = request.POST.getlist('inclinacion[]')
        fechas_inicio = request.POST.getlist('fecha_inicio_perforacion[]')
        fechas_termino = request.POST.getlist('fecha_termino_perforacion[]')
        tipos_sondaje = request.POST.getlist('tipo_sondaje[]')
        metodologias_sondaje = request.POST.getlist('metodologia_sondaje[]')
        metodologias_geofisica = request.POST.getlist('metodologia_geofisica[]')
        diametros_sondaje = request.POST.getlist('diametro_sondaje[]')
        habilitaciones = request.POST.getlist('habilitacion[]')
        monolitos = request.POST.getlist('monolito[]')
        tapados = request.POST.getlist('tapado[]')
        contratistas = request.POST.getlist('contratista[]')
        marcas_maquina1 = request.POST.getlist('marca_maquina1[]')
        modelos_maquina1 = request.POST.getlist('modelo_maquina1[]')
        ppus1 = request.POST.getlist('ppu1[]')
        marcas_maquina2 = request.POST.getlist('marca_maquina2[]')
        modelos_maquina2 = request.POST.getlist('modelo_maquina2[]')
        ppus2 = request.POST.getlist('ppu2[]')
        observaciones = request.POST.getlist('observacion[]')
        images = request.FILES.getlist('image[]')  # Lista de todas las imágenes subidas
        errors = []

        if not proyecto_id:
            errors.append("Debe seleccionar un proyecto.")
        if not tipo_prospeccion:
            errors.append("Debe seleccionar un tipo de prospección.")
        if not Proyectos.objects.filter(id=proyecto_id).exists():
            errors.append("El proyecto seleccionado no existe.")
        if not id_prospecciones:
            errors.append("Debe ingresar al menos un ID de prospección.")

        # Procesar cada fila
        for i in range(len(id_prospecciones)):
            id_pros = id_prospecciones[i]
            if not id_pros:
                errors.append(f"Fila {i+1}: Falta ID de prospección.")
                continue

            try:
                coord_este = None
                if i < len(coordenadas_este) and coordenadas_este[i]:
                    coord_este = float(coordenadas_este[i])

                coord_norte = None
                if i < len(coordenadas_norte) and coordenadas_norte[i]:
                    coord_norte = float(coordenadas_norte[i])

                prospeccion_obj = Prospecciones(
                    id_proyecto_id=proyecto_id,
                    tipo_prospeccion=tipo_prospeccion,
                    id_prospeccion=id_pros,
                    area=areas[i] if i < len(areas) and areas[i] else None,
                    coordenada_este=coord_este,
                    coordenada_norte=coord_norte,
                    elevacion=int(elevaciones[i]) if i < len(elevaciones) and elevaciones[i] else None,
                    profundidad=float(profundidades[i]) if i < len(profundidades) and profundidades[i] else None,
                    inclinacion=float(inclinaciones[i]) if i < len(inclinaciones) and inclinaciones[i] and tipo_prospeccion == 'sondajes' else None,
                    fecha_inicio_perforacion=fechas_inicio[i] if i < len(fechas_inicio) and fechas_inicio[i] else None,
                    fecha_termino_perforacion=fechas_termino[i] if i < len(fechas_termino) and fechas_termino[i] else None,
                    tipo_sondaje=tipos_sondaje[i] if i < len(tipos_sondaje) and tipos_sondaje[i] and tipo_prospeccion == 'sondajes' else None,
                    metodologia_sondaje=metodologias_sondaje[i] if i < len(metodologias_sondaje) and metodologias_sondaje[i] and tipo_prospeccion == 'sondajes' else None,
                    metodologia_geofisica=metodologias_geofisica[i] if i < len(metodologias_geofisica) and metodologias_geofisica[i] and tipo_prospeccion == 'geofisica' else None,
                    diametro_sondaje=diametros_sondaje[i] if i < len(diametros_sondaje) and diametros_sondaje[i] and tipo_prospeccion == 'sondajes' else None,
                    habilitacion=habilitaciones[i] if i < len(habilitaciones) and habilitaciones[i] and tipo_prospeccion == 'sondajes' else None,
                    monolito=monolitos[i] if i < len(monolitos) and monolitos[i] and tipo_prospeccion == 'sondajes' else None,
                    tapado=tapados[i] if i < len(tapados) and tapados[i] and (tipo_prospeccion in ['sondajes', 'calicatas']) else None,
                    contratista=contratistas[i] if i < len(contratistas) and contratistas[i] else None,
                    marca_maquina1=marcas_maquina1[i] if i < len(marcas_maquina1) and marcas_maquina1[i] else None,
                    modelo_maquina1=modelos_maquina1[i] if i < len(modelos_maquina1) and modelos_maquina1[i] else None,
                    ppu1=ppus1[i] if i < len(ppus1) and ppus1[i] else None,
                    marca_maquina2=marcas_maquina2[i] if i < len(marcas_maquina2) and marcas_maquina2[i] else None,
                    modelo_maquina2=modelos_maquina2[i] if i < len(modelos_maquina2) and modelos_maquina2[i] else None,
                    ppu2=ppus2[i] if i < len(ppus2) and ppus2[i] else None,
                    observacion=observaciones[i] if i < len(observaciones) and observaciones[i] else None,
                    user=request.user
                )
                prospeccion_obj.save()
                logger.info(f"Prospección guardada: ID={prospeccion_obj.id_prospeccion}")

                # Asociar imágenes utilizando ManyToManyField
                for img in images:
                    imagen_obj = ProspeccionImage.objects.create(prospeccion=prospeccion_obj, image=img)
                    prospeccion_obj.imagenes.add(imagen_obj)
                logger.info(f"Guardadas {len(images)} imágenes para {prospeccion_obj.id_prospeccion}")

            except ValueError as e:
                errors.append(f"Fila {i+1}: Valor inválido: {e}")
            except Exception as e:
                errors.append(f"Fila {i+1}: Error al guardar: {e}")

        if errors:
            return render(request, 'terreno/prospecciones/agregar_prospecciones.html', {
                'proyectos': proyectos,
                'tipo_prospeccion_choices': tipo_prospeccion_choices,
                'tipo_sondaje_choices': tipo_sondaje_choices,
                'metodologia_sondaje_choices': metodologia_sondaje_choices,
                'metodologia_geofisica_choices': metodologia_geofisica_choices,
                'diametro_sondaje_choices': diametro_sondaje_choices,
                'habilitacion_choices': habilitacion_choices,
                'monolito_choices': monolito_choices,
                'tapado_choices': tapado_choices,
                'errors': errors
            })

        return redirect('listar_prospecciones')

    return render(request, 'terreno/prospecciones/agregar_prospecciones.html', {
        'proyectos': proyectos,
        'tipo_prospeccion_choices': tipo_prospeccion_choices,
        'tipo_sondaje_choices': tipo_sondaje_choices,
        'metodologia_sondaje_choices': metodologia_sondaje_choices,
        'metodologia_geofisica_choices': metodologia_geofisica_choices,
        'diametro_sondaje_choices': diametro_sondaje_choices,
        'habilitacion_choices': habilitacion_choices,
        'monolito_choices': monolito_choices,
        'tapado_choices': tapado_choices,
    })
    
#Agregar y eliminar las imagenes del html editar

logger = logging.getLogger(__name__)

@login_required
def agregar_imagen(request):
    prospeccion_id = request.POST.get('prospeccion_id')
    try:
        prospeccion = Prospecciones.objects.get(id=prospeccion_id)
        if 'image' not in request.FILES:
            return JsonResponse({'success': False, 'error': 'No se proporcionó ninguna imagen'})

        imagen = ProspeccionImage(image=request.FILES['image'])
        imagen.prospeccion = prospeccion  # Establecer la relación con la prospección
        imagen.save()  # Ahora guardar la imagen con la relación establecida
        prospeccion.imagenes.add(imagen)  # Añadir la imagen a la relación ManyToMany

        Historial.objects.create(
            usuario=request.user,
            accion=f'Imágenes Asociadas -> AGREGADA 1 imagen a la prospección {prospeccion.id_prospeccion}',
            fecha=timezone.now()
        )
        logger.info(f"Usuario {request.user} agregó una imagen a la prospección {prospeccion.id_prospeccion}")

        return JsonResponse({
            'success': True,
            'image_id': imagen.id,
            'image_url': imagen.image.url
        })
    except Prospecciones.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'La prospección no existe'})
    except Exception as e:
        logger.error(f"Error al agregar imagen: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def eliminar_imagen(request):
    if request.method == 'POST':
        image_id = request.POST.get('image_id')
        image = get_object_or_404(ProspeccionImage, id=image_id)
        prospeccion = image.prospeccion
        prospeccion.imagenes.remove(image)
        image.delete()
        
        # Registrar en el historial
        Historial.objects.create(
            usuario=request.user,
            accion=f'Imágenes Asociadas -> ELIMINADA 1 imagen de la prospección {prospeccion.id_prospeccion}',
            fecha=timezone.now()
        )
        logger.info(f"Usuario {request.user} eliminó una imagen de la prospección {prospeccion.id_prospeccion}")
        
        return JsonResponse({'success': True})
    return JsonResponse({'success': False})
    
    
# 2. Listar Prospecciones
@login_required
def listar_prospecciones(request):
    prospecciones = Prospecciones.objects.all().order_by('id')
    query = request.GET.get('q', '')

    if query:
        prospeccion_model = apps.get_model('administrador', 'Prospecciones')
        campos = [field.name for field in prospeccion_model._meta.get_fields() if isinstance(field, (models.CharField, models.TextField, models.DateField, models.DecimalField, models.ForeignKey, models.IntegerField))]
        query_filter = Q()
        try:
            query_numeric = float(query)
            numeric_search = True
        except ValueError:
            numeric_search = False

        for field in campos:
            field_instance = prospeccion_model._meta.get_field(field)
            if isinstance(field_instance, (models.DateField, models.DateTimeField)):
                query_filter |= Q(**{f"{field}__year__icontains": query})
                query_filter |= Q(**{f"{field}__month__icontains": query})
                query_filter |= Q(**{f"{field}__day__icontains": query})
            elif isinstance(field_instance, models.ForeignKey):
                related_model = field_instance.related_model
                related_fields = [f"{field}__{related_field.name}" for related_field in related_model._meta.get_fields() if isinstance(related_field, (models.CharField, models.TextField))]
                for related_field in related_fields:
                    query_filter |= Q(**{f"{related_field}__icontains": query})
            elif isinstance(field_instance, (models.DecimalField, models.IntegerField)) and numeric_search:
                query_filter |= Q(**{f"{field}": query_numeric})
            else:
                query_filter |= Q(**{f"{field}__icontains": query})

        prospecciones = prospecciones.filter(query_filter)

    paginator = Paginator(prospecciones, 1000)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'terreno/prospecciones/prospecciones_table.html', {'page_obj': page_obj})

    return render(request, 'terreno/prospecciones/listar_prospecciones.html', {
        'page_obj': page_obj,
        'query': query
    })

# 3. Editar Prospecciones
@login_required
def editar_prospecciones(request, id):
    prospeccion_obj = get_object_or_404(Prospecciones, id=id)
    proyectos = Proyectos.objects.all()

    if request.method == 'POST':
        form = ProspeccionesForm(request.POST, instance=prospeccion_obj)
        if form.is_valid():
            try:
                prospeccion_obj = form.save(commit=False)
                prospeccion_obj.user = request.user
                prospeccion_obj.save()
                logger.info(f"Prospección editada: ID {prospeccion_obj.id_prospeccion}")
                return redirect('listar_prospecciones')
            except Exception as e:
                logger.error(f"Error al editar: {e}")
                form.add_error(None, f"Error al editar: {e}")
    else:
        form = ProspeccionesForm(instance=prospeccion_obj)

    return render(request, 'terreno/prospecciones/editar_prospecciones.html', {
        'form': form,
        'prospeccion_obj': prospeccion_obj,
        'proyectos': proyectos
    })
    
# 4. Ver Prospecciones

# Configurar el registro
logger = logging.getLogger(__name__)

@login_required
def ver_prospecciones(request, id):
    prospeccion_obj = get_object_or_404(Prospecciones, id=id)

    # Historial de cambios
    history_records = []
    for record in prospeccion_obj.history.all():
        changes = []
        previous = record.prev_record
        if previous:
            for field in record.history_object._meta.fields:
                old_value = getattr(previous, field.name, None)
                new_value = getattr(record, field.name, None)
                if old_value != new_value:
                    changes.append({
                        'field_name': field.verbose_name,
                        'old_value': old_value if old_value is not None else '-',
                        'new_value': new_value if new_value is not None else '-',
                    })
        history_records.append({
            'date': record.history_date,
            'user': record.history_user,
            'type': record.history_type,
            'reason': record.history_change_reason,
            'changes': changes if changes else None,
        })

    # Conversión de coordenadas UTM a latitud/longitud
    lat, lon = None, None
    if prospeccion_obj.coordenada_este and prospeccion_obj.coordenada_norte:
        try:
            # Usar zona_utm si existe, sino default a 19S (Calama)
            zona = getattr(prospeccion_obj, 'zona_utm', "19S")
            zone_number = int(zona[:-1])  # Extrae el número (19)
            south = zona.endswith('S')    # Determina hemisferio
            utm_proj = Proj(proj='utm', zone=zone_number, datum='WGS84', south=south)
            wgs84_proj = Proj(proj='latlong', datum='WGS84')

            este = float(prospeccion_obj.coordenada_este)  # Convertir a float
            norte = float(prospeccion_obj.coordenada_norte)  # Convertir a float

            # Validación de rangos estrictos para UTM, ajustados para zona 19S
            if (200000 <= este <= 800000) and (7000000 <= norte <= 8000000):
                lon, lat = transform(utm_proj, wgs84_proj, este, norte)
            else:
                logger.warning(f"Coordenadas fuera de rango para {prospeccion_obj.id_prospeccion}: Este={este}, Norte={norte}, Zona={zona}")
        except ValueError as ve:
            logger.error(f"Error de valor al convertir coordenadas para {prospeccion_obj.id_prospeccion}: {ve}")
        except Exception as e:
            logger.error(f"Error inesperado al convertir coordenadas para {prospeccion_obj.id_prospeccion}: {e}")

    context = {
        'prospeccion': prospeccion_obj,
        'history_records': history_records,
        'lat': lat,
        'lon': lon,
    }
    return render(request, 'terreno/prospecciones/ver_prospecciones.html', context)


# 5. Eliminar Prospecciones
@login_required
def eliminar_prospecciones(request, id):
    prospeccion_obj = get_object_or_404(Prospecciones, id=id)
    if request.method == 'POST':
        prospeccion_obj.delete()
        return redirect('listar_prospecciones')
    return render(request, 'terreno/prospecciones/eliminar_prospecciones.html', {
        'prospeccion': prospeccion_obj
    })

# 6. Exportar a Excel
@login_required
def export_to_excel_prospecciones(request):
    query = request.GET.get('q', '').strip()
    headers = [
        'ID Prospección', 'ID Proyecto', 'Tipo Prospección', 'Tipo Sondaje', 'Metodología Sondaje',
        'Metodología Geofísica', 'Área', 'Fecha Inicio Perforación', 'Fecha Término Perforación',
        'Coord. Este', 'Coord. Norte', 'Elevación', 'Profundidad', 'Inclinación', 'Diámetro Sondaje',
        'Habilitación', 'Monolito', 'Tapado', 'Contratista', 'Marca Máquina 1', 'Modelo Máquina 1',
        'PPU 1', 'Marca Máquina 2', 'Modelo Máquina 2', 'PPU 2', 'Observación', 'Usuario'
    ]

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="prospecciones_filtrado.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Prospecciones"
    ws.append(headers)

    try:
        prospecciones_list = Prospecciones.objects.select_related('id_proyecto', 'user').all()

        if query:
            filtros = Q()
            try:
                query_numeric = float(query)
                filtros |= (
                    Q(coordenada_este=query_numeric) |
                    Q(coordenada_norte=query_numeric) |
                    Q(elevacion=query_numeric) |
                    Q(profundidad=query_numeric) |
                    Q(inclinacion=query_numeric)
                )
            except ValueError:
                pass

            filtros |= (
                Q(id_prospeccion__icontains=query) |
                Q(id_proyecto__id__icontains=query) |
                Q(tipo_prospeccion__icontains=query) |
                Q(tipo_sondaje__icontains=query) |
                Q(metodologia_sondaje__icontains=query) |
                Q(metodologia_geofisica__icontains=query) |
                Q(area__icontains=query) |
                Q(contratista__icontains=query) |
                Q(observacion__icontains=query) |
                Q(user__username__icontains=query)
            )
            prospecciones_list = prospecciones_list.filter(filtros)

        for prospeccion in prospecciones_list:
            row = [
                str(prospeccion.id_prospeccion),
                str(prospeccion.id_proyecto.id) if prospeccion.id_proyecto else '',
                str(prospeccion.tipo_prospeccion or ''),
                str(prospeccion.tipo_sondaje or ''),
                str(prospeccion.metodologia_sondaje or ''),
                str(prospeccion.metodologia_geofisica or ''),
                str(prospeccion.area or ''),
                str(prospeccion.fecha_inicio_perforacion or ''),
                str(prospeccion.fecha_termino_perforacion or ''),
                str(prospeccion.coordenada_este) if prospeccion.coordenada_este is not None else '',
                str(prospeccion.coordenada_norte) if prospeccion.coordenada_norte is not None else '',
                str(prospeccion.elevacion) if prospeccion.elevacion is not None else '',
                str(prospeccion.profundidad) if prospeccion.profundidad is not None else '',
                str(prospeccion.inclinacion) if prospeccion.inclinacion is not None else '',
                str(prospeccion.diametro_sondaje or ''),
                str(prospeccion.habilitacion or ''),
                str(prospeccion.monolito or ''),
                str(prospeccion.tapado or ''),
                str(prospeccion.contratista or ''),
                str(prospeccion.marca_maquina1 or ''),
                str(prospeccion.modelo_maquina1 or ''),
                str(prospeccion.ppu1 or ''),
                str(prospeccion.marca_maquina2 or ''),
                str(prospeccion.modelo_maquina2 or ''),
                str(prospeccion.ppu2 or ''),
                str(prospeccion.observacion or ''),
                str(prospeccion.user.username) if prospeccion.user else 'Sin usuario'
            ]
            ws.append(row)

        wb.save(response)
        return response

    except Exception as e:
        logger.error(f"Error en export_to_excel_prospecciones: {str(e)}")
        ws.append(["Error al generar el reporte", str(e)])
        wb.save(response)
        return response

# 7. Exportar a PDF
@login_required
def export_to_pdf_prospecciones(request):
    query = request.GET.get('q', '').strip()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="prospecciones_filtrado.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(letter), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()

    title = Paragraph("Lista de Prospecciones", styles['Heading1'].clone('Title', alignment=1, fontSize=12, spaceAfter=10))
    elements.append(title)
    elements.append(Paragraph("<br/>", styles['Normal']))

    headers = [
        'ID Prospección', 'Proyecto', 'Tipo', 'Tipo Sondaje', 'Met. Sondaje',
        'Met. Geofísica', 'Área', 'F. Inicio', 'F. Término', 'Coord. E',
        'Coord. N', 'Elev.', 'Prof.', 'Incl.', 'Diám. Sondaje', 'Habil.',
        'Monolito', 'Tapado', 'Contratista', 'Marca M1', 'Modelo M1', 'PPU 1',
        'Marca M2', 'Modelo M2', 'PPU 2', 'Observación', 'Usuario'
    ]
    data = [headers]

    try:
        prospecciones_list = Prospecciones.objects.select_related('id_proyecto', 'user').all()

        if query:
            filtros = Q()
            try:
                query_numeric = float(query)
                filtros |= (
                    Q(coordenada_este=query_numeric) |
                    Q(coordenada_norte=query_numeric) |
                    Q(elevacion=query_numeric) |
                    Q(profundidad=query_numeric) |
                    Q(inclinacion=query_numeric)
                )
            except ValueError:
                pass

            filtros |= (
                Q(id_prospeccion__icontains=query) |
                Q(id_proyecto__id__icontains=query) |
                Q(tipo_prospeccion__icontains=query) |
                Q(tipo_sondaje__icontains=query) |
                Q(metodologia_sondaje__icontains=query) |
                Q(metodologia_geofisica__icontains=query) |
                Q(area__icontains=query) |
                Q(contratista__icontains=query) |
                Q(observacion__icontains=query) |
                Q(user__username__icontains=query)
            )
            prospecciones_list = prospecciones_list.filter(filtros)

        for prospeccion in prospecciones_list:
            row = [
                str(prospeccion.id_prospeccion),
                str(prospeccion.id_proyecto.id) if prospeccion.id_proyecto else '',
                str(prospeccion.tipo_prospeccion or ''),
                str(prospeccion.tipo_sondaje or ''),
                str(prospeccion.metodologia_sondaje or ''),
                str(prospeccion.metodologia_geofisica or ''),
                str(prospeccion.area or ''),
                str(prospeccion.fecha_inicio_perforacion or ''),
                str(prospeccion.fecha_termino_perforacion or ''),
                str(prospeccion.coordenada_este) if prospeccion.coordenada_este is not None else '',
                str(prospeccion.coordenada_norte) if prospeccion.coordenada_norte is not None else '',
                str(prospeccion.elevacion) if prospeccion.elevacion is not None else '',
                str(prospeccion.profundidad) if prospeccion.profundidad is not None else '',
                str(prospeccion.inclinacion) if prospeccion.inclinacion is not None else '',
                str(prospeccion.diametro_sondaje or ''),
                str(prospeccion.habilitacion or ''),
                str(prospeccion.monolito or ''),
                str(prospeccion.tapado or ''),
                str(prospeccion.contratista or ''),
                str(prospeccion.marca_maquina1 or ''),
                str(prospeccion.modelo_maquina1 or ''),
                str(prospeccion.ppu1 or ''),
                str(prospeccion.marca_maquina2 or ''),
                str(prospeccion.modelo_maquina2 or ''),
                str(prospeccion.ppu2 or ''),
                str(prospeccion.observacion or ''),
                str(prospeccion.user.username) if prospeccion.user else 'Sin usuario'
            ]
            data.append(row)

        table = Table(data, colWidths=[0.6*inch]*27)  # Ajusta el ancho de las columnas según necesites
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(table)
        doc.build(elements)
        return response

    except Exception as e:
        logger.error(f"Error en export_to_pdf_prospecciones: {str(e)}")
        doc.build([Paragraph(f"Error al generar el reporte: {str(e)}", styles['Normal'])])
        return response

# 8. Gráficos Prospecciones
def generar_grafico_profundidad_area(df):
    areas_unicas = df['area'].unique()
    colores = {area: (random.random(), random.random(), random.random()) for area in areas_unicas}
    fig, ax = plt.subplots()

    for area in areas_unicas:
        datos_area = df[df['area'] == area]
        ax.scatter(
            datos_area['coordenada_este'],
            datos_area['profundidad'],
            color=colores[area],
            s=100,
            label=area,
            marker='+'
        )

    ax.xaxis.set_ticks_position("top")
    ax.xaxis.set_label_position("top")
    ax.invert_yaxis()
    ax.set_xlabel("Coordenada Este")
    ax.set_ylabel("Profundidad (m)")
    ax.grid(True)
    plt.subplots_adjust(right=0.75)
    num_columns = (len(areas_unicas) + 24) // 25
    legend = ax.legend(loc='center left', bbox_to_anchor=(1, 0.5), ncol=num_columns)
    plt.title("GRÁFICO AGRUPADO POR ÁREA (DISPERSIÓN)", pad=30)

    buffer = BytesIO()
    fig.savefig(buffer, format='png', dpi=100, bbox_inches='tight', bbox_extra_artists=[legend])
    buffer.seek(0)
    plt.close()
    return base64.b64encode(buffer.getvalue()).decode('utf-8')

def generar_grafico_profundidad_prospeccion(df):
    df['id_prospeccion'] = df['id_prospeccion'].astype(str)
    prospecciones_unicas = df['id_prospeccion'].unique()
    colores = {prospeccion: (random.random(), random.random(), random.random()) for prospeccion in prospecciones_unicas}

    fig, ax = plt.subplots()

    for prospeccion in prospecciones_unicas:
        datos_prospeccion = df[df['id_prospeccion'] == prospeccion]
        ax.scatter(
            datos_prospeccion['coordenada_este'],
            datos_prospeccion['profundidad'],
            color=colores[prospeccion],
            s=100,
            label=prospeccion,
            marker='+'
        )

    ax.xaxis.set_ticks_position("top")
    ax.xaxis.set_label_position("top")
    ax.invert_yaxis()
    ax.set_xlabel("Coordenada Este")
    ax.set_ylabel("Profundidad (m)")
    ax.grid(True)
    plt.subplots_adjust(right=0.75)
    num_columns = (len(prospecciones_unicas) + 24) // 25
    legend = ax.legend(loc='center left', bbox_to_anchor=(1, 0.5), ncol=num_columns)
    plt.title("GRÁFICO AGRUPADO POR PROSPECCIÓN (DISPERSIÓN)", pad=30)

    buffer = BytesIO()
    fig.savefig(buffer, format='png', dpi=100, bbox_inches='tight', bbox_extra_artists=[legend])
    buffer.seek(0)
    plt.close()
    return base64.b64encode(buffer.getvalue()).decode('utf-8')

@login_required
def graficos_prospecciones(request):
    id_proyectos = request.GET.getlist('id_proyectos')
    tipos_prospeccion = request.GET.getlist('tipos_prospeccion')
    areas = request.GET.getlist('areas')
    id_prospecciones = request.GET.getlist('id_prospecciones')

    query = Prospecciones.objects.all()
    if id_proyectos:
        query = query.filter(id_proyecto__in=id_proyectos)
    if tipos_prospeccion:
        query = query.filter(tipo_prospeccion__in=tipos_prospeccion)
    if areas:
        query = query.filter(area__in=areas)
    if id_prospecciones:
        query = query.filter(id_prospeccion__in=id_prospecciones)

    proyectos = query.values('id_proyecto').distinct()
    tipos_prospeccion_inicial = query.values_list('tipo_prospeccion', flat=True).distinct().exclude(tipo_prospeccion__isnull=True)
    areas_inicial = query.values_list('area', flat=True).distinct().exclude(area__isnull=True).exclude(area='')
    id_prospecciones_inicial = query.values_list('id_prospeccion', flat=True).distinct()

    df = pd.DataFrame(list(query.values('coordenada_este', 'profundidad', 'area', 'id_prospeccion')))

    context = {
        'proyectos': proyectos,
        'tipos_prospeccion_inicial': tipos_prospeccion_inicial,
        'areas_inicial': areas_inicial,
        'id_prospecciones_inicial': id_prospecciones_inicial,
        'selected_id_proyectos': json.dumps(id_proyectos),
        'selected_tipos_prospeccion': json.dumps(tipos_prospeccion),
        'selected_areas': json.dumps(areas),
        'selected_id_prospecciones': json.dumps(id_prospecciones),
    }

    if df.empty:
        context['error'] = "No hay datos disponibles para los filtros seleccionados."
    else:
        context['image_base64_area'] = generar_grafico_profundidad_area(df)
        context['image_base64_prospeccion'] = generar_grafico_profundidad_prospeccion(df)

    return render(request, 'terreno/prospecciones/graficos_prospecciones.html', context)

#### FUNCIONES COMUNES PARA MÓDULOS DE LABORATORIO ##############################################
logger = logging.getLogger(__name__)

# Funciones comunes
@login_required
def obtener_proyectos_prospecciones(request, as_json=False):
    """
    Obtiene una lista de proyectos únicos asociados a Prospecciones (ajustado para usar Prospecciones en lugar de Granulometria).
    """
    proyectos = Prospecciones.objects.values('id_proyecto').distinct()
    proyectos_list = [proj['id_proyecto'] for proj in proyectos if proj['id_proyecto'] is not None]
    proyectos_dict = {proj: proj for proj in proyectos_list}
    if as_json:
        return JsonResponse(proyectos_dict, safe=False)
    return proyectos_list

@login_required
def obtener_tipos_prospeccion(request):
    """
    Obtiene tipos de prospección únicos basados en id_proyectos (múltiples) o id_proyecto (único).
    """
    id_proyectos = request.GET.get('id_proyectos')  # Para múltiples proyectos (separados por coma)
    id_proyecto = request.GET.get('id_proyecto')    # Para un solo proyecto
    query = Prospecciones.objects.all()

    if id_proyectos:
        query = query.filter(id_proyecto__in=id_proyectos.split(','))
    elif id_proyecto:
        query = query.filter(id_proyecto=id_proyecto)

    tipos_list = list(query.values_list('tipo_prospeccion', flat=True).distinct().exclude(tipo_prospeccion__isnull=True))
    return JsonResponse({'options': tipos_list}, safe=False)

@login_required
def obtener_id_prospecciones(request):
    """
    Obtiene IDs de prospecciones basados en id_proyectos, tipos_prospeccion y áreas.
    """
    id_proyectos = request.GET.get('id_proyectos')  # Para múltiples proyectos
    id_proyecto = request.GET.get('id_proyecto')    # Para un solo proyecto
    tipos_prospeccion = request.GET.get('tipos_prospeccion')
    tipo_prospeccion = request.GET.get('tipo_prospeccion')
    areas = request.GET.get('areas')
    query = Prospecciones.objects.all()

    if id_proyectos:
        query = query.filter(id_proyecto__in=id_proyectos.split(','))
    elif id_proyecto:
        query = query.filter(id_proyecto=id_proyecto)

    if tipos_prospeccion:
        query = query.filter(tipo_prospeccion__in=tipos_prospeccion.split(','))
    elif tipo_prospeccion:
        query = query.filter(tipo_prospeccion=tipo_prospeccion)

    if areas:
        query = query.filter(area__in=areas.split(','))

    id_prospecciones_list = list(query.values_list('id_prospeccion', flat=True).distinct())
    return JsonResponse({'options': id_prospecciones_list}, safe=False)

@login_required
def obtener_area(request):
    """
    Obtiene áreas únicas basadas en id_prospecciones (múltiples) o prospeccion_id (único).
    """
    id_prospecciones = request.GET.get('id_prospecciones')  # Para múltiples prospecciones
    prospeccion_id = request.GET.get('prospeccion_id')      # Para una sola prospección

    if id_prospecciones:
        query = Prospecciones.objects.filter(id_prospeccion__in=id_prospecciones.split(','))
        areas_list = list(query.values_list('area', flat=True).distinct().exclude(area__isnull=True).exclude(area=''))
        return JsonResponse({'options': areas_list})
    elif prospeccion_id:
        try:
            prospeccion = Prospecciones.objects.get(id_prospeccion=prospeccion_id)
            area = prospeccion.area or ''
            return JsonResponse({'area': area})
        except Prospecciones.DoesNotExist:
            return JsonResponse({'area': ''})

    return JsonResponse({'options': []})

####### GRAVEDAD ESPECIFICA #######################

#1 Agregar
logger = logging.getLogger(__name__)

@login_required
def agregar_gravedad_especifica(request):
    proyectos_con_muestreo = Proyectos.objects.filter(muestreo__isnull=False).distinct()
    tipo_prospeccion_choices = []
    prospecciones = []
    muestras = []
    area = ""

    if request.method == 'POST':
        logger.info(f"Datos POST recibidos: {request.POST}")
        proyecto_id = request.POST.get('id_proyecto')
        tipo_prospeccion = request.POST.get('tipo_prospeccion')
        id_prospecciones = request.POST.getlist('id_prospeccion[]')
        id_muestras = request.POST.getlist('id_muestra[]')
        gravedades = request.POST.getlist('gravedad_especifica[]')
        areas = request.POST.getlist('area[]')
        errors = []

        if not proyecto_id:
            errors.append("Debe seleccionar un proyecto.")
            return render(request, 'laboratorio/ensayos/gravedad_especifica/agregar_gravedad_especifica.html', {
                'proyectos': proyectos_con_muestreo,
                'errors': errors
            })

        muestreos = Muestreo.objects.filter(id_proyecto_id=proyecto_id)
        if not muestreos.exists():
            errors.append("No hay muestreos asociados a este proyecto.")
            return render(request, 'laboratorio/ensayos/gravedad_especifica/agregar_gravedad_especifica.html', {
                'proyectos': proyectos_con_muestreo,
                'errors': errors
            })

        if tipo_prospeccion:
            muestreos = muestreos.filter(tipo_prospeccion=tipo_prospeccion)
            if not muestreos.exists():
                errors.append("No hay muestreos con este tipo de prospección.")

        if id_prospecciones and id_prospecciones[0]:
            muestreos = muestreos.filter(id_prospeccion__id_prospeccion__in=id_prospecciones)
            if not muestreos.exists():
                errors.append("No hay muestreos con las prospecciones seleccionadas.")

        if id_muestras and id_muestras[0]:
            muestreos = muestreos.filter(id_muestra__in=id_muestras)
            if not muestreos.exists():
                errors.append("No hay muestreos con las muestras seleccionadas.")

        if not (len(id_prospecciones) == len(id_muestras) == len(gravedades) == len(areas)):
            errors.append("El número de prospecciones, muestras, gravedades específicas y áreas no coincide.")
            logger.error(f"Longitudes inconsistentes: prospecciones={len(id_prospecciones)}, muestras={len(id_muestras)}, gravedades={len(gravedades)}, areas={len(areas)}")
        else:
            for id_pros, id_muestra, gravedad, area_val in zip(id_prospecciones, id_muestras, gravedades, areas):
                if not id_pros or not id_muestra or not gravedad or not area_val:
                    errors.append(f"Falta ID de prospección ({id_pros}), ID de muestra ({id_muestra}), gravedad específica ({gravedad}) o área ({area_val}).")
                    continue
                
                try:
                    muestreo = muestreos.get(id_prospeccion__id_prospeccion=id_pros, id_muestra=id_muestra)
                    profundidad_desde = float(muestreo.profundidad_desde or 0)  # Si es None, usa 0
                    profundidad_hasta = float(muestreo.profundidad_hasta or 0)  # Si es None, usa 0
                    profundidad_promedio = (profundidad_desde + profundidad_hasta) / 2

                    gravedad_obj = gravedad_especifica(
                        id_proyecto=muestreo.id_proyecto,
                        tipo_prospeccion=muestreo.tipo_prospeccion,
                        id_prospeccion=muestreo.id_prospeccion,
                        id_muestra=muestreo.id_muestra,
                        profundidad_desde=muestreo.profundidad_desde,
                        profundidad_hasta=muestreo.profundidad_hasta,
                        profundidad_promedio=profundidad_promedio,
                        gravedad_especifica=float(gravedad),
                        area=area_val,
                        user=request.user
                    )
                    gravedad_obj.save()
                    logger.info(f"Gravedad específica guardada: ID={gravedad_obj.id}, Muestra={id_muestra}, Gravedad={gravedad}, Profundidad Promedio={profundidad_promedio}")
                except Muestreo.DoesNotExist:
                    errors.append(f"Muestra {id_muestra} con prospección {id_pros} no encontrada.")
                except ValueError as e:
                    errors.append(f"Valor inválido para gravedad específica {gravedad}: {e}")
                except Exception as e:
                    errors.append(f"Error al guardar muestra {id_muestra}: {e}")

        if errors:
            tipo_prospeccion_choices = Muestreo.objects.filter(id_proyecto_id=proyecto_id).values_list('tipo_prospeccion', flat=True).distinct()
            prospecciones = Muestreo.objects.filter(id_proyecto_id=proyecto_id).values_list('id_prospeccion__id_prospeccion', flat=True).distinct()
            muestras = Muestreo.objects.filter(id_proyecto_id=proyecto_id).values_list('id_muestra', flat=True).distinct()
            return render(request, 'laboratorio/ensayos/gravedad_especifica/agregar_gravedad_especifica.html', {
                'proyectos': proyectos_con_muestreo,
                'tipo_prospeccion_choices': tipo_prospeccion_choices,
                'prospecciones': prospecciones,
                'muestras': muestras,
                'area': area,
                'errors': errors
            })

        return redirect('listar_gravedad_especifica')

    proyecto_id = request.GET.get('id_proyecto')
    tipo_prospeccion = request.GET.get('tipo_prospeccion')
    id_prospeccion = request.GET.get('id_prospeccion')
    id_muestra = request.GET.get('id_muestra')

    muestreos = Muestreo.objects.all()
    if proyecto_id:
        muestreos = muestreos.filter(id_proyecto_id=proyecto_id)
        tipo_prospeccion_choices = muestreos.values_list('tipo_prospeccion', flat=True).distinct()
        prospecciones = muestreos.values_list('id_prospeccion__id_prospeccion', flat=True).distinct()
        muestras = muestreos.values_list('id_muestra', flat=True).distinct()
        area = muestreos.first().area if muestreos.exists() else ""

    if tipo_prospeccion:
        muestreos = muestreos.filter(tipo_prospeccion=tipo_prospeccion)
        prospecciones = muestreos.values_list('id_prospeccion__id_prospeccion', flat=True).distinct()
        muestras = muestreos.values_list('id_muestra', flat=True).distinct()

    if id_prospeccion:
        muestreos = muestreos.filter(id_prospeccion__id_prospeccion=id_prospeccion)
        muestras = muestreos.values_list('id_muestra', flat=True).distinct()
        area = muestreos.first().area if muestreos.exists() else ""

    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if is_ajax:
        data = {
            'tipo_prospeccion_choices': list(tipo_prospeccion_choices),
            'prospecciones': list(prospecciones),
            'muestras': list(muestras),
            'area': area
        }
        if id_muestra:
            muestreo = muestreos.filter(id_muestra=id_muestra).first()
            if muestreo:
                data['profundidad_desde'] = str(muestreo.profundidad_desde) if muestreo.profundidad_desde is not None else '0'
                data['profundidad_hasta'] = str(muestreo.profundidad_hasta) if muestreo.profundidad_hasta is not None else '0'
        return JsonResponse(data)

    return render(request, 'laboratorio/ensayos/gravedad_especifica/agregar_gravedad_especifica.html', {
        'proyectos': proyectos_con_muestreo,
        'tipo_prospeccion_choices': tipo_prospeccion_choices,
        'prospecciones': prospecciones,
        'muestras': muestras,
        'area': area
    })
    
#2 Listar
@login_required
def listar_gravedad_especifica(request):
    gravedad_list = gravedad_especifica.objects.all().order_by('id')  # Ordenar por 'id'
    query = request.GET.get('q', '')

    if query:
        gravedad_model = apps.get_model('administrador', 'gravedad_especifica') 
        # Incluir campos de texto y numéricos
        campos = [field.name for field in gravedad_model._meta.get_fields() if isinstance(field, (CharField, TextField, DateField, DateTimeField, DecimalField, FloatField, ForeignKey))]
        query_filter = Q()

        # Intentar convertir el query a float para búsquedas numéricas
        try:
            query_numeric = float(query)
            numeric_search = True
        except ValueError:
            numeric_search = False

        for field in campos:
            field_instance = gravedad_model._meta.get_field(field)
            if isinstance(field_instance, (DateField, DateTimeField)):
                query_filter |= Q(**{f"{field}__year__icontains": query})
                query_filter |= Q(**{f"{field}__month__icontains": query})
                query_filter |= Q(**{f"{field}__day__icontains": query})
            elif isinstance(field_instance, ForeignKey):
                related_model = field_instance.related_model
                related_fields = [f"{field}__{related_field.name}" for related_field in related_model._meta.get_fields() if isinstance(related_field, (CharField, TextField))]
                for related_field in related_fields:
                    query_filter |= Q(**{f"{related_field}__icontains": query})
            elif isinstance(field_instance, (DecimalField, FloatField)) and numeric_search:
                # Para campos numéricos, usar coincidencia exacta
                query_filter |= Q(**{f"{field}": query_numeric})
            else:
                # Para campos de texto
                query_filter |= Q(**{f"{field}__icontains": query})

        gravedad_list = gravedad_list.filter(query_filter)

    paginator = Paginator(gravedad_list, 1000)  # 1000 registros por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Si es una solicitud AJAX (opcional)
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'laboratorio/ensayos/gravedad_especifica/gravedad_table.html', {'page_obj': page_obj})

    return render(request, 'laboratorio/ensayos/gravedad_especifica/listar_gravedad_especifica.html', {
        'page_obj': page_obj,
        'query': query
    })


#3 Editar
@login_required
def editar_gravedad_especifica(request, id):
    gravedad_obj = get_object_or_404(gravedad_especifica, id=id)
    muestreo = Muestreo.objects.filter(id_muestra=gravedad_obj.id_muestra).first()
    proyectos = Proyectos.objects.filter(muestreo__isnull=False).distinct()

    if request.method == 'POST':
        form = GravedadEspecificaForm(request.POST, instance=gravedad_obj)
        if form.is_valid():
            try:
                gravedad_obj = form.save()
                logger.info(f"Gravedad específica editada: ID {gravedad_obj.id}")
                return redirect('listar_gravedad_especifica')
            except Exception as e:
                logger.error(f"Error al editar: {e}")
                form.add_error(None, f"Error al editar: {e}")
    else:
        form = GravedadEspecificaForm(instance=gravedad_obj)
    
    return render(request, 'laboratorio/ensayos/gravedad_especifica/editar_gravedad_especifica.html', {
        'form': form,
        'gravedad_especifica_obj': gravedad_obj,
        'proyectos': proyectos
    })

#4 Ver
@login_required
def ver_gravedad_especifica(request, id):
    gravedad_obj = get_object_or_404(gravedad_especifica, id=id)
    
    # Preparar el historial con los cambios
    history_records = []
    for record in gravedad_obj.history.all():
        changes = []
        previous = record.prev_record
        if previous:
            # Comparar cada campo entre el registro actual y el anterior
            for field in record.history_object._meta.fields:
                old_value = getattr(previous, field.name, None)
                new_value = getattr(record, field.name, None)
                if old_value != new_value:
                    changes.append({
                        'field_name': field.verbose_name,
                        'old_value': old_value if old_value is not None else '-',
                        'new_value': new_value if new_value is not None else '-',
                    })
        history_records.append({
            'date': record.history_date,
            'user': record.history_user,
            'type': record.history_type,
            'reason': record.history_change_reason,
            'changes': changes if changes else None,
        })

    return render(request, 'laboratorio/ensayos/gravedad_especifica/ver_gravedad_especifica.html', {
        'gravedad': gravedad_obj,
        'history_records': history_records,
    })

#5 Eliminar
@login_required
def eliminar_gravedad_especifica(request, id):
    gravedad_obj = get_object_or_404(gravedad_especifica, id=id)
    if request.method == 'POST':
        gravedad_obj.delete()
        return redirect('listar_gravedad_especifica')
    return render(request, 'laboratorio/ensayos/gravedad_especifica/eliminar_gravedad_especifica.html', {
        'gravedad_especifica': gravedad_obj
    })

#6 Exportar excel
logger = logging.getLogger(__name__)

@login_required
def export_to_excel_gravedad_especifica(request):
    query = request.GET.get('q', '').strip()
    headers = ['ID', 'ID Proyecto', 'ID Prospección', 'Tipo Prospección', 'ID Muestra', 
               'Profundidad Desde', 'Profundidad Hasta', 'Profundidad Promedio', 
               'Gravedad Específica', 'Área', 'Usuario']

    # Crear el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="gravedad_especifica_filtrado.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Gravedad Específica"
    ws.append(headers)

    # Obtener queryset base
    try:
        gravedad_list = gravedad_especifica.objects.select_related(
            'id_proyecto', 'id_prospeccion', 'user'
        ).prefetch_related('id_muestra__muestreo_set').all()

        logger.debug(f"Registros iniciales: {gravedad_list.count()}, query: '{query}'")

        # Aplicar filtros si hay query
        if query:
            filtros = Q()
            
            # Campos numéricos
            try:
                query_numeric = float(query)
                filtros |= (
                    Q(profundidad_desde=query_numeric) |
                    Q(profundidad_hasta=query_numeric) |
                    Q(profundidad_promedio=query_numeric) |
                    Q(gravedad_especifica=query_numeric)
                )
            except ValueError:
                pass

            # Campos de texto y relaciones
            filtros |= (
                Q(id__icontains=query) |
                Q(id_proyecto__id__icontains=query) |
                Q(id_prospeccion__id_prospeccion__icontains=query) |
                Q(tipo_prospeccion__icontains=query) |
                Q(id_muestra__icontains=query) |
                Q(user__username__icontains=query)
            )

            gravedad_list = gravedad_list.filter(filtros)
        
        logger.debug(f"Registros después del filtro: {gravedad_list.count()}")

        # Construir las filas
        for gravedad in gravedad_list:
            try:
                muestreo = Muestreo.objects.filter(id_muestra=gravedad.id_muestra).first()
                
                row = [
                    str(gravedad.id) if gravedad.id is not None else '',
                    str(gravedad.id_proyecto.id) if gravedad.id_proyecto else '',
                    str(gravedad.id_prospeccion.id_prospeccion) if gravedad.id_prospeccion else '',
                    str(gravedad.tipo_prospeccion or ''),
                    str(gravedad.id_muestra or ''),
                    str(gravedad.profundidad_desde) if gravedad.profundidad_desde is not None else '',
                    str(gravedad.profundidad_hasta) if gravedad.profundidad_hasta is not None else '',
                    str(gravedad.profundidad_promedio) if gravedad.profundidad_promedio is not None else '',
                    str(gravedad.gravedad_especifica) if gravedad.gravedad_especifica is not None else '',
                    str(muestreo.area) if muestreo and hasattr(muestreo, 'area') else '',
                    str(gravedad.user.username) if gravedad.user else 'Sin usuario'
                ]
                ws.append(row)
            except Exception as e:
                logger.error(f"Error al procesar registro {gravedad.id}: {str(e)}")
                continue

        wb.save(response)
        return response

    except Exception as e:
        logger.error(f"Error en export_to_excel_gravedad_especifica: {str(e)}")
        ws.append(["Error al generar el reporte", str(e)])
        wb.save(response)
        return response
    
    
#7 Exportar PDF
logger = logging.getLogger(__name__)

@login_required
def export_to_pdf_gravedad_especifica(request):
    query = request.GET.get('q', '').strip()
    
    # Configurar respuesta HTTP para PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="gravedad_especifica_filtrado.pdf"'

    # Configurar documento PDF (carta horizontal)
    doc = SimpleDocTemplate(response, pagesize=landscape(letter), 
                          rightMargin=30, leftMargin=30, 
                          topMargin=30, bottomMargin=30)
    
    # Elementos del PDF
    elements = []
    styles = getSampleStyleSheet()
    
    # Título ajustado y reducido
    title = Paragraph("Lista de Gravedad Específica", 
                     styles['Heading1'].clone('Title', 
                                            alignment=1,  # Centrado
                                            fontSize=12,  # Tamaño reducido
                                            spaceAfter=10))  # Espacio después del título
    elements.append(title)
    elements.append(Paragraph("<br/>", styles['Normal']))  # Espacio reducido

    # Encabezados ajustados y reducidos
    headers = ['ID', 'Proyecto', 'Prospección', 'Tipo', 'Muestra', 
               'Prof. Ini', 'Prof. Fin', 'Prof. Prom', 
               'Grav. Esp', 'Área', 'Usuario']
    
    # Datos
    data = [headers]
    
    try:
        # Obtener queryset base
        gravedad_list = gravedad_especifica.objects.select_related(
            'id_proyecto', 'id_prospeccion', 'user'
        ).prefetch_related('id_muestra__muestreo_set').all()

        logger.debug(f"Registros iniciales: {gravedad_list.count()}, query: '{query}'")

        # Aplicar filtros si hay query
        if query:
            filtros = Q()
            try:
                query_numeric = float(query)
                filtros |= (
                    Q(profundidad_desde=query_numeric) |
                    Q(profundidad_hasta=query_numeric) |
                    Q(profundidad_promedio=query_numeric) |
                    Q(gravedad_especifica=query_numeric)
                )
            except ValueError:
                pass

            filtros |= (
                Q(id__icontains=query) |
                Q(id_proyecto__id__icontains=query) |
                Q(id_prospeccion__id_prospeccion__icontains=query) |
                Q(tipo_prospeccion__icontains=query) |
                Q(id_muestra__icontains=query) |
                Q(user__username__icontains=query)
            )

            gravedad_list = gravedad_list.filter(filtros)
        
        logger.debug(f"Registros después del filtro: {gravedad_list.count()}")

        # Construir las filas
        for gravedad in gravedad_list:
            try:
                muestreo = Muestreo.objects.filter(id_muestra=gravedad.id_muestra).first()
                
                row = [
                    str(gravedad.id) if gravedad.id is not None else '',
                    str(gravedad.id_proyecto.id) if gravedad.id_proyecto else '',
                    str(gravedad.id_prospeccion.id_prospeccion) if gravedad.id_prospeccion else '',
                    str(gravedad.tipo_prospeccion or ''),
                    str(gravedad.id_muestra or ''),
                    str(gravedad.profundidad_desde) if gravedad.profundidad_desde is not None else '',
                    str(gravedad.profundidad_hasta) if gravedad.profundidad_hasta is not None else '',
                    str(gravedad.profundidad_promedio) if gravedad.profundidad_promedio is not None else '',
                    str(gravedad.gravedad_especifica) if gravedad.gravedad_especifica is not None else '',
                    str(muestreo.area) if muestreo and hasattr(muestreo, 'area') else '',
                    str(gravedad.user.username) if gravedad.user else 'Sin usuario'
                ]
                data.append(row)
            except Exception as e:
                logger.error(f"Error al procesar registro {gravedad.id}: {str(e)}")
                continue

        # Crear tabla con tamaños ajustados
        table = Table(data, colWidths=[0.5*inch] + [0.8*inch]*10)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),    # Tamaño reducido para encabezados
            ('FONTSIZE', (0, 1), (-1, -1), 8),   # Tamaño de datos
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),  # Padding reducido
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        elements.append(table)
        doc.build(elements)
        return response

    except Exception as e:
        logger.error(f"Error en export_to_pdf_gravedad_especifica: {str(e)}")
        doc.build([Paragraph(f"Error al generar el reporte: {str(e)}", styles['Normal'])])
        return response

#8 Garficos Gravedad especifica
# Funciones de gráficos por área
def generar_grafico_gravedad_area(df):
    areas_unicas = df['area'].unique()
    colores = {area: (random.random(), random.random(), random.random()) for area in areas_unicas}
    fig, ax = plt.subplots()
    
    for area in areas_unicas:
        datos_area = df[df['area'] == area]
        ax.scatter(
            datos_area['gravedad_especifica'],
            datos_area['profundidad_promedio'],
            color=colores[area],
            s=100,
            label=area,
            marker='+'
        )
    
    ax.xaxis.set_ticks_position("top")
    ax.xaxis.set_label_position("top")
    ax.invert_yaxis()
    ax.set_xlabel("Gravedad Específica")
    ax.set_ylabel("Profundidad (m)")
    ax.grid(True)
    plt.subplots_adjust(right=0.75)
    num_columns = (len(areas_unicas) + 24) // 25
    legend = ax.legend(loc='center left', bbox_to_anchor=(1, 0.5), ncol=num_columns)
    plt.title("GRÁFICO AGRUPADO POR ÁREA (DISPERSIÓN)", pad=30)
    
    buffer = BytesIO()
    fig.savefig(buffer, format='png', dpi=100, bbox_inches='tight', bbox_extra_artists=[legend])
    buffer.seek(0)
    plt.close()
    image_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
    return image_base64

# Funciones de gráficos por prospección
def generar_grafico_gravedad_prospeccion(df):
    df['id_prospeccion'] = df['id_prospeccion'].astype(str)
    df['id_prospeccion_unico'] = df.groupby('id_prospeccion').cumcount() + 1
    df['etiqueta'] = df['id_prospeccion'] + '-' + df['id_prospeccion_unico'].astype(str)
    prospecciones_unicas = df['etiqueta'].unique()
    colores = {prospeccion: (random.random(), random.random(), random.random()) for prospeccion in prospecciones_unicas}
    
    fig, ax = plt.subplots()
    
    for prospeccion in prospecciones_unicas:
        datos_prospeccion = df[df['etiqueta'] == prospeccion]
        ax.scatter(
            datos_prospeccion['gravedad_especifica'],
            datos_prospeccion['profundidad_promedio'],
            color=colores[prospeccion],
            s=100,
            label=prospeccion,
            marker='+'
        )
    
    ax.xaxis.set_ticks_position("top")
    ax.xaxis.set_label_position("top")
    ax.invert_yaxis()
    ax.set_xlabel("Gravedad Específica")
    ax.set_ylabel("Profundidad (m)")
    ax.grid(True)
    plt.subplots_adjust(right=0.75)
    num_columns = (len(prospecciones_unicas) + 24) // 25
    legend = ax.legend(loc='center left', bbox_to_anchor=(1, 0.5), ncol=num_columns)
    plt.title("GRÁFICO AGRUPADO POR PROSPECCIÓN (DISPERSIÓN)", pad=30)
    
    buffer = BytesIO()
    fig.savefig(buffer, format='png', dpi=100, bbox_inches='tight', bbox_extra_artists=[legend])
    buffer.seek(0)
    plt.close()
    image_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
    return image_base64

# Vista principal para gráficos de Gravedad Específica
@login_required
def graficos_gravedad_especifica(request):
    id_proyectos = request.GET.getlist('id_proyectos')
    tipos_prospeccion = request.GET.getlist('tipos_prospeccion')
    areas = request.GET.getlist('areas')
    id_prospecciones = request.GET.getlist('id_prospecciones')

    query = gravedad_especifica.objects.all()  # Nombre correcto del modelo
    if id_proyectos:
        query = query.filter(id_proyecto__in=id_proyectos)
    if tipos_prospeccion:
        query = query.filter(tipo_prospeccion__in=tipos_prospeccion)
    if areas:
        query = query.filter(area__in=areas)
    if id_prospecciones:
        query = query.filter(id_prospeccion__in=id_prospecciones)

    proyectos = query.values('id_proyecto').distinct()
    tipos_prospeccion_inicial = query.values_list('tipo_prospeccion', flat=True).distinct().exclude(tipo_prospeccion__isnull=True)
    areas_inicial = query.values_list('area', flat=True).distinct().exclude(area__isnull=True).exclude(area='')
    id_prospecciones_inicial = query.values_list('id_prospeccion', flat=True).distinct().exclude(id_prospeccion__isnull=True)

    # Ajustamos los nombres de los campos para ForeignKey (usamos id_prospeccion_id)
    df = pd.DataFrame(list(query.values('gravedad_especifica', 'profundidad_promedio', 'area', 'id_prospeccion_id')))
    df = df.rename(columns={'id_prospeccion_id': 'id_prospeccion'})  # Renombramos para consistencia en las funciones

    context = {
        'proyectos': proyectos,
        'tipos_prospeccion_inicial': tipos_prospeccion_inicial,
        'areas_inicial': areas_inicial,
        'id_prospecciones_inicial': id_prospecciones_inicial,
        'selected_id_proyectos': json.dumps(id_proyectos),
        'selected_tipos_prospeccion': json.dumps(tipos_prospeccion),
        'selected_areas': json.dumps(areas),
        'selected_id_prospecciones': json.dumps(id_prospecciones),
    }

    if df.empty:
        context['error'] = "No hay datos disponibles para los filtros seleccionados."
    else:
        context['image_base64_area'] = generar_grafico_gravedad_area(df)
        context['image_base64_prospeccion'] = generar_grafico_gravedad_prospeccion(df)

    return render(request, 'laboratorio/ensayos/gravedad_especifica/graficos_gravedad_especifica.html', context)

# Funciones auxiliares para Gravedad Específica
def obtener_tipos_prospeccion_gravedad(request):
    id_proyectos = request.GET.get('id_proyectos')
    query = gravedad_especifica.objects.all()  # Nombre correcto del modelo
    if id_proyectos:
        query = query.filter(id_proyecto__in=id_proyectos.split(','))
    tipos_list = list(query.values_list('tipo_prospeccion', flat=True).distinct().exclude(tipo_prospeccion__isnull=True))
    return JsonResponse({'options': tipos_list}, safe=False)

def obtener_id_prospecciones_gravedad(request):
    id_proyectos = request.GET.get('id_proyectos')
    tipos_prospeccion = request.GET.get('tipos_prospeccion')
    areas = request.GET.get('areas')
    query = gravedad_especifica.objects.all()  # Nombre correcto del modelo
    if id_proyectos:
        query = query.filter(id_proyecto__in=id_proyectos.split(','))
    if tipos_prospeccion:
        query = query.filter(tipo_prospeccion__in=tipos_prospeccion.split(','))
    if areas:
        query = query.filter(area__in=areas.split(','))
    id_prospecciones_list = list(query.values_list('id_prospeccion', flat=True).distinct().exclude(id_prospeccion__isnull=True))
    return JsonResponse({'options': id_prospecciones_list}, safe=False)

def obtener_area_gravedad(request):
    id_prospecciones = request.GET.get('id_prospecciones')
    if id_prospecciones:
        query = gravedad_especifica.objects.filter(id_prospeccion__in=id_prospecciones.split(','))  # Nombre correcto del modelo
        areas_list = list(query.values_list('area', flat=True).distinct().exclude(area__isnull=True).exclude(area=''))
        return JsonResponse({'options': areas_list})
    return JsonResponse({'options': []})


#### Limites Atterberg 

####### LÍMITES DE ATTERBERG #######################
# views.py
logger = logging.getLogger(__name__)

# 1 Agregar
@login_required
def agregar_limites_atterberg(request):
    proyectos_con_muestreo = Proyectos.objects.filter(muestreo__isnull=False).distinct()
    tipo_prospeccion_choices = []
    prospecciones = []
    muestras = []
    area = ""

    if request.method == 'POST':
        logger.info(f"Datos POST recibidos: {request.POST}")
        proyecto_id = request.POST.get('id_proyecto')
        tipo_prospeccion = request.POST.get('tipo_prospeccion')
        id_prospecciones = request.POST.getlist('id_prospeccion[]')
        id_muestras = request.POST.getlist('id_muestra[]')
        limites_liquido = request.POST.getlist('limite_liquido[]')
        limites_plastico = request.POST.getlist('limite_plastico[]')
        metodos = request.POST.getlist('metodo[]')
        acanalados = request.POST.getlist('acanalado[]')
        areas = request.POST.getlist('area[]')
        errors = []

        if not proyecto_id:
            errors.append("Debe seleccionar un proyecto.")
            return render(request, 'laboratorio/ensayos/limites_atterberg/agregar_limites_atterberg.html', {
                'proyectos': proyectos_con_muestreo,
                'errors': errors
            })

        muestreos = Muestreo.objects.filter(id_proyecto_id=proyecto_id)
        if not muestreos.exists():
            errors.append("No hay muestreos asociados a este proyecto.")
            return render(request, 'laboratorio/ensayos/limites_atterberg/agregar_limites_atterberg.html', {
                'proyectos': proyectos_con_muestreo,
                'errors': errors
            })

        if tipo_prospeccion:
            muestreos = muestreos.filter(tipo_prospeccion=tipo_prospeccion)
            if not muestreos.exists():
                errors.append("No hay muestreos con este tipo de prospección.")

        if id_prospecciones and id_prospecciones[0]:
            muestreos = muestreos.filter(id_prospeccion__id_prospeccion__in=id_prospecciones)
            if not muestreos.exists():
                errors.append("No hay muestreos con las prospecciones seleccionadas.")

        if id_muestras and id_muestras[0]:
            muestreos = muestreos.filter(id_muestra__in=id_muestras)
            if not muestreos.exists():
                errors.append("No hay muestreos con las muestras seleccionadas.")

        if not (len(id_prospecciones) == len(id_muestras) == len(limites_liquido) == len(limites_plastico) == len(metodos) == len(acanalados) == len(areas)):
            errors.append("El número de prospecciones, muestras, límites líquido, límites plástico, métodos, acanalados y áreas no coincide.")
            logger.error(f"Longitudes inconsistentes: prospecciones={len(id_prospecciones)}, muestras={len(id_muestras)}, límites líquido={len(limites_liquido)}, límites plástico={len(limites_plastico)}, métodos={len(metodos)}, acanalados={len(acanalados)}, áreas={len(areas)}")
        else:
            for id_pros, id_muestra, ll, lp, metodo, acanalado, area_val in zip(id_prospecciones, id_muestras, limites_liquido, limites_plastico, metodos, acanalados, areas):
                if not id_pros or not id_muestra or not ll or not lp:
                    errors.append(f"Falta ID de prospección ({id_pros}), ID de muestra ({id_muestra}), límite líquido ({ll}) o límite plástico ({lp}).")
                    continue
                
                try:
                    muestreo = muestreos.get(id_prospeccion__id_prospeccion=id_pros, id_muestra=id_muestra)
                    profundidad_desde = float(muestreo.profundidad_desde or 0)  # Si es None, usa 0
                    profundidad_hasta = float(muestreo.profundidad_hasta or 0)  # Si es None, usa 0
                    profundidad_promedio = (profundidad_desde + profundidad_hasta) / 2

                    indice_plasticidad = float(ll) - float(lp) if ll and lp else "NP"
                    limites_obj = Limites_atterberg(
                        id_proyecto=muestreo.id_proyecto,
                        tipo_prospeccion=muestreo.tipo_prospeccion,
                        id_prospeccion=muestreo.id_prospeccion,
                        id_muestra=muestreo.id_muestra,
                        profundidad_desde=muestreo.profundidad_desde,
                        profundidad_hasta=muestreo.profundidad_hasta,
                        profundidad_promedio=profundidad_promedio,
                        limite_liquido=float(ll),
                        limite_plastico=float(lp),
                        indice_plasticidad=str(indice_plasticidad),
                        metodo=metodo,
                        acanalado=acanalado,
                        area=area_val,
                        user=request.user
                    )
                    limites_obj.save()
                    logger.info(f"Límites de Atterberg guardados: ID={limites_obj.id}, Muestra={id_muestra}, LL={ll}, LP={lp}, IP={indice_plasticidad}, Profundidad Promedio={profundidad_promedio}")
                except Muestreo.DoesNotExist:
                    errors.append(f"Muestra {id_muestra} con prospección {id_pros} no encontrada.")
                except ValueError as e:
                    errors.append(f"Valor inválido para límite líquido {ll} o límite plástico {lp}: {e}")
                except Exception as e:
                    errors.append(f"Error al guardar muestra {id_muestra}: {e}")

        if errors:
            tipo_prospeccion_choices = Muestreo.objects.filter(id_proyecto_id=proyecto_id).values_list('tipo_prospeccion', flat=True).distinct()
            prospecciones = Muestreo.objects.filter(id_proyecto_id=proyecto_id).values_list('id_prospeccion__id_prospeccion', flat=True).distinct()
            muestras = Muestreo.objects.filter(id_proyecto_id=proyecto_id).values_list('id_muestra', flat=True).distinct()
            return render(request, 'laboratorio/ensayos/limites_atterberg/agregar_limites_atterberg.html', {
                'proyectos': proyectos_con_muestreo,
                'tipo_prospeccion_choices': tipo_prospeccion_choices,
                'prospecciones': prospecciones,
                'muestras': muestras,
                'area': area,
                'errors': errors
            })

        return redirect('listar_limites_atterberg')

    proyecto_id = request.GET.get('id_proyecto')
    tipo_prospeccion = request.GET.get('tipo_prospeccion')
    id_prospeccion = request.GET.get('id_prospeccion')
    id_muestra = request.GET.get('id_muestra')

    muestreos = Muestreo.objects.all()
    if proyecto_id:
        muestreos = muestreos.filter(id_proyecto_id=proyecto_id)
        tipo_prospeccion_choices = muestreos.values_list('tipo_prospeccion', flat=True).distinct()
        prospecciones = muestreos.values_list('id_prospeccion__id_prospeccion', flat=True).distinct()
        muestras = muestreos.values_list('id_muestra', flat=True).distinct()
        area = muestreos.first().area if muestreos.exists() else ""

    if tipo_prospeccion:
        muestreos = muestreos.filter(tipo_prospeccion=tipo_prospeccion)
        prospecciones = muestreos.values_list('id_prospeccion__id_prospeccion', flat=True).distinct()
        muestras = muestreos.values_list('id_muestra', flat=True).distinct()

    if id_prospeccion:
        muestreos = muestreos.filter(id_prospeccion__id_prospeccion=id_prospeccion)
        muestras = muestreos.values_list('id_muestra', flat=True).distinct()
        area = muestreos.first().area if muestreos.exists() else ""

    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if is_ajax:
        data = {
            'tipo_prospeccion_choices': list(tipo_prospeccion_choices),
            'prospecciones': list(prospecciones),
            'muestras': list(muestras),
            'area': area
        }
        if id_muestra:
            muestreo = muestreos.filter(id_muestra=id_muestra).first()
            if muestreo:
                data['profundidad_desde'] = str(muestreo.profundidad_desde) if muestreo.profundidad_desde is not None else '0'
                data['profundidad_hasta'] = str(muestreo.profundidad_hasta) if muestreo.profundidad_hasta is not None else '0'
        return JsonResponse(data)

    return render(request, 'laboratorio/ensayos/limites_atterberg/agregar_limites_atterberg.html', {
        'proyectos': proyectos_con_muestreo,
        'tipo_prospeccion_choices': tipo_prospeccion_choices,
        'prospecciones': prospecciones,
        'muestras': muestras,
        'area': area
    })

# 2 Listar
@login_required
def listar_limites_atterberg(request):
    limites_list = Limites_atterberg.objects.all().order_by('id')
    query = request.GET.get('q', '')

    if query:
        limites_model = apps.get_model('administrador', 'Limites_atterberg')
        campos = [field.name for field in limites_model._meta.get_fields() if isinstance(field, (models.CharField, models.TextField, models.DateField, models.DateTimeField, models.DecimalField, models.FloatField, models.ForeignKey))]
        query_filter = Q()

        try:
            query_numeric = float(query)
            numeric_search = True
        except ValueError:
            numeric_search = False

        for field in campos:
            field_instance = limites_model._meta.get_field(field)
            if isinstance(field_instance, (models.DateField, models.DateTimeField)):
                query_filter |= Q(**{f"{field}__year__icontains": query})
                query_filter |= Q(**{f"{field}__month__icontains": query})
                query_filter |= Q(**{f"{field}__day__icontains": query})
            elif isinstance(field_instance, models.ForeignKey):
                related_model = field_instance.related_model
                related_fields = [f"{field}__{related_field.name}" for related_field in related_model._meta.get_fields() if isinstance(related_field, (models.CharField, models.TextField))]
                for related_field in related_fields:
                    query_filter |= Q(**{f"{related_field}__icontains": query})
            elif isinstance(field_instance, (models.DecimalField, models.FloatField)) and numeric_search:
                query_filter |= Q(**{f"{field}": query_numeric})
            else:
                query_filter |= Q(**{f"{field}__icontains": query})

        limites_list = limites_list.filter(query_filter)

    paginator = Paginator(limites_list, 1000)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'laboratorio/ensayos/limites_atterberg/limites_atterberg_table.html', {'page_obj': page_obj})

    return render(request, 'laboratorio/ensayos/limites_atterberg/listar_limites_atterberg.html', {
        'page_obj': page_obj,
        'query': query
    })

# 3 Editar
@login_required
def editar_limites_atterberg(request, id):
    limites_obj = get_object_or_404(Limites_atterberg, id=id)
    muestreo = Muestreo.objects.filter(id_muestra=limites_obj.id_muestra).first()
    proyectos = Proyectos.objects.filter(muestreo__isnull=False).distinct()

    if request.method == 'POST':
        form = LimitesAtterbergForm(request.POST, instance=limites_obj)
        if form.is_valid():
            try:
                limites_obj = form.save()
                logger.info(f"Límites de Atterberg editados: ID {limites_obj.id}")
                return redirect('listar_limites_atterberg')
            except Exception as e:
                logger.error(f"Error al editar: {e}")
                form.add_error(None, f"Error al editar: {e}")
    else:
        form = LimitesAtterbergForm(instance=limites_obj)
    
    return render(request, 'laboratorio/ensayos/limites_atterberg/editar_limites_atterberg.html', {
        'form': form,
        'limites_atterberg_obj': limites_obj,
        'proyectos': proyectos
    })

# 4 Ver
@login_required
def ver_limites_atterberg(request, id):
    limites_obj = get_object_or_404(Limites_atterberg, id=id)
    
    history_records = []
    for record in limites_obj.history.all():
        changes = []
        previous = record.prev_record
        if previous:
            for field in record.history_object._meta.fields:
                old_value = getattr(previous, field.name, None)
                new_value = getattr(record, field.name, None)
                if old_value != new_value:
                    changes.append({
                        'field_name': field.verbose_name,
                        'old_value': old_value if old_value is not None else '-',
                        'new_value': new_value if new_value is not None else '-',
                    })
        history_records.append({
            'date': record.history_date,
            'user': record.history_user,
            'type': record.history_type,
            'reason': record.history_change_reason,
            'changes': changes if changes else None,
        })

    return render(request, 'laboratorio/ensayos/limites_atterberg/ver_limites_atterberg.html', {
        'limites_atterberg': limites_obj,
        'history_records': history_records,
    })

# 5 Eliminar
@login_required
def eliminar_limites_atterberg(request, id):
    limites_obj = get_object_or_404(Limites_atterberg, id=id)
    if request.method == 'POST':
        limites_obj.delete()
        return redirect('listar_limites_atterberg')
    return render(request, 'laboratorio/ensayos/limites_atterberg/eliminar_limites_atterberg.html', {
        'limites_atterberg': limites_obj
    })

# 6 Exportar Excel
@login_required
def export_to_excel_limites_atterberg(request):
    query = request.GET.get('q', '').strip()
    headers = ['ID', 'ID Proyecto', 'ID Prospección', 'Tipo Prospección', 'ID Muestra', 
               'Profundidad Desde', 'Profundidad Hasta', 'Profundidad Promedio', 
               'Límite Líquido', 'Límite Plástico', 'Índice de Plasticidad', 
               'Método', 'Acanalado', 'Área', 'Usuario']

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="limites_atterberg_filtrado.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Límites de Atterberg"
    ws.append(headers)

    try:
        limites_list = Limites_atterberg.objects.select_related(
            'id_proyecto', 'id_prospeccion', 'user'
        ).prefetch_related('id_muestra__muestreo_set').all()

        logger.debug(f"Registros iniciales: {limites_list.count()}, query: '{query}'")

        if query:
            filtros = Q()
            try:
                query_numeric = float(query)
                filtros |= (
                    Q(profundidad_desde=query_numeric) |
                    Q(profundidad_hasta=query_numeric) |
                    Q(profundidad_promedio=query_numeric) |
                    Q(limite_liquido=query_numeric) |
                    Q(limite_plastico=query_numeric)
                )
            except ValueError:
                pass

            filtros |= (
                Q(id__icontains=query) |
                Q(id_proyecto__id__icontains=query) |
                Q(id_prospeccion__id_prospeccion__icontains=query) |
                Q(tipo_prospeccion__icontains=query) |
                Q(id_muestra__icontains=query) |
                Q(indice_plasticidad__icontains=query) |
                Q(metodo__icontains=query) |
                Q(acanalado__icontains=query) |
                Q(area__icontains=query) |
                Q(user__username__icontains=query)
            )

            limites_list = limites_list.filter(filtros)
        
        logger.debug(f"Registros después del filtro: {limites_list.count()}")

        for limites in limites_list:
            try:
                muestreo = Muestreo.objects.filter(id_muestra=limites.id_muestra).first()
                
                row = [
                    str(limites.id) if limites.id is not None else '',
                    str(limites.id_proyecto.id) if limites.id_proyecto else '',
                    str(limites.id_prospeccion.id_prospeccion) if limites.id_prospeccion else '',
                    str(limites.tipo_prospeccion or ''),
                    str(limites.id_muestra or ''),
                    str(limites.profundidad_desde) if limites.profundidad_desde is not None else '',
                    str(limites.profundidad_hasta) if limites.profundidad_hasta is not None else '',
                    str(limites.profundidad_promedio) if limites.profundidad_promedio is not None else '',
                    str(limites.limite_liquido) if limites.limite_liquido is not None else '',
                    str(limites.limite_plastico) if limites.limite_plastico is not None else '',
                    str(limites.indice_plasticidad or ''),
                    str(limites.metodo or ''),
                    str(limites.acanalado or ''),
                    str(muestreo.area) if muestreo and hasattr(muestreo, 'area') else '',
                    str(limites.user.username) if limites.user else 'Sin usuario'
                ]
                ws.append(row)
            except Exception as e:
                logger.error(f"Error al procesar registro {limites.id}: {str(e)}")
                continue

        wb.save(response)
        return response

    except Exception as e:
        logger.error(f"Error en export_to_excel_limites_atterberg: {str(e)}")
        ws.append(["Error al generar el reporte", str(e)])
        wb.save(response)
        return response

# 7 Exportar PDF
@login_required
def export_to_pdf_limites_atterberg(request):
    query = request.GET.get('q', '').strip()
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="limites_atterberg_filtrado.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(letter), 
                          rightMargin=30, leftMargin=30, 
                          topMargin=30, bottomMargin=30)
    
    elements = []
    styles = getSampleStyleSheet()
    
    title = Paragraph("Lista de Límites de Atterberg", 
                     styles['Heading1'].clone('Title', alignment=1, fontSize=12, spaceAfter=10))
    elements.append(title)
    elements.append(Paragraph("<br/>", styles['Normal']))

    headers = ['ID', 'Proyecto', 'Prospección', 'Tipo', 'Muestra', 
               'Prof. Ini', 'Prof. Fin', 'Prof. Prom', 
               'LL', 'LP', 'IP', 'Método', 'Acanalado', 'Área', 'Usuario']
    
    data = [headers]
    
    try:
        limites_list = Limites_atterberg.objects.select_related(
            'id_proyecto', 'id_prospeccion', 'user'
        ).prefetch_related('id_muestra__muestreo_set').all()

        logger.debug(f"Registros iniciales: {limites_list.count()}, query: '{query}'")

        if query:
            filtros = Q()
            try:
                query_numeric = float(query)
                filtros |= (
                    Q(profundidad_desde=query_numeric) |
                    Q(profundidad_hasta=query_numeric) |
                    Q(profundidad_promedio=query_numeric) |
                    Q(limite_liquido=query_numeric) |
                    Q(limite_plastico=query_numeric)
                )
            except ValueError:
                pass

            filtros |= (
                Q(id__icontains=query) |
                Q(id_proyecto__id__icontains=query) |
                Q(id_prospeccion__id_prospeccion__icontains=query) |
                Q(tipo_prospeccion__icontains=query) |
                Q(id_muestra__icontains=query) |
                Q(indice_plasticidad__icontains=query) |
                Q(metodo__icontains=query) |
                Q(acanalado__icontains=query) |
                Q(area__icontains=query) |
                Q(user__username__icontains=query)
            )

            limites_list = limites_list.filter(filtros)
        
        logger.debug(f"Registros después del filtro: {limites_list.count()}")

        for limites in limites_list:
            try:
                muestreo = Muestreo.objects.filter(id_muestra=limites.id_muestra).first()
                
                row = [
                    str(limites.id) if limites.id is not None else '',
                    str(limites.id_proyecto.id) if limites.id_proyecto else '',
                    str(limites.id_prospeccion.id_prospeccion) if limites.id_prospeccion else '',
                    str(limites.tipo_prospeccion or ''),
                    str(limites.id_muestra or ''),
                    str(limites.profundidad_desde) if limites.profundidad_desde is not None else '',
                    str(limites.profundidad_hasta) if limites.profundidad_hasta is not None else '',
                    str(limites.profundidad_promedio) if limites.profundidad_promedio is not None else '',
                    str(limites.limite_liquido) if limites.limite_liquido is not None else '',
                    str(limites.limite_plastico) if limites.limite_plastico is not None else '',
                    str(limites.indice_plasticidad or ''),
                    str(limites.metodo or ''),
                    str(limites.acanalado or ''),
                    str(muestreo.area) if muestreo and hasattr(muestreo, 'area') else '',
                    str(limites.user.username) if limites.user else 'Sin usuario'
                ]
                data.append(row)
            except Exception as e:
                logger.error(f"Error al procesar registro {limites.id}: {str(e)}")
                continue

        table = Table(data, colWidths=[0.5*inch] + [0.8*inch]*14)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        elements.append(table)
        doc.build(elements)
        return response

    except Exception as e:
        logger.error(f"Error en export_to_pdf_limites_atterberg: {str(e)}")
        doc.build([Paragraph(f"Error al generar el reporte: {str(e)}", styles['Normal'])])
        return response
# 8 Gráficos Límites de Atterberg
def generar_grafico_carta_plasticidad_area(df):
    areas = df['area'].unique()
    colores = {area: (random.random(), random.random(), random.random()) for area in areas}
    size = pd.to_numeric(df['indice_plasticidad'], errors='coerce').fillna(0) * 5

    fig, ax = plt.subplots()

    for area in areas:
        datos_area = df[df['area'] == area]
        ax.scatter(
            datos_area['limite_liquido'],
            pd.to_numeric(datos_area['indice_plasticidad'], errors='coerce'),
            c=[colores[area]],
            s=size[datos_area.index],
            label=area
        )

    ax.set_xlim(0, 80)
    ax.set_ylim(0, 50)

    x1 = [15.7777777777777, 74.6666666666666]
    y1 = [7, 60]
    x2 = [25.4794520547945, 102.191780821917]
    y2 = [4, 60]
    x3 = [0, 29.5890410958904]
    y3 = [7, 7]
    x4 = [0, 25.479452054794]
    y4 = [4, 4]

    ax.plot(x1, y1, color='green', linestyle='-', linewidth=1, zorder=0)
    ax.plot(x2, y2, color='blue', linestyle='-', linewidth=1, zorder=0)
    ax.plot(x3, y3, color='red', linestyle='-', linewidth=1, zorder=0)
    ax.plot(x4, y4, color='purple', linestyle='-', linewidth=1, zorder=0)

    def add_text_with_box(ax, x, y, text, color, pad=0.5, boxstyle='square'):
        ax.text(x, y, text, color=color, va='bottom', ha='center', fontsize=8,
                bbox=dict(facecolor='white', edgecolor="black", boxstyle=boxstyle, pad=pad))

    add_text_with_box(ax, 13, 52, "PLASTICIDAD BAJA", color='black', pad=0.3)
    add_text_with_box(ax, 40, 52, 'PLASTICIDAD MEDIA', color='black', pad=0.3)
    add_text_with_box(ax, 66, 52, 'PLASTICIDAD ALTA', color='black', pad=0.3)

    ax.text(3, 8, 'IP=7', color='red', va='bottom', ha='center', fontsize=10, bbox=dict(facecolor='white', edgecolor='none', alpha=0))
    ax.text(3, 1, 'IP=4', color='purple', va='bottom', ha='center', fontsize=10, bbox=dict(facecolor='white', edgecolor='none', alpha=0))
    ax.text(8, 3.3, 'Arcilla/Limo\nCL-ML', color='black', va='bottom', ha='center', fontsize=10, bbox=dict(facecolor='white', edgecolor='none', alpha=0))
    ax.text(29, 12.5, 'Arcilla\nCL u OL', color='black', va='bottom', ha='center', fontsize=10, bbox=dict(facecolor='white', edgecolor='none', alpha=0))
    ax.text(40, 20, 'Arcilla\nCL u OL', color='black', va='bottom', ha='center', fontsize=10, bbox=dict(facecolor='white', edgecolor='none', alpha=0))
    ax.text(60, 37.5, 'Arcilla\nCH u OH', color='black', va='bottom', ha='center', fontsize=10, bbox=dict(facecolor='white', edgecolor='none', alpha=0))
    ax.text(40, 5, 'Limo\nML u OL', color='black', va='bottom', ha='center', fontsize=10, bbox=dict(facecolor='white', edgecolor='none', alpha=0))
    ax.text(60, 17.5, 'Limo\nMH u OH', color='black', va='bottom', ha='center', fontsize=10, bbox=dict(facecolor='white', edgecolor='none', alpha=0))
    ax.text(50, 35, '"U" Line\nIP=0,73*(LL-8)', color='green', va='bottom', ha='center', fontsize=10, rotation=47, bbox=dict(facecolor='none', edgecolor='none', alpha=0))
    ax.text(65, 23.5, '"A" Line\nIP=0,9*(LL-8)', color='blue', va='bottom', ha='center', fontsize=10, rotation=44, bbox=dict(facecolor='none', edgecolor='none', alpha=0))

    ax.set_xlabel('Límite Líquido (%)')
    ax.set_ylabel('Índice de Plasticidad')
    ax.legend(loc='center left', bbox_to_anchor=(1, 0.5))
    plt.title("CARTA DE PLASTICIDAD", pad=30)
    plt.grid(True)

    buffer = BytesIO()
    fig.savefig(buffer, format='png', dpi=100, bbox_inches="tight", pad_inches=0.5)
    buffer.seek(0)
    plt.close()
    image_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
    return image_base64

def generar_grafico_carta_plasticidad_prospeccion(df):
    df['id_prospeccion'] = df['id_prospeccion'].astype(str)
    df['id_prospeccion_unico'] = df.groupby('id_prospeccion').cumcount() + 1
    df['etiqueta'] = df['id_prospeccion'] + '-' + df['id_prospeccion_unico'].astype(str)
    prospecciones_unicas = df['etiqueta'].unique()
    colores = {prospeccion: (random.random(), random.random(), random.random()) for prospeccion in prospecciones_unicas}
    size = pd.to_numeric(df['indice_plasticidad'], errors='coerce').fillna(0) * 5

    fig, ax = plt.subplots()

    for prospeccion in prospecciones_unicas:
        datos_prospeccion = df[df['etiqueta'] == prospeccion]
        ax.scatter(
            datos_prospeccion['limite_liquido'],
            pd.to_numeric(datos_prospeccion['indice_plasticidad'], errors='coerce'),
            c=[colores[prospeccion]],
            s=size[datos_prospeccion.index],
            label=prospeccion
        )

    ax.set_xlim(0, 80)
    ax.set_ylim(0, 50)

    x1 = [15.7777777777777, 74.6666666666666]
    y1 = [7, 60]
    x2 = [25.4794520547945, 102.191780821917]
    y2 = [4, 60]
    x3 = [0, 29.5890410958904]
    y3 = [7, 7]
    x4 = [0, 25.479452054794]
    y4 = [4, 4]

    ax.plot(x1, y1, color='green', linestyle='-', linewidth=1, zorder=0)
    ax.plot(x2, y2, color='blue', linestyle='-', linewidth=1, zorder=0)
    ax.plot(x3, y3, color='red', linestyle='-', linewidth=1, zorder=0)
    ax.plot(x4, y4, color='purple', linestyle='-', linewidth=1, zorder=0)

    def add_text_with_box(ax, x, y, text, color, pad=0.5, boxstyle='square'):
        ax.text(x, y, text, color=color, va='bottom', ha='center', fontsize=8,
                bbox=dict(facecolor='white', edgecolor="black", boxstyle=boxstyle, pad=pad))

    add_text_with_box(ax, 13, 52, "PLASTICIDAD BAJA", color='black', pad=0.3)
    add_text_with_box(ax, 40, 52, 'PLASTICIDAD MEDIA', color='black', pad=0.3)
    add_text_with_box(ax, 66, 52, 'PLASTICIDAD ALTA', color='black', pad=0.3)

    ax.text(3, 8, 'IP=7', color='red', va='bottom', ha='center', fontsize=10, bbox=dict(facecolor='white', edgecolor='none', alpha=0))
    ax.text(3, 1, 'IP=4', color='purple', va='bottom', ha='center', fontsize=10, bbox=dict(facecolor='white', edgecolor='none', alpha=0))
    ax.text(8, 3.3, 'Arcilla/Limo\nCL-ML', color='black', va='bottom', ha='center', fontsize=10, bbox=dict(facecolor='white', edgecolor='none', alpha=0))
    ax.text(29, 12.5, 'Arcilla\nCL u OL', color='black', va='bottom', ha='center', fontsize=10, bbox=dict(facecolor='white', edgecolor='none', alpha=0))
    ax.text(40, 20, 'Arcilla\nCL u OL', color='black', va='bottom', ha='center', fontsize=10, bbox=dict(facecolor='white', edgecolor='none', alpha=0))
    ax.text(60, 37.5, 'Arcilla\nCH u OH', color='black', va='bottom', ha='center', fontsize=10, bbox=dict(facecolor='white', edgecolor='none', alpha=0))
    ax.text(40, 5, 'Limo\nML u OL', color='black', va='bottom', ha='center', fontsize=10, bbox=dict(facecolor='white', edgecolor='none', alpha=0))
    ax.text(60, 17.5, 'Limo\nMH u OH', color='black', va='bottom', ha='center', fontsize=10, bbox=dict(facecolor='white', edgecolor='none', alpha=0))
    ax.text(50, 35, '"U" Line\nIP=0,73*(LL-8)', color='green', va='bottom', ha='center', fontsize=10, rotation=47, bbox=dict(facecolor='none', edgecolor='none', alpha=0))
    ax.text(65, 23.5, '"A" Line\nIP=0,9*(LL-8)', color='blue', va='bottom', ha='center', fontsize=10, rotation=44, bbox=dict(facecolor='none', edgecolor='none', alpha=0))

    ax.set_xlabel('Límite Líquido (%)')
    ax.set_ylabel('Índice de Plasticidad')
    ax.legend(loc='center left', bbox_to_anchor=(1, 0.5))
    plt.title("CARTA DE PLASTICIDAD", pad=30)
    plt.grid(True)

    buffer = BytesIO()
    fig.savefig(buffer, format='png', dpi=100, bbox_inches="tight", pad_inches=0.5)
    buffer.seek(0)
    plt.close()
    image_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
    return image_base64

@login_required
def graficos_limites_atterberg(request):
    id_proyectos = request.GET.getlist('id_proyectos')
    tipos_prospeccion = request.GET.getlist('tipos_prospeccion')
    areas = request.GET.getlist('areas')
    id_prospecciones = request.GET.getlist('id_prospecciones')

    query = Limites_atterberg.objects.all()
    if id_proyectos:
        query = query.filter(id_proyecto__in=id_proyectos)
    if tipos_prospeccion:
        query = query.filter(tipo_prospeccion__in=tipos_prospeccion)
    if areas:
        query = query.filter(area__in=areas)
    if id_prospecciones:
        query = query.filter(id_prospeccion__in=id_prospecciones)

    proyectos = query.values('id_proyecto').distinct()
    tipos_prospeccion_inicial = query.values_list('tipo_prospeccion', flat=True).distinct().exclude(tipo_prospeccion__isnull=True)
    areas_inicial = query.values_list('area', flat=True).distinct().exclude(area__isnull=True).exclude(area='')
    id_prospecciones_inicial = query.values_list('id_prospeccion', flat=True).distinct().exclude(id_prospeccion__isnull=True)

    df = pd.DataFrame(list(query.values('limite_liquido', 'limite_plastico', 'indice_plasticidad', 'area', 'id_muestra', 'id_prospeccion')))

    context = {
        'proyectos': proyectos,
        'tipos_prospeccion_inicial': tipos_prospeccion_inicial,
        'areas_inicial': areas_inicial,
        'id_prospecciones_inicial': id_prospecciones_inicial,
        'selected_id_proyectos': json.dumps(id_proyectos),
        'selected_tipos_prospeccion': json.dumps(tipos_prospeccion),
        'selected_areas': json.dumps(areas),
        'selected_id_prospecciones': json.dumps(id_prospecciones),
    }

    if df.empty:
        context['error'] = "No hay datos disponibles para los filtros seleccionados."
    else:
        context['image_base64_area'] = generar_grafico_carta_plasticidad_area(df)
        context['image_base64_prospeccion'] = generar_grafico_carta_plasticidad_prospeccion(df)

    return render(request, 'laboratorio/ensayos/limites_atterberg/graficos_limites_atterberg.html', context)

# 9 Funciones Auxiliares (sin cambios, se mantienen igual)
def obtener_tipos_prospeccion_limites_atterberg(request):
    id_proyectos = request.GET.get('id_proyectos')
    query = Limites_atterberg.objects.all()
    if id_proyectos:
        query = query.filter(id_proyecto__in=id_proyectos.split(','))
    tipos_list = list(query.values_list('tipo_prospeccion', flat=True).distinct().exclude(tipo_prospeccion__isnull=True))
    return JsonResponse({'options': tipos_list}, safe=False)

def obtener_id_prospecciones_limites_atterberg(request):
    id_proyectos = request.GET.get('id_proyectos')
    tipos_prospeccion = request.GET.get('tipos_prospeccion')
    areas = request.GET.get('areas')
    query = Limites_atterberg.objects.all()
    if id_proyectos:
        query = query.filter(id_proyecto__in=id_proyectos.split(','))
    if tipos_prospeccion:
        query = query.filter(tipo_prospeccion__in=tipos_prospeccion.split(','))
    if areas:
        query = query.filter(area__in=areas.split(','))
    id_prospecciones_list = list(query.values_list('id_prospeccion', flat=True).distinct().exclude(id_prospeccion__isnull=True))
    return JsonResponse({'options': id_prospecciones_list}, safe=False)

def obtener_area_limites_atterberg(request):
    id_prospecciones = request.GET.get('id_prospecciones')
    if id_prospecciones:
        query = Limites_atterberg.objects.filter(id_prospeccion__in=id_prospecciones.split(','))
        areas_list = list(query.values_list('area', flat=True).distinct().exclude(area__isnull=True).exclude(area=''))
        return JsonResponse({'options': areas_list})
    return JsonResponse({'options': []})

###USCS##################################################


logger = logging.getLogger(__name__)

@login_required
def agregar_uscs(request):
    proyectos_con_muestreo = Proyectos.objects.filter(muestreo__isnull=False).distinct()
    tipo_prospeccion_choices = []
    prospecciones = []
    muestras = []
    area = ""
    uscs_choices = uscs._meta.get_field('uscs').choices

    if request.method == 'POST':
        logger.info(f"Datos POST recibidos: {request.POST}")
        proyecto_id = request.POST.get('id_proyecto')
        tipo_prospeccion = request.POST.get('tipo_prospeccion')
        id_prospecciones = request.POST.getlist('id_prospeccion[]')
        id_muestras = request.POST.getlist('id_muestra[]')
        uscs_values = request.POST.getlist('uscs[]')
        areas = request.POST.getlist('area[]')
        errors = []

        if not proyecto_id:
            errors.append("Debe seleccionar un proyecto.")
            return render(request, 'laboratorio/ensayos/uscs/agregar_uscs.html', {
                'proyectos': proyectos_con_muestreo,
                'uscs_choices': uscs_choices,
                'errors': errors
            })

        muestreos = Muestreo.objects.filter(id_proyecto_id=proyecto_id)
        if not muestreos.exists():
            errors.append("No hay muestreos asociados a este proyecto.")
            return render(request, 'laboratorio/ensayos/uscs/agregar_uscs.html', {
                'proyectos': proyectos_con_muestreo,
                'uscs_choices': uscs_choices,
                'errors': errors
            })

        if tipo_prospeccion:
            muestreos = muestreos.filter(tipo_prospeccion=tipo_prospeccion)
            if not muestreos.exists():
                errors.append("No hay muestreos con este tipo de prospección.")

        if id_prospecciones and id_prospecciones[0]:
            muestreos = muestreos.filter(id_prospeccion__id_prospeccion__in=id_prospecciones)
            if not muestreos.exists():
                errors.append("No hay muestreos con las prospecciones seleccionadas.")

        if id_muestras and id_muestras[0]:
            muestreos = muestreos.filter(id_muestra__in=id_muestras)
            if not muestreos.exists():
                errors.append("No hay muestreos con las muestras seleccionadas.")

        if not (len(id_prospecciones) == len(id_muestras) == len(uscs_values) == len(areas)):
            errors.append("El número de prospecciones, muestras, clasificaciones USCS y áreas no coincide.")
            logger.error(f"Longitudes inconsistentes: prospecciones={len(id_prospecciones)}, muestras={len(id_muestras)}, uscs={len(uscs_values)}, areas={len(areas)}")
        else:
            for id_pros, id_muestra, uscs_val, area_val in zip(id_prospecciones, id_muestras, uscs_values, areas):
                if not id_pros or not id_muestra or not uscs_val or not area_val:
                    errors.append(f"Falta ID de prospección ({id_pros}), ID de muestra ({id_muestra}), clasificación USCS ({uscs_val}) o área ({area_val}).")
                    continue
                
                try:
                    muestreo = muestreos.get(id_prospeccion__id_prospeccion=id_pros, id_muestra=id_muestra)
                    profundidad_desde = float(muestreo.profundidad_desde or 0)  # Si es None, usa 0
                    profundidad_hasta = float(muestreo.profundidad_hasta or 0)  # Si es None, usa 0
                    profundidad_promedio = (profundidad_desde + profundidad_hasta) / 2

                    uscs_obj = uscs(
                        id_proyecto=muestreo.id_proyecto,
                        tipo_prospeccion=muestreo.tipo_prospeccion,
                        id_prospeccion=muestreo.id_prospeccion,
                        id_muestra=muestreo.id_muestra,
                        profundidad_desde=muestreo.profundidad_desde,
                        profundidad_hasta=muestreo.profundidad_hasta,
                        profundidad_promedio=profundidad_promedio,
                        uscs=uscs_val,
                        area=area_val,
                        user=request.user
                    )
                    uscs_obj.save()
                    logger.info(f"USCS guardado: ID={uscs_obj.id}, Muestra={id_muestra}, USCS={uscs_val}, Profundidad Promedio={profundidad_promedio}")
                except Muestreo.DoesNotExist:
                    errors.append(f"Muestra {id_muestra} con prospección {id_pros} no encontrada.")
                except ValueError as e:
                    errors.append(f"Valor inválido para USCS {uscs_val}: {e}")
                except Exception as e:
                    errors.append(f"Error al guardar muestra {id_muestra}: {e}")

        if errors:
            tipo_prospeccion_choices = Muestreo.objects.filter(id_proyecto_id=proyecto_id).values_list('tipo_prospeccion', flat=True).distinct()
            prospecciones = Muestreo.objects.filter(id_proyecto_id=proyecto_id).values_list('id_prospeccion__id_prospeccion', flat=True).distinct()
            muestras = Muestreo.objects.filter(id_proyecto_id=proyecto_id).values_list('id_muestra', flat=True).distinct()
            return render(request, 'laboratorio/ensayos/uscs/agregar_uscs.html', {
                'proyectos': proyectos_con_muestreo,
                'tipo_prospeccion_choices': tipo_prospeccion_choices,
                'prospecciones': prospecciones,
                'muestras': muestras,
                'area': area,
                'uscs_choices': uscs_choices,
                'errors': errors
            })

        return redirect('listar_uscs')

    proyecto_id = request.GET.get('id_proyecto')
    tipo_prospeccion = request.GET.get('tipo_prospeccion')
    id_prospeccion = request.GET.get('id_prospeccion')
    id_muestra = request.GET.get('id_muestra')

    muestreos = Muestreo.objects.all()
    if proyecto_id:
        muestreos = muestreos.filter(id_proyecto_id=proyecto_id)
        tipo_prospeccion_choices = list(muestreos.values_list('tipo_prospeccion', flat=True).distinct().exclude(tipo_prospeccion__isnull=True).exclude(tipo_prospeccion=''))
        prospecciones = list(muestreos.values_list('id_prospeccion__id_prospeccion', flat=True).distinct().exclude(id_prospeccion__id_prospeccion__isnull=True))
        muestras = list(muestreos.values_list('id_muestra', flat=True).distinct().exclude(id_muestra__isnull=True).exclude(id_muestra=''))
        area = ""

    if tipo_prospeccion:
        muestreos = muestreos.filter(tipo_prospeccion=tipo_prospeccion)
        prospecciones = list(muestreos.values_list('id_prospeccion__id_prospeccion', flat=True).distinct().exclude(id_prospeccion__id_prospeccion__isnull=True))
        muestras = list(muestreos.values_list('id_muestra', flat=True).distinct().exclude(id_muestra__isnull=True).exclude(id_muestra=''))

    if id_prospeccion:
        muestreos = muestreos.filter(id_prospeccion__id_prospeccion=id_prospeccion)
        muestras = list(muestreos.values_list('id_muestra', flat=True).distinct().exclude(id_muestra__isnull=True).exclude(id_muestra=''))
        area = muestreos.values_list('area', flat=True).first()

    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if is_ajax:
        data = {
            'tipo_prospeccion_choices': tipo_prospeccion_choices,
            'prospecciones': prospecciones,
            'muestras': muestras,
            'area': area,
        }
        if id_muestra:
            muestreo = muestreos.filter(id_muestra=id_muestra).first()
            if muestreo:
                data['profundidad_desde'] = str(muestreo.profundidad_desde) if muestreo.profundidad_desde is not None else '0'
                data['profundidad_hasta'] = str(muestreo.profundidad_hasta) if muestreo.profundidad_hasta is not None else '0'
                data['area'] = muestreo.area if muestreo.area else ''
        return JsonResponse(data)

    return render(request, 'laboratorio/ensayos/uscs/agregar_uscs.html', {
        'proyectos': proyectos_con_muestreo,
        'tipo_prospeccion_choices': tipo_prospeccion_choices,
        'prospecciones': prospecciones,
        'muestras': muestras,
        'area': area,
        'uscs_choices': uscs_choices
    })

@login_required
def listar_uscs(request):
    uscs_list = uscs.objects.all().order_by('id')
    query = request.GET.get('q', '')

    if query:
        uscs_model = apps.get_model('administrador', 'uscs')
        campos = [field.name for field in uscs_model._meta.get_fields() if isinstance(field, (CharField, TextField, DateField, DateTimeField, DecimalField, FloatField, ForeignKey))]
        query_filter = Q()

        try:
            query_numeric = float(query)
            numeric_search = True
        except ValueError:
            numeric_search = False

        for field in campos:
            field_instance = uscs_model._meta.get_field(field)
            if isinstance(field_instance, (DateField, DateTimeField)):
                query_filter |= Q(**{f"{field}__year__icontains": query})
                query_filter |= Q(**{f"{field}__month__icontains": query})
                query_filter |= Q(**{f"{field}__day__icontains": query})
            elif isinstance(field_instance, ForeignKey):
                related_model = field_instance.related_model
                related_fields = [f"{field}__{related_field.name}" for related_field in related_model._meta.get_fields() if isinstance(related_field, (CharField, TextField))]
                for related_field in related_fields:
                    query_filter |= Q(**{f"{related_field}__icontains": query})
            elif isinstance(field_instance, (DecimalField, FloatField)) and numeric_search:
                query_filter |= Q(**{f"{field}": query_numeric})
            else:
                query_filter |= Q(**{f"{field}__icontains": query})

        uscs_list = uscs_list.filter(query_filter)

    paginator = Paginator(uscs_list, 1000)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'laboratorio/ensayos/uscs/uscs_table.html', {'page_obj': page_obj})

    return render(request, 'laboratorio/ensayos/uscs/listar_uscs.html', {
        'page_obj': page_obj,
        'query': query
    })

@login_required
def editar_uscs(request, id):
    uscs_obj = get_object_or_404(uscs, id=id)
    muestreo = Muestreo.objects.filter(id_muestra=uscs_obj.id_muestra).first()
    proyectos = Proyectos.objects.filter(muestreo__isnull=False).distinct()

    if request.method == 'POST':
        form = UscsForm(request.POST, instance=uscs_obj)  # Asumimos que existe un UscsForm
        if form.is_valid():
            try:
                uscs_obj = form.save()
                logger.info(f"USCS editado: ID {uscs_obj.id}")
                return redirect('listar_uscs')
            except Exception as e:
                logger.error(f"Error al editar: {e}")
                form.add_error(None, f"Error al editar: {e}")
    else:
        form = UscsForm(instance=uscs_obj)
    
    return render(request, 'laboratorio/ensayos/uscs/editar_uscs.html', {
        'form': form,
        'uscs_obj': uscs_obj,
        'proyectos': proyectos
    })

@login_required
def ver_uscs(request, id):
    uscs_obj = get_object_or_404(uscs, id=id)
    
    # Preparar el historial con los cambios
    history_records = []
    for record in uscs_obj.history.all():
        changes = []
        previous = record.prev_record
        if previous:
            # Comparar cada campo entre el registro actual y el anterior
            for field in record.history_object._meta.fields:
                old_value = getattr(previous, field.name, None)
                new_value = getattr(record, field.name, None)
                if old_value != new_value:
                    changes.append({
                        'field_name': field.verbose_name,
                        'old_value': old_value if old_value is not None else '-',
                        'new_value': new_value if new_value is not None else '-',
                    })
        history_records.append({
            'date': record.history_date,
            'user': record.history_user,
            'type': record.history_type,
            'reason': record.history_change_reason,
            'changes': changes if changes else None,
        })

    return render(request, 'laboratorio/ensayos/uscs/ver_uscs.html', {
        'uscs': uscs_obj,
        'history_records': history_records,
    })

@login_required
def eliminar_uscs(request, id):
    uscs_obj = get_object_or_404(uscs, id=id)
    if request.method == 'POST':
        uscs_obj.delete()
        return redirect('listar_uscs')
    return render(request, 'laboratorio/ensayos/uscs/eliminar_uscs.html', {
        'uscs': uscs_obj
    })

@login_required
def export_to_excel_uscs(request):
    query = request.GET.get('q', '').strip()
    headers = ['ID', 'ID Proyecto', 'ID Prospección', 'Tipo Prospección', 'ID Muestra', 
               'Profundidad Desde', 'Profundidad Hasta', 'Profundidad Promedio', 
               'USCS', 'Área', 'Usuario']

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="uscs_filtrado.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "USCS"
    ws.append(headers)

    try:
        uscs_list = uscs.objects.select_related('id_proyecto', 'id_prospeccion', 'user').prefetch_related('id_muestra__muestreo_set').all()
        logger.debug(f"Registros iniciales: {uscs_list.count()}, query: '{query}'")

        if query:
            filtros = Q()
            try:
                query_numeric = float(query)
                filtros |= (
                    Q(profundidad_desde=query_numeric) |
                    Q(profundidad_hasta=query_numeric) |
                    Q(profundidad_promedio=query_numeric)
                )
            except ValueError:
                pass

            filtros |= (
                Q(id__icontains=query) |
                Q(id_proyecto__id__icontains=query) |
                Q(id_prospeccion__id_prospeccion__icontains=query) |
                Q(tipo_prospeccion__icontains=query) |
                Q(id_muestra__icontains=query) |
                Q(uscs__icontains=query) |
                Q(user__username__icontains=query)
            )

            uscs_list = uscs_list.filter(filtros)
        
        logger.debug(f"Registros después del filtro: {uscs_list.count()}")

        for uscs_obj in uscs_list:
            try:
                muestreo = Muestreo.objects.filter(id_muestra=uscs_obj.id_muestra).first()
                row = [
                    str(uscs_obj.id) if uscs_obj.id is not None else '',
                    str(uscs_obj.id_proyecto.id) if uscs_obj.id_proyecto else '',
                    str(uscs_obj.id_prospeccion.id_prospeccion) if uscs_obj.id_prospeccion else '',
                    str(uscs_obj.tipo_prospeccion or ''),
                    str(uscs_obj.id_muestra or ''),
                    str(uscs_obj.profundidad_desde) if uscs_obj.profundidad_desde is not None else '',
                    str(uscs_obj.profundidad_hasta) if uscs_obj.profundidad_hasta is not None else '',
                    str(uscs_obj.profundidad_promedio) if uscs_obj.profundidad_promedio is not None else '',
                    str(uscs_obj.uscs) if uscs_obj.uscs is not None else '',
                    str(muestreo.area) if muestreo and hasattr(muestreo, 'area') else '',
                    str(uscs_obj.user.username) if uscs_obj.user else 'Sin usuario'
                ]
                ws.append(row)
            except Exception as e:
                logger.error(f"Error al procesar registro {uscs_obj.id}: {str(e)}")
                continue

        wb.save(response)
        return response

    except Exception as e:
        logger.error(f"Error en export_to_excel_uscs: {str(e)}")
        ws.append(["Error al generar el reporte", str(e)])
        wb.save(response)
        return response

@login_required
def export_to_pdf_uscs(request):
    query = request.GET.get('q', '').strip()
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="uscs_filtrado.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(letter), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()
    title = Paragraph("Lista de USCS", styles['Heading1'].clone('Title', alignment=1, fontSize=12, spaceAfter=10))
    elements.append(title)
    elements.append(Paragraph("<br/>", styles['Normal']))

    headers = ['ID', 'Proyecto', 'Prospección', 'Tipo', 'Muestra', 'Prof. Ini', 'Prof. Fin', 'Prof. Prom', 'USCS', 'Área', 'Usuario']
    data = [headers]

    try:
        uscs_list = uscs.objects.select_related('id_proyecto', 'id_prospeccion', 'user').prefetch_related('id_muestra__muestreo_set').all()
        logger.debug(f"Registros iniciales: {uscs_list.count()}, query: '{query}'")

        if query:
            filtros = Q()
            try:
                query_numeric = float(query)
                filtros |= (
                    Q(profundidad_desde=query_numeric) |
                    Q(profundidad_hasta=query_numeric) |
                    Q(profundidad_promedio=query_numeric)
                )
            except ValueError:
                pass

            filtros |= (
                Q(id__icontains=query) |
                Q(id_proyecto__id__icontains=query) |
                Q(id_prospeccion__id_prospeccion__icontains=query) |
                Q(tipo_prospeccion__icontains=query) |
                Q(id_muestra__icontains=query) |
                Q(uscs__icontains=query) |
                Q(user__username__icontains=query)
            )

            uscs_list = uscs_list.filter(filtros)
        
        logger.debug(f"Registros después del filtro: {uscs_list.count()}")

        for uscs_obj in uscs_list:
            try:
                muestreo = Muestreo.objects.filter(id_muestra=uscs_obj.id_muestra).first()
                row = [
                    str(uscs_obj.id) if uscs_obj.id is not None else '',
                    str(uscs_obj.id_proyecto.id) if uscs_obj.id_proyecto else '',
                    str(uscs_obj.id_prospeccion.id_prospeccion) if uscs_obj.id_prospeccion else '',
                    str(uscs_obj.tipo_prospeccion or ''),
                    str(uscs_obj.id_muestra or ''),
                    str(uscs_obj.profundidad_desde) if uscs_obj.profundidad_desde is not None else '',
                    str(uscs_obj.profundidad_hasta) if uscs_obj.profundidad_hasta is not None else '',
                    str(uscs_obj.profundidad_promedio) if uscs_obj.profundidad_promedio is not None else '',
                    str(uscs_obj.uscs) if uscs_obj.uscs is not None else '',
                    str(muestreo.area) if muestreo and hasattr(muestreo, 'area') else '',
                    str(uscs_obj.user.username) if uscs_obj.user else 'Sin usuario'
                ]
                data.append(row)
            except Exception as e:
                logger.error(f"Error al procesar registro {uscs_obj.id}: {str(e)}")
                continue

        table = Table(data, colWidths=[0.5*inch] + [0.8*inch]*10)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        elements.append(table)
        doc.build(elements)
        return response

    except Exception as e:
        logger.error(f"Error en export_to_pdf_uscs: {str(e)}")
        doc.build([Paragraph(f"Error al generar el reporte: {str(e)}", styles['Normal'])])
        return response

# Gráficos USCS (mismo formato que gravedad_especifica: dispersión)

logger = logging.getLogger(__name__)

# Vista para gráficos USCS
def graficos_uscs(request):
    id_proyectos = request.GET.getlist('id_proyecto')
    tipos_prospeccion = request.GET.getlist('tipo_prospeccion')
    areas = request.GET.getlist('area')
    id_prospecciones = request.GET.getlist('id_prospeccion')
    
    query = uscs.objects.all()
    if id_proyectos:
        query = query.filter(id_proyecto__in=id_proyectos)
    if tipos_prospeccion:
        query = query.filter(tipo_prospeccion__in=tipos_prospeccion)
    if areas:
        query = query.filter(area__in=areas)
    if id_prospecciones:
        query = query.filter(id_prospeccion__in=id_prospecciones)
    
    proyectos = query.values('id_proyecto').distinct()
    tipos_prospeccion_inicial = (query.values_list('tipo_prospeccion', flat=True)
                                 .distinct().exclude(tipo_prospeccion__isnull=True).exclude(tipo_prospeccion=''))
    areas_inicial = (query.values_list('area', flat=True)
                     .distinct().exclude(area__isnull=True).exclude(area=''))
    id_prospecciones_inicial = (query.values_list('id_prospeccion', flat=True)
                                .distinct().exclude(id_prospeccion__isnull=True).exclude(id_prospeccion=''))
    
    df = pd.DataFrame.from_records(query.values('id_proyecto', 'id_prospeccion', 'id_muestra', 'uscs', 'area'))
    
    context = {
        'proyectos': proyectos,
        'tipos_prospeccion_inicial': tipos_prospeccion_inicial,
        'areas_inicial': areas_inicial,
        'id_prospecciones_inicial': id_prospecciones_inicial,
        'selected_id_proyectos': json.dumps(id_proyectos),
        'selected_tipos_prospeccion': json.dumps(tipos_prospeccion),
        'selected_areas': json.dumps(areas),
        'selected_id_prospecciones': json.dumps(id_prospecciones),
    }
    
    if df.empty:
        context['error'] = "No hay datos para graficar."
        return render(request, 'laboratorio/ensayos/uscs/graficos_uscs.html', context)
    
    # Gráfico 1: Distribución porcentual de clasificaciones USCS (anillo)
    def generar_grafico_distribucion_uscs(df):
        uscs_counts = df['uscs'].value_counts(normalize=True) * 100
        labels = uscs_counts.index
        sizes = uscs_counts.values
        colors = [np.random.rand(3,) for _ in range(len(labels))]  # Colores aleatorios como en humedad
        
        fig, ax = plt.subplots()
        wedges, texts, autotexts = ax.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%', 
                                          startangle=90, wedgeprops=dict(width=0.3, edgecolor='white'))
        ax.grid(True)  # Grilla como en humedad
        plt.subplots_adjust(right=0.75)  # Márgenes como en humedad
        num_columns = (len(labels) + 24) // 25  # Cálculo de columnas como en humedad
        legend = ax.legend(loc='center left', bbox_to_anchor=(1.05, 0.5), ncol=num_columns)
        ax.set_title("Distribución Porcentual de Clasificaciones USCS")
        buffer = BytesIO()
        plt.savefig(buffer, format='png', dpi=100, bbox_inches='tight', bbox_extra_artists=[legend])
        buffer.seek(0)
        plt.close()
        return base64.b64encode(buffer.getvalue()).decode('utf-8')
    
    # Gráfico 2: Clasificaciones por área (barras apiladas)
    def generar_grafico_uscs_por_area(df):
        pivot = df.pivot_table(index='area', columns='uscs', aggfunc='size', fill_value=0)
        colors = [np.random.rand(3,) for _ in range(len(pivot.columns))]  # Colores aleatorios como en humedad
        
        fig, ax = plt.subplots()
        pivot.plot(kind='bar', stacked=True, ax=ax, color=colors, edgecolor='white', linewidth=0.5)
        ax.set_xlabel("Área")
        ax.set_ylabel("Cantidad")
        ax.grid(True)  # Grilla como en humedad
        plt.subplots_adjust(right=0.75)  # Márgenes como en humedad
        num_columns = (len(pivot.columns) + 24) // 25  # Cálculo de columnas como en humedad
        legend = ax.legend(loc='center left', bbox_to_anchor=(1, 0.5), ncol=num_columns)
        ax.set_title("Clasificaciones USCS por Área")
        plt.xticks(rotation=45, ha='right')
        buffer = BytesIO()
        plt.savefig(buffer, format='png', dpi=85, bbox_inches='tight', bbox_extra_artists=[legend])
        buffer.seek(0)
        plt.close()
        return base64.b64encode(buffer.getvalue()).decode('utf-8')
    
    image_base64_distribucion = generar_grafico_distribucion_uscs(df)
    image_base64_area = generar_grafico_uscs_por_area(df)
    
    context.update({
        'image_base64_distribucion': image_base64_distribucion,
        'image_base64_area': image_base64_area,
    })
    return render(request, 'laboratorio/ensayos/uscs/graficos_uscs.html', context)

# Obtener tipos de prospección desde Muestreo
def obtener_tipos_prospeccion_muestreo(request):
    id_proyecto = request.GET.get('id_proyecto')
    tipos = Muestreo.objects.all()
    if id_proyecto:
        tipos = tipos.filter(id_proyecto=id_proyecto)
    tipos_list = list(tipos.values_list('tipo_prospeccion', flat=True)
                     .distinct()
                     .exclude(tipo_prospeccion__isnull=True)
                     .exclude(tipo_prospeccion=''))
    return JsonResponse({'options': tipos_list}, safe=False)

# Obtener IDs de prospección desde Muestreo
def obtener_id_prospecciones_muestreo(request):
    id_proyecto = request.GET.get('id_proyecto')
    tipo_prospeccion = request.GET.get('tipo_prospeccion')
    query = Muestreo.objects.all()
    if id_proyecto:
        query = query.filter(id_proyecto=id_proyecto)
    if tipo_prospeccion:
        query = query.filter(tipo_prospeccion=tipo_prospeccion)
    id_prospecciones_list = list(query.values_list('id_prospeccion__id_prospeccion', flat=True)
                                 .distinct()
                                 .exclude(id_prospeccion__isnull=True))
    return JsonResponse({'options': id_prospecciones_list}, safe=False)

# Obtener área desde Muestreo
def obtener_area_muestreo(request):
    prospeccion_id = request.GET.get('prospeccion_id')
    if prospeccion_id:
        try:
            muestreo = Muestreo.objects.filter(id_prospeccion__id_prospeccion=prospeccion_id).first()
            if muestreo:
                return JsonResponse({'area': muestreo.area})
        except Muestreo.DoesNotExist:
            pass
    return JsonResponse({'area': ''})

# Obtener IDs de muestra desde Muestreo
def obtener_id_muestras_muestreo(request):
    id_prospeccion = request.GET.get('id_prospeccion')
    muestras = Muestreo.objects.all()
    if id_prospeccion:
        muestras = muestras.filter(id_prospeccion__id_prospeccion=id_prospeccion)
    muestras_list = list(muestras.values_list('id_muestra', flat=True)
                        .distinct()
                        .exclude(id_muestra__isnull=True)
                        .exclude(id_muestra=''))
    return JsonResponse({'options': muestras_list}, safe=False)

def obtener_profundidades_muestreo(request):
    id_muestra = request.GET.get('id_muestra')
    if (id_muestra):
        try:
            muestreo = Muestreo.objects.get(id_muestra=id_muestra)
            return JsonResponse({
                'profundidad_desde': str(muestreo.profundidad_desde) if muestreo.profundidad_desde else '',
                'profundidad_hasta': str(muestreo.profundidad_hasta) if muestreo.profundidad_hasta else ''
            })
        except Muestreo.DoesNotExist:
            return JsonResponse({'profundidad_desde': '', 'profundidad_hasta': ''})
    return JsonResponse({'profundidad_desde': '', 'profundidad_hasta': ''})

# Funciones auxiliares (iguales que en Código 1, renombradas para uscs)
def obtener_tipos_prospeccion_uscs(request):
    id_proyectos = request.GET.get('id_proyectos')
    query = uscs.objects.all()
    if id_proyectos:
        query = query.filter(id_proyecto__in=id_proyectos.split(','))
    tipos_list = list(query.values_list('tipo_prospeccion', flat=True).distinct().exclude(tipo_prospeccion__isnull=True))
    return JsonResponse({'options': tipos_list}, safe=False)

def obtener_id_prospecciones_uscs(request):
    id_proyectos = request.GET.get('id_proyectos')
    tipos_prospeccion = request.GET.get('tipos_prospeccion')
    areas = request.GET.get('areas')
    query = uscs.objects.all()
    if id_proyectos:
        query = query.filter(id_proyecto__in=id_proyectos.split(','))
    if tipos_prospeccion:
        query = query.filter(tipo_prospeccion__in=tipos_prospeccion.split(','))
    if areas:
        query = query.filter(area__in=areas.split(','))
    id_prospecciones_list = list(query.values_list('id_prospeccion', flat=True).distinct().exclude(id_prospeccion__isnull=True))
    return JsonResponse({'options': id_prospecciones_list}, safe=False)

def obtener_area_uscs(request):
    id_prospecciones = request.GET.get('id_prospecciones')
    if id_prospecciones:
        query = uscs.objects.filter(id_prospeccion__in=id_prospecciones.split(','))
        areas_list = list(query.values_list('area', flat=True).distinct().exclude(area__isnull=True).exclude(area=''))
        return JsonResponse({'options': areas_list})
    return JsonResponse({'options': []})



#### HUMEDAD ##############################################
import logging
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator
from django.db.models import Q
from .models import Proyectos, Muestreo, Humedad
from .forms import HumedadForm
from openpyxl import Workbook
from reportlab.lib.pagesizes import landscape, letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from io import BytesIO
import base64
import json

logger = logging.getLogger(__name__)

# Agregar Humedad
@login_required
def agregar_humedad(request):
    proyectos_con_muestreo = Proyectos.objects.filter(muestreo__isnull=False).distinct()
    tipo_prospeccion_choices = []
    prospecciones = []
    muestras = []
    area = ""

    if request.method == 'POST':
        logger.info(f"Datos POST recibidos: {request.POST}")
        proyecto_id = request.POST.get('id_proyecto')
        tipo_prospeccion = request.POST.get('tipo_prospeccion')
        id_prospecciones = request.POST.getlist('id_prospeccion[]')
        id_muestras = request.POST.getlist('id_muestra[]')
        humedades = request.POST.getlist('humedad[]')
        areas = request.POST.getlist('area[]')
        errors = []

        if not proyecto_id:
            errors.append("Debe seleccionar un proyecto.")
            return render(request, 'laboratorio/ensayos/humedad/agregar_humedad.html', {
                'proyectos': proyectos_con_muestreo,
                'errors': errors
            })

        muestreos = Muestreo.objects.filter(id_proyecto_id=proyecto_id)
        if not muestreos.exists():
            errors.append("No hay muestreos asociados a este proyecto.")
            return render(request, 'laboratorio/ensayos/humedad/agregar_humedad.html', {
                'proyectos': proyectos_con_muestreo,
                'errors': errors
            })

        if tipo_prospeccion:
            muestreos = muestreos.filter(tipo_prospeccion=tipo_prospeccion)
            if not muestreos.exists():
                errors.append("No hay muestreos con este tipo de prospección.")

        if id_prospecciones and id_prospecciones[0]:
            muestreos = muestreos.filter(id_prospeccion__id_prospeccion__in=id_prospecciones)
            if not muestreos.exists():
                errors.append("No hay muestreos con las prospecciones seleccionadas.")

        if id_muestras and id_muestras[0]:
            muestreos = muestreos.filter(id_muestra__in=id_muestras)
            if not muestreos.exists():
                errors.append("No hay muestreos con las muestras seleccionadas.")

        if not (len(id_prospecciones) == len(id_muestras) == len(humedades) == len(areas)):
            errors.append("El número de prospecciones, muestras, humedades y áreas no coincide.")
            logger.error(f"Longitudes inconsistentes: prospecciones={len(id_prospecciones)}, muestras={len(id_muestras)}, humedades={len(humedades)}, areas={len(areas)}")
        else:
            for id_pros, id_muestra, humedad, area_val in zip(id_prospecciones, id_muestras, humedades, areas):
                if not id_pros or not id_muestra or not humedad or not area_val:
                    errors.append(f"Falta ID de prospección ({id_pros}), ID de muestra ({id_muestra}), humedad ({humedad}) o área ({area_val}).")
                    continue
                
                try:
                    muestreo = muestreos.get(id_prospeccion__id_prospeccion=id_pros, id_muestra=id_muestra)
                    profundidad_desde = float(muestreo.profundidad_desde or 0)  # Si es None, usa 0
                    profundidad_hasta = float(muestreo.profundidad_hasta or 0)  # Si es None, usa 0
                    profundidad_promedio = (profundidad_desde + profundidad_hasta) / 2

                    humedad_obj = Humedad(
                        id_proyecto=muestreo.id_proyecto,
                        tipo_prospeccion=muestreo.tipo_prospeccion,
                        id_prospeccion=muestreo.id_prospeccion,
                        id_muestra=muestreo.id_muestra,
                        profundidad_desde=muestreo.profundidad_desde,
                        profundidad_hasta=muestreo.profundidad_hasta,
                        profundidad_promedio=profundidad_promedio,
                        humedad=float(humedad),
                        area=area_val,
                        user=request.user
                    )
                    humedad_obj.save()
                    logger.info(f"Humedad guardada: ID={humedad_obj.id}, Muestra={id_muestra}, Humedad={humedad}, Profundidad Promedio={profundidad_promedio}")
                except Muestreo.DoesNotExist:
                    errors.append(f"Muestra {id_muestra} con prospección {id_pros} no encontrada.")
                except ValueError as e:
                    errors.append(f"Valor inválido para humedad {humedad}: {e}")
                except Exception as e:
                    errors.append(f"Error al guardar muestra {id_muestra}: {e}")

        if errors:
            tipo_prospeccion_choices = Muestreo.objects.filter(id_proyecto_id=proyecto_id).values_list('tipo_prospeccion', flat=True).distinct()
            prospecciones = Muestreo.objects.filter(id_proyecto_id=proyecto_id).values_list('id_prospeccion__id_prospeccion', flat=True).distinct()
            muestras = Muestreo.objects.filter(id_proyecto_id=proyecto_id).values_list('id_muestra', flat=True).distinct()
            return render(request, 'laboratorio/ensayos/humedad/agregar_humedad.html', {
                'proyectos': proyectos_con_muestreo,
                'tipo_prospeccion_choices': tipo_prospeccion_choices,
                'prospecciones': prospecciones,
                'muestras': muestras,
                'area': area,
                'errors': errors
            })

        return redirect('listar_humedad')

    proyecto_id = request.GET.get('id_proyecto')
    tipo_prospeccion = request.GET.get('tipo_prospeccion')
    id_prospeccion = request.GET.get('id_prospeccion')
    id_muestra = request.GET.get('id_muestra')

    muestreos = Muestreo.objects.all()
    if proyecto_id:
        muestreos = muestreos.filter(id_proyecto_id=proyecto_id)
        tipo_prospeccion_choices = muestreos.values_list('tipo_prospeccion', flat=True).distinct()
        prospecciones = muestreos.values_list('id_prospeccion__id_prospeccion', flat=True).distinct()
        muestras = muestreos.values_list('id_muestra', flat=True).distinct()
        area = muestreos.first().area if muestreos.exists() else ""

    if tipo_prospeccion:
        muestreos = muestreos.filter(tipo_prospeccion=tipo_prospeccion)
        prospecciones = muestreos.values_list('id_prospeccion__id_prospeccion', flat=True).distinct()
        muestras = muestreos.values_list('id_muestra', flat=True).distinct()

    if id_prospeccion:
        muestreos = muestreos.filter(id_prospeccion__id_prospeccion=id_prospeccion)
        muestras = muestreos.values_list('id_muestra', flat=True).distinct()
        area = muestreos.first().area if muestreos.exists() else ""

    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if is_ajax:
        data = {
            'tipo_prospeccion_choices': list(tipo_prospeccion_choices),
            'prospecciones': list(prospecciones),
            'muestras': list(muestras),
            'area': area
        }
        if id_muestra:
            muestreo = muestreos.filter(id_muestra=id_muestra).first()
            if muestreo:
                data['profundidad_desde'] = str(muestreo.profundidad_desde) if muestreo.profundidad_desde is not None else '0'
                data['profundidad_hasta'] = str(muestreo.profundidad_hasta) if muestreo.profundidad_hasta is not None else '0'
        return JsonResponse(data)

    return render(request, 'laboratorio/ensayos/humedad/agregar_humedad.html', {
        'proyectos': proyectos_con_muestreo,
        'tipo_prospeccion_choices': tipo_prospeccion_choices,
        'prospecciones': prospecciones,
        'muestras': muestras,
        'area': area
    })
    
    
# Listar Humedad
@login_required
def listar_humedad(request):
    humedad_list = Humedad.objects.all().order_by('id')
    query = request.GET.get('q', '')

    if query:
        filtros = Q()
        try:
            query_numeric = float(query)
            filtros |= (
                Q(profundidad_desde=query_numeric) |
                Q(profundidad_hasta=query_numeric) |
                Q(profundidad_promedio=query_numeric) |
                Q(humedad=query_numeric)
            )
        except ValueError:
            pass

        filtros |= (
            Q(id_proyecto__id__icontains=query) |
            Q(tipo_prospeccion__icontains=query) |
            Q(id_prospeccion__id_prospeccion__icontains=query) |
            Q(id_muestra__icontains=query) |
            Q(area__icontains=query) |
            Q(user__username__icontains=query)
        )
        humedad_list = humedad_list.filter(filtros)

    paginator = Paginator(humedad_list, 1000)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'laboratorio/ensayos/humedad/humedad_table.html', {'page_obj': page_obj})

    return render(request, 'laboratorio/ensayos/humedad/listar_humedad.html', {
        'page_obj': page_obj,
        'query': query
    })

# Editar Humedad
@login_required
def editar_humedad(request, id):
    humedad_obj = get_object_or_404(Humedad, id=id)
    proyectos = Proyectos.objects.filter(muestreo__isnull=False).distinct()

    if request.method == 'POST':
        form = HumedadForm(request.POST, instance=humedad_obj)
        if form.is_valid():
            try:
                humedad_obj = form.save()
                logger.info(f"Humedad editada: ID {humedad_obj.id}")
                return redirect('listar_humedad')
            except Exception as e:
                logger.error(f"Error al editar: {e}")
                form.add_error(None, f"Error al editar: {e}")
    else:
        form = HumedadForm(instance=humedad_obj)

    return render(request, 'laboratorio/ensayos/humedad/editar_humedad.html', {
        'form': form,
        'humedad_obj': humedad_obj,
        'proyectos': proyectos
    })

# Ver Humedad
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from administrador.models import Humedad

@login_required
def ver_humedad(request, id):
    humedad_obj = get_object_or_404(Humedad, id=id)
    
    # Preparar el historial con los cambios
    history_records = []
    for record in humedad_obj.history.all():
        changes = []
        previous = record.prev_record
        if previous:
            # Comparar cada campo entre el registro actual y el anterior
            for field in record.history_object._meta.fields:
                old_value = getattr(previous, field.name, None)
                new_value = getattr(record, field.name, None)
                if old_value != new_value:
                    changes.append({
                        'field_name': field.verbose_name,
                        'old_value': old_value if old_value is not None else '-',
                        'new_value': new_value if new_value is not None else '-',
                    })
        history_records.append({
            'date': record.history_date,
            'user': record.history_user,
            'type': record.history_type,
            'reason': record.history_change_reason,
            'changes': changes if changes else None,
        })

    return render(request, 'laboratorio/ensayos/humedad/ver_humedad.html', {
        'humedad': humedad_obj,
        'history_records': history_records,
    })

# Eliminar Humedad
@login_required
def eliminar_humedad(request, id):
    humedad_obj = get_object_or_404(Humedad, id=id)
    if request.method == 'POST':
        humedad_obj.delete()
        return redirect('listar_humedad')
    return render(request, 'laboratorio/ensayos/humedad/eliminar_humedad.html', {
        'humedad': humedad_obj
    })

# Exportar a Excel
@login_required
def export_to_excel_humedad(request):
    query = request.GET.get('q', '').strip()
    headers = ['ID', 'ID Proyecto', 'ID Prospección', 'Tipo Prospección', 'ID Muestra', 
               'Profundidad Desde', 'Profundidad Hasta', 'Profundidad Promedio', 
               'Humedad', 'Área', 'Usuario']

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="humedad_filtrado.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Humedad"
    ws.append(headers)

    try:
        humedad_list = Humedad.objects.select_related('id_proyecto', 'id_prospeccion', 'user').all()
        if query:
            filtros = Q()
            try:
                query_numeric = float(query)
                filtros |= (
                    Q(profundidad_desde=query_numeric) |
                    Q(profundidad_hasta=query_numeric) |
                    Q(profundidad_promedio=query_numeric) |
                    Q(humedad=query_numeric)
                )
            except ValueError:
                pass

            filtros |= (
                Q(id__icontains=query) |
                Q(id_proyecto__id__icontains=query) |
                Q(id_prospeccion__id_prospeccion__icontains=query) |
                Q(tipo_prospeccion__icontains=query) |
                Q(id_muestra__icontains=query) |
                Q(area__icontains=query) |
                Q(user__username__icontains=query)
            )
            humedad_list = humedad_list.filter(filtros)

        for humedad in humedad_list:
            row = [
                str(humedad.id),
                str(humedad.id_proyecto.id) if humedad.id_proyecto else '',
                str(humedad.id_prospeccion.id_prospeccion) if humedad.id_prospeccion else '',
                str(humedad.tipo_prospeccion or ''),
                str(humedad.id_muestra or ''),
                str(humedad.profundidad_desde) if humedad.profundidad_desde is not None else '',
                str(humedad.profundidad_hasta) if humedad.profundidad_hasta is not None else '',
                str(humedad.profundidad_promedio) if humedad.profundidad_promedio is not None else '',
                str(humedad.humedad) if humedad.humedad is not None else '',
                str(humedad.area) if humedad.area else '',
                str(humedad.user.username) if humedad.user else 'Sin usuario'
            ]
            ws.append(row)

        wb.save(response)
        return response
    except Exception as e:
        logger.error(f"Error en export_to_excel_humedad: {str(e)}")
        ws.append(["Error al generar el reporte", str(e)])
        wb.save(response)
        return response

# Exportar a PDF
@login_required
def export_to_pdf_humedad(request):
    query = request.GET.get('q', '').strip()
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="humedad_filtrado.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(letter), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()
    title = Paragraph("Lista de Humedad", styles['Heading1'].clone('Title', alignment=1, fontSize=12, spaceAfter=10))
    elements.append(title)
    elements.append(Paragraph("<br/>", styles['Normal']))

    headers = ['ID', 'Proyecto', 'Prospección', 'Tipo', 'Muestra', 'Prof. Ini', 'Prof. Fin', 'Prof. Prom', 'Humedad', 'Área', 'Usuario']
    data = [headers]

    try:
        humedad_list = Humedad.objects.select_related('id_proyecto', 'id_prospeccion', 'user').all()
        if query:
            filtros = Q()
            try:
                query_numeric = float(query)
                filtros |= (
                    Q(profundidad_desde=query_numeric) |
                    Q(profundidad_hasta=query_numeric) |
                    Q(profundidad_promedio=query_numeric) |
                    Q(humedad=query_numeric)
                )
            except ValueError:
                pass

            filtros |= (
                Q(id__icontains=query) |
                Q(id_proyecto__id__icontains=query) |
                Q(id_prospeccion__id_prospeccion__icontains=query) |
                Q(tipo_prospeccion__icontains=query) |
                Q(id_muestra__icontains=query) |
                Q(area__icontains=query) |
                Q(user__username__icontains=query)
            )
            humedad_list = humedad_list.filter(filtros)

        for humedad in humedad_list:
            row = [
                str(humedad.id),
                str(humedad.id_proyecto.id) if humedad.id_proyecto else '',
                str(humedad.id_prospeccion.id_prospeccion) if humedad.id_prospeccion else '',
                str(humedad.tipo_prospeccion or ''),
                str(humedad.id_muestra or ''),
                str(humedad.profundidad_desde) if humedad.profundidad_desde is not None else '',
                str(humedad.profundidad_hasta) if humedad.profundidad_hasta is not None else '',
                str(humedad.profundidad_promedio) if humedad.profundidad_promedio is not None else '',
                str(humedad.humedad) if humedad.humedad is not None else '',
                str(humedad.area) if humedad.area else '',
                str(humedad.user.username) if humedad.user else 'Sin usuario'
            ]
            data.append(row)

        table = Table(data, colWidths=[0.5*inch] + [0.8*inch]*10)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(table)
        doc.build(elements)
        return response
    except Exception as e:
        logger.error(f"Error en export_to_pdf_humedad: {str(e)}")
        doc.build([Paragraph(f"Error al generar el reporte: {str(e)}", styles['Normal'])])
        return response

# Gráficos Humedad
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
from io import BytesIO
import base64
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.http import JsonResponse
from administrador.models import Humedad  # Nombre correcto del modelo
import json
import random

# Funciones de gráficos por área
def generar_grafico_humedad_area(df):
    areas_unicas = df['area'].unique()
    colores = {area: (random.random(), random.random(), random.random()) for area in areas_unicas}
    fig, ax = plt.subplots()
    
    for area in areas_unicas:
        datos_area = df[df['area'] == area]
        ax.scatter(
            datos_area['humedad'],
            datos_area['profundidad_promedio'],
            color=colores[area],
            s=100,
            label=area,
            marker='s'
        )
    
    ax.xaxis.set_ticks_position("top")
    ax.xaxis.set_label_position("top")
    ax.invert_yaxis()
    ax.set_xlabel("Contenido de Humedad (%)")
    ax.set_ylabel("Profundidad (m)")
    ax.grid(True)
    plt.subplots_adjust(right=0.75)
    num_columns = (len(areas_unicas) + 24) // 25
    legend = ax.legend(loc='center left', bbox_to_anchor=(1, 0.5), ncol=num_columns)
    plt.title("GRÁFICO AGRUPADO POR ÁREA (DISPERSIÓN)", pad=30)
    
    buffer = BytesIO()
    fig.savefig(buffer, format='png', dpi=100, bbox_inches='tight', bbox_extra_artists=[legend])
    buffer.seek(0)
    plt.close()
    image_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
    return image_base64

# Funciones de gráficos por prospección
def generar_grafico_humedad_prospeccion(df):
    df['id_prospeccion'] = df['id_prospeccion'].astype(str)
    df['id_prospeccion_unico'] = df.groupby('id_prospeccion').cumcount() + 1
    df['etiqueta'] = df['id_prospeccion'] + '-' + df['id_prospeccion_unico'].astype(str)
    prospecciones_unicas = df['etiqueta'].unique()
    colores = {prospeccion: (random.random(), random.random(), random.random()) for prospeccion in prospecciones_unicas}
    
    fig, ax = plt.subplots()
    
    for prospeccion in prospecciones_unicas:
        datos_prospeccion = df[df['etiqueta'] == prospeccion]
        ax.scatter(
            datos_prospeccion['humedad'],
            datos_prospeccion['profundidad_promedio'],
            color=colores[prospeccion],
            s=100,
            label=prospeccion,
            marker='s'
        )
    
    ax.xaxis.set_ticks_position("top")
    ax.xaxis.set_label_position("top")
    ax.invert_yaxis()
    ax.set_xlabel("Contenido de Humedad (%)")
    ax.set_ylabel("Profundidad (m)")
    ax.grid(True)
    plt.subplots_adjust(right=0.75)
    num_columns = (len(prospecciones_unicas) + 24) // 25
    legend = ax.legend(loc='center left', bbox_to_anchor=(1, 0.5), ncol=num_columns)
    plt.title("GRÁFICO AGRUPADO POR PROSPECCIÓN (DISPERSIÓN)", pad=30)
    
    buffer = BytesIO()
    fig.savefig(buffer, format='png', dpi=100, bbox_inches='tight', bbox_extra_artists=[legend])
    buffer.seek(0)
    plt.close()
    image_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
    return image_base64

# Vista principal para gráficos de Humedad
@login_required
def graficos_humedad(request):
    id_proyectos = request.GET.getlist('id_proyectos')
    tipos_prospeccion = request.GET.getlist('tipos_prospeccion')
    areas = request.GET.getlist('areas')
    id_prospecciones = request.GET.getlist('id_prospecciones')

    query = Humedad.objects.all()  # Nombre correcto del modelo
    if id_proyectos:
        query = query.filter(id_proyecto__in=id_proyectos)
    if tipos_prospeccion:
        query = query.filter(tipo_prospeccion__in=tipos_prospeccion)
    if areas:
        query = query.filter(area__in=areas)
    if id_prospecciones:
        query = query.filter(id_prospeccion__in=id_prospecciones)

    proyectos = query.values('id_proyecto').distinct()
    tipos_prospeccion_inicial = query.values_list('tipo_prospeccion', flat=True).distinct().exclude(tipo_prospeccion__isnull=True)
    areas_inicial = query.values_list('area', flat=True).distinct().exclude(area__isnull=True).exclude(area='')
    id_prospecciones_inicial = query.values_list('id_prospeccion', flat=True).distinct().exclude(id_prospeccion__isnull=True)

    # Ajustamos los nombres de los campos para ForeignKey (usamos id_prospeccion_id)
    df = pd.DataFrame(list(query.values('humedad', 'profundidad_promedio', 'area', 'id_prospeccion_id')))
    df = df.rename(columns={'id_prospeccion_id': 'id_prospeccion'})  # Renombramos para consistencia en las funciones

    context = {
        'proyectos': proyectos,
        'tipos_prospeccion_inicial': tipos_prospeccion_inicial,
        'areas_inicial': areas_inicial,
        'id_prospecciones_inicial': id_prospecciones_inicial,
        'selected_id_proyectos': json.dumps(id_proyectos),
        'selected_tipos_prospeccion': json.dumps(tipos_prospeccion),
        'selected_areas': json.dumps(areas),
        'selected_id_prospecciones': json.dumps(id_prospecciones),
    }

    if df.empty:
        context['error'] = "No hay datos disponibles para los filtros seleccionados."
    else:
        context['image_base64_area'] = generar_grafico_humedad_area(df)
        context['image_base64_prospeccion'] = generar_grafico_humedad_prospeccion(df)

    return render(request, 'laboratorio/ensayos/humedad/graficos_humedad.html', context)

# Funciones auxiliares para Humedad
def obtener_tipos_prospeccion_humedad(request):
    id_proyectos = request.GET.get('id_proyectos')
    query = Humedad.objects.all()  # Nombre correcto del modelo
    if id_proyectos:
        query = query.filter(id_proyecto__in=id_proyectos.split(','))
    tipos_list = list(query.values_list('tipo_prospeccion', flat=True).distinct().exclude(tipo_prospeccion__isnull=True))
    return JsonResponse({'options': tipos_list}, safe=False)

def obtener_id_prospecciones_humedad(request):
    id_proyectos = request.GET.get('id_proyectos')
    tipos_prospeccion = request.GET.get('tipos_prospeccion')
    areas = request.GET.get('areas')
    query = Humedad.objects.all()  # Nombre correcto del modelo
    if id_proyectos:
        query = query.filter(id_proyecto__in=id_proyectos.split(','))
    if tipos_prospeccion:
        query = query.filter(tipo_prospeccion__in=tipos_prospeccion.split(','))
    if areas:
        query = query.filter(area__in=areas.split(','))
    id_prospecciones_list = list(query.values_list('id_prospeccion', flat=True).distinct().exclude(id_prospeccion__isnull=True))
    return JsonResponse({'options': id_prospecciones_list}, safe=False)

def obtener_area_humedad(request):
    id_prospecciones = request.GET.get('id_prospecciones')
    if id_prospecciones:
        query = Humedad.objects.filter(id_prospeccion__in=id_prospecciones.split(','))  # Nombre correcto del modelo
        areas_list = list(query.values_list('area', flat=True).distinct().exclude(area__isnull=True).exclude(area=''))
        return JsonResponse({'options': areas_list})
    return JsonResponse({'options': []})

# # AGREGAR HUMEDAD

# logger = logging.getLogger(__name__)

# @login_required
# def agregar_humedad(request):
#     # Filtrar proyectos que tienen registros en Muestreo
#     proyectos_con_muestreo = Proyectos.objects.filter(muestreo__isnull=False).distinct()

#     if request.method == 'POST':
#         post_data = request.POST.copy()
#         proyecto_id = post_data.get('id_proyecto')
#         tipo_prospeccion = post_data.get('tipo_prospeccion')
#         area = post_data.get('area')

#         # Validar y asignar proyecto
#         if proyecto_id:
#             try:
#                 proyecto = Proyectos.objects.get(id=proyecto_id)
#                 post_data['id_proyecto'] = proyecto.id
#             except Proyectos.DoesNotExist:
#                 post_data['id_proyecto'] = None

#         # Manejar múltiples muestras y valores de humedad
#         id_prospecciones = post_data.getlist('id_prospeccion')  # Lista de prospecciones por fila
#         muestras = post_data.getlist('id_muestra')
#         humedades = post_data.getlist('humedad')
#         profundidades_desde = post_data.getlist('profundidad_desde')
#         profundidades_hasta = post_data.getlist('profundidad_hasta')
#         errors = []

#         for id_pros, muestra, humedad_val, desde, hasta in zip(id_prospecciones, muestras, humedades, profundidades_desde, profundidades_hasta):
#             if not id_pros or not muestra or not humedad_val:
#                 errors.append(f"Falta ID de prospección, ID de muestra o valor de humedad para una entrada.")
#                 continue
            
#             # Validar y asignar prospección
#             try:
#                 prospeccion = Prospecciones.objects.get(id_prospeccion=id_pros)
#                 post_data['id_prospeccion'] = prospeccion.id_prospeccion
#             except Prospecciones.DoesNotExist:
#                 errors.append(f"Prospección {id_pros} no encontrada.")
#                 continue

#             post_data['id_muestra'] = muestra
#             post_data['humedad'] = humedad_val
#             post_data['tipo_prospeccion'] = tipo_prospeccion
#             post_data['area'] = area

#             # Crear instancia de Humedad manualmente
#             try:
#                 humedad_obj = Humedad(
#                     id_proyecto=proyecto if proyecto_id else None,
#                     tipo_prospeccion=tipo_prospeccion,
#                     id_prospeccion=prospeccion,
#                     area=area,
#                     humedad=humedad_val,
#                     profundidad_promedio=(float(desde) + float(hasta)) / 2 if desde and hasta else 0,
#                     user=request.user
#                 )
#                 humedad_obj.save()
#                 logger.info(f"Humedad creada: ID {humedad_obj.id}")
#             except Exception as e:
#                 logger.error(f"Error al guardar humedad para muestra {muestra}: {e}")
#                 errors.append(f"Error al guardar muestra {muestra}: {e}")

#         if errors:
#             return render(request, 'laboratorio/ensayos/humedad/agregar_humedad.html', {
#                 'proyectos': proyectos_con_muestreo,
#                 'errors': errors
#             })
#         return redirect('listar_humedad')

#     return render(request, 'laboratorio/ensayos/humedad/agregar_humedad.html', {
#         'proyectos': proyectos_con_muestreo
#     })

# def listar_humedad(request):
#     humedades = Humedad.objects.all().order_by('id')
#     query = request.GET.get('q', '')
#     if query:
#         humedades = humedades.filter(
#             Q(id_proyecto__id__icontains=query) |
#             Q(id_prospeccion__id_prospeccion__icontains=query) |
#             Q(tipo_prospeccion__icontains=query) |
#             Q(humedad__icontains=query) |
#             Q(profundidad_promedio__icontains=query) |
#             Q(area__icontains=query)
#         )
#     paginator = Paginator(humedades, 1000)
#     page_number = request.GET.get('page')
#     page_obj = paginator.get_page(page_number)
#     if request.headers.get('x-requested-with') == 'XMLHttpRequest':
#         return render(request, 'laboratorio/ensayos/humedad/humedad_table.html', {'page_obj': page_obj})
#     return render(request, 'laboratorio/ensayos/humedad/listar_humedad.html', {'page_obj': page_obj, 'query': query})

# from django.http import HttpResponse
# from openpyxl import Workbook

# def export_to_excel_humedad(request):
#     # Obtener parámetros del request
#     query = request.GET.get('q', '')
#     headers_str = request.GET.get('headers', '')
#     headers = headers_str.split(',') if headers_str else ['ID', 'Tipo Prospección', 'ID Proyecto', 'ID Prospección', 'Humedad', 'Profundidad Promedio', 'Área']

#     # Crear respuesta para Excel
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = 'attachment; filename="humedad_filtrada.xlsx"'

#     # Crear el libro de Excel
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Humedad Filtrada"

#     # Agregar encabezados dinámicos
#     ws.append(headers)

#     # Obtener datos filtrados
#     humedades = Humedad.objects.all()
#     if query:
#         humedades = humedades.filter(id_prospeccion__id_prospeccion__icontains=query)

#     # Mapear campos según encabezados
#     for humedad in humedades:
#         row = []
#         for header in headers:
#             if header == 'ID':
#                 value = humedad.id
#             elif header == 'Tipo Prospección':
#                 value = humedad.tipo_prospeccion
#             elif header == 'ID Proyecto':
#                 value = str(humedad.id_proyecto)
#             elif header == 'ID Prospección':
#                 value = str(humedad.id_prospeccion)
#             elif header == 'Humedad (%)':
#                 value = humedad.humedad
#             elif header == 'Profundidad Promedio (m)':
#                 value = humedad.profundidad_promedio
#             elif header == 'Área':
#                 value = humedad.area
#             elif header == 'Usuario':
#                 value = humedad.user.username if humedad.user else 'Sin usuario'
#             else:
#                 value = '-'  # Valor por defecto si el encabezado no coincide
#             row.append(value)
#         ws.append(row)

#     # Guardar y devolver respuesta
#     wb.save(response)
#     return response

# def ver_humedad(request, id):
#     humedad = get_object_or_404(Humedad, id=id)
#     return render(request, 'laboratorio/ensayos/humedad/ver_humedad.html', {'humedad': humedad})

# from django.http import HttpResponse
# from reportlab.pdfgen import canvas

# def export_to_pdf_humedad(request):
#     # Obtener parámetros del request
#     query = request.GET.get('q', '')
#     headers_str = request.GET.get('headers', '')
#     headers = headers_str.split(',') if headers_str else ['ID', 'Tipo Prospección', 'ID Proyecto', 'ID Prospección', 'Humedad', 'Profundidad Promedio', 'Área']

#     # Crear respuesta para PDF
#     response = HttpResponse(content_type='application/pdf')
#     response['Content-Disposition'] = 'attachment; filename="humedad_filtrada.pdf"'

#     # Crear el objeto PDF
#     p = canvas.Canvas(response)

#     # Agregar título y encabezados dinámicos
#     p.drawString(100, 750, "Reporte de Humedad Filtrada")
#     p.drawString(100, 730, " | ".join(headers))

#     # Obtener datos filtrados
#     humedades = Humedad.objects.all()
#     if query:
#         humedades = humedades.filter(id_prospeccion__id_prospeccion__icontains=query)

#     # Rellenar datos
#     y = 710
#     for humedad in humedades:
#         row = []
#         for header in headers:
#             if header == 'ID':
#                 value = str(humedad.id)
#             elif header == 'Tipo Prospección':
#                 value = humedad.tipo_prospeccion
#             elif header == 'ID Proyecto':
#                 value = str(humedad.id_proyecto)
#             elif header == 'ID Prospección':
#                 value = str(humedad.id_prospeccion)
#             elif header == 'Humedad (%)':
#                 value = str(humedad.humedad)
#             elif header == 'Profundidad Promedio (m)':
#                 value = str(humedad.profundidad_promedio)
#             elif header == 'Área':
#                 value = humedad.area
#             elif header == 'Usuario':
#                 value = humedad.user.username if humedad.user else 'Sin usuario'
#             else:
#                 value = '-'
#             row.append(value)
#         p.drawString(100, y, " | ".join(row))
#         y -= 20
#         if y < 50:
#             p.showPage()
#             y = 750

#     # Finalizar el PDF
#     p.showPage()
#     p.save()
#     return response

# @login_required
# def editar_humedad(request, id):
#     humedad = get_object_or_404(Humedad, id=id)
#     proyectos = Proyectos.objects.all()
#     if request.method == 'POST':
#         form = HumedadForm(request.POST, instance=humedad)
#         if form.is_valid():
#             try:
#                 humedad = form.save()
#                 logger.info(f"Humedad editada: ID {humedad.id}, Área: {humedad.area}")
#                 return redirect('listar_humedad')
#             except Exception as e:
#                 logger.error(f"Error al editar humedad {humedad.id}: {e}")
#                 form.add_error(None, f"Error al editar: {e}")
#     else:
#         form = HumedadForm(instance=humedad)
#     return render(request, 'laboratorio/ensayos/humedad/editar_humedad.html', {'form': form, 'humedad': humedad, 'proyectos': proyectos})


# def eliminar_humedad(request, id):
#     humedad = get_object_or_404(Humedad, id=id)
#     if request.method == 'POST':
#         humedad.delete()
#         return redirect('listar_humedad')
#     return render(request, 'laboratorio/ensayos/humedad/eliminar_humedad.html', {'humedad': humedad})


# # Funciones de gráficos por proyecto
# def generar_grafico_humedad_area(df):
#     areas_unicas = df['area'].unique()
#     colores = {area: np.random.rand(3,) for area in areas_unicas}
#     fig, ax = plt.subplots()
#     for area in areas_unicas:
#         datos_area = df[df['area'] == area]
#         ax.scatter(
#             datos_area['humedad'],
#             datos_area['profundidad_promedio'],
#             color=colores[area],
#             s=100,
#             label=area,
#             marker='+'
#         )
#     ax.xaxis.set_ticks_position("top")
#     ax.xaxis.set_label_position("top")
#     ax.invert_yaxis()
#     ax.set_xlabel("Humedad (%)")
#     ax.set_ylabel("Profundidad (m)")
#     ax.grid(True)
#     plt.subplots_adjust(right=0.75)
#     num_columns = (len(areas_unicas) + 24) // 25
#     legend = ax.legend(loc='center left', bbox_to_anchor=(1, 0.5), ncol=num_columns)
#     buffer = BytesIO()
#     fig.savefig(buffer, format='png', dpi=100, bbox_inches='tight', bbox_extra_artists=[legend])
#     buffer.seek(0)
#     plt.close()
#     image_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
#     return image_base64

# # Funciones de gráficos por prospeccion
# def generar_grafico_humedad_prospeccion(df):
#     df['id_prospeccion'] = df['id_prospeccion'].astype(str)
#     df['id_prospeccion_unico'] = df.groupby('id_prospeccion').cumcount() + 1
#     df['etiqueta'] = df['id_prospeccion'] + '-' + df['id_prospeccion_unico'].astype(str)
#     prospecciones_unicas = df['etiqueta'].unique()
#     colores = {prospeccion: np.random.rand(3,) for prospeccion in prospecciones_unicas}
#     fig, ax = plt.subplots()
#     for prospeccion in prospecciones_unicas:
#         datos_prospeccion = df[df['etiqueta'] == prospeccion]
#         ax.scatter(
#             datos_prospeccion['humedad'],
#             datos_prospeccion['profundidad_promedio'],
#             color=colores[prospeccion],
#             s=100,
#             label=prospeccion,
#             marker='+'
#         )
#     ax.xaxis.set_ticks_position("top")
#     ax.xaxis.set_label_position("top")
#     ax.invert_yaxis()
#     ax.set_xlabel("Humedad (%)")
#     ax.set_ylabel("Profundidad (m)")
#     ax.grid(True)
#     plt.subplots_adjust(right=0.75)
#     num_columns = (len(prospecciones_unicas) + 24) // 25
#     legend = ax.legend(loc='center left', bbox_to_anchor=(1, 0.5), ncol=num_columns)
#     buffer = BytesIO()
#     fig.savefig(buffer, format='png', dpi=100, bbox_inches='tight', bbox_extra_artists=[legend])
#     buffer.seek(0)
#     plt.close()
#     image_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
#     return image_base64


# # Funciones de gráficos
# def graficos_humedad(request):
#     id_proyectos = request.GET.getlist('id_proyectos')
#     tipos_prospeccion = request.GET.getlist('tipos_prospeccion')
#     areas = request.GET.getlist('areas')
#     id_prospecciones = request.GET.getlist('id_prospecciones')
#     query = Humedad.objects.all()
#     if id_proyectos:
#         query = query.filter(id_proyecto__in=id_proyectos)
#     if tipos_prospeccion:
#         query = query.filter(tipo_prospeccion__in=tipos_prospeccion)
#     if areas:
#         query = query.filter(area__in=areas)
#     if id_prospecciones:
#         query = query.filter(id_prospeccion__in=id_prospecciones)
#     proyectos = query.values('id_proyecto').distinct()
#     tipos_prospeccion_inicial = (query.values_list('tipo_prospeccion', flat=True)
#                                  .distinct().exclude(tipo_prospeccion__isnull=True).exclude(tipo_prospeccion=''))
#     areas_inicial = (query.values_list('area', flat=True)
#                      .distinct().exclude(area__isnull=True).exclude(area=''))
#     id_prospecciones_inicial = (query.values_list('id_prospeccion', flat=True)
#                                 .distinct().exclude(id_prospeccion__isnull=True).exclude(id_prospeccion=''))
#     humedades = Humedad.objects.all()
#     if id_proyectos:
#         humedades = humedades.filter(id_proyecto__in=id_proyectos)
#     if tipos_prospeccion:
#         humedades = humedades.filter(tipo_prospeccion__in=tipos_prospeccion)
#     if areas:
#         humedades = humedades.filter(area__in=areas)
#     if id_prospecciones:
#         humedades = humedades.filter(id_prospeccion__in=id_prospecciones)
#     df = pd.DataFrame.from_records(humedades.values('id_proyecto', 'id_prospeccion', 'humedad', 'profundidad_promedio', 'area'))
#     if df.empty:
#         context = {
#             'error': "No hay datos para graficar.",
#             'proyectos': proyectos,
#             'tipos_prospeccion_inicial': tipos_prospeccion_inicial,
#             'areas_inicial': areas_inicial,
#             'id_prospecciones_inicial': id_prospecciones_inicial,
#             'selected_id_proyectos': json.dumps(id_proyectos),
#             'selected_tipos_prospeccion': json.dumps(tipos_prospeccion),
#             'selected_areas': json.dumps(areas),
#             'selected_id_prospecciones': json.dumps(id_prospecciones),
#         }
#         return render(request, 'laboratorio/ensayos/humedad/graficos_humedad.html', context)
#     image_base64_area = generar_grafico_humedad_area(df)
#     image_base64_prospeccion = generar_grafico_humedad_prospeccion(df)
#     context = {
#         'image_base64_area': image_base64_area,
#         'image_base64_prospeccion': image_base64_prospeccion,
#         'proyectos': proyectos,
#         'tipos_prospeccion_inicial': tipos_prospeccion_inicial,
#         'areas_inicial': areas_inicial,
#         'id_prospecciones_inicial': id_prospecciones_inicial,
#         'selected_id_proyectos': json.dumps(id_proyectos),
#         'selected_tipos_prospeccion': json.dumps(tipos_prospeccion),
#         'selected_areas': json.dumps(areas),
#         'selected_id_prospecciones': json.dumps(id_prospecciones),
#     }
#     return render(request, 'laboratorio/ensayos/humedad/graficos_humedad.html', context)


# ##Obtener funciones API para humedad
# # Obtener tipos de prospección para Humedad
# def obtener_tipos_prospeccion_humedad(request):
#     id_proyectos = request.GET.getlist('id_proyecto')
#     tipos = Humedad.objects.all()
#     if id_proyectos:
#         tipos = tipos.filter(id_proyecto__in=id_proyectos)
#     tipos_list = list(tipos.values_list('tipo_prospeccion', flat=True)
#                      .distinct()
#                      .exclude(tipo_prospeccion__isnull=True)
#                      .exclude(tipo_prospeccion=''))
#     print(f"Tipos de prospección para id_proyectos={id_proyectos}: {tipos_list}")
#     return JsonResponse({'options': tipos_list}, safe=False)

# # Obtener áreas para Humedad
# def obtener_areas_humedad(request):
#     id_proyectos = request.GET.getlist('id_proyecto')
#     areas = Humedad.objects.all()
#     if id_proyectos:
#         areas = areas.filter(id_proyecto__in=id_proyectos)
#     areas_list = list(areas.values_list('area', flat=True)
#                      .distinct()
#                      .exclude(area__isnull=True)
#                      .exclude(area=''))
#     print(f"Áreas para id_proyectos={id_proyectos}: {areas_list}")
#     return JsonResponse({'options': areas_list}, safe=False)

# # Obtener IDs de prospección para Humedad
# def obtener_id_prospecciones_humedad(request):
#     id_proyectos = request.GET.getlist('id_proyecto')
#     tipos_prospeccion = request.GET.getlist('tipo_prospeccion')
#     areas = request.GET.getlist('area')
#     query = Humedad.objects.all()
#     if id_proyectos:
#         query = query.filter(id_proyecto__in=id_proyectos)
#     if tipos_prospeccion:
#         query = query.filter(tipo_prospeccion__in=tipos_prospeccion)
#     if areas:
#         query = query.filter(area__in=areas)
#     id_prospecciones_list = list(query.values_list('id_prospeccion', flat=True)
#                                 .distinct()
#                                 .exclude(id_prospeccion__isnull=True)
#                                 .exclude(id_prospeccion=''))
#     print(f"IDs de prospección para id_proyectos={id_proyectos}, tipos={tipos_prospeccion}, áreas={areas}: {id_prospecciones_list}")
#     return JsonResponse({'options': id_prospecciones_list}, safe=False)

# # Obtener proyectos para Humedad
# def obtener_id_proyecto_humedad(request):
#     tipos_prospeccion = request.GET.getlist('tipo_prospeccion')
#     areas = request.GET.getlist('area')
#     id_prospecciones = request.GET.getlist('id_prospeccion')
#     query = Humedad.objects.all()
#     if tipos_prospeccion:
#         query = query.filter(tipo_prospeccion__in=tipos_prospeccion)
#     if areas:
#         query = query.filter(area__in=areas)
#     if id_prospecciones:
#         query = query.filter(id_prospeccion__in=id_prospecciones)
#     proyectos_list = list(query.values_list('id_proyecto', flat=True)
#                          .distinct()
#                          .exclude(id_proyecto__isnull=True)
#                          .exclude(id_proyecto=''))
#     print(f"Proyectos para tipos={tipos_prospeccion}, áreas={areas}, id_prospecciones={id_prospecciones}: {proyectos_list}")
#     return JsonResponse({'options': proyectos_list}, safe=False)



#### GRANULOMETRIA ##############################################
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.apps import apps
from django.core.paginator import Paginator
from django.db.models import Q
from .models import Granulometria, Proyectos, Prospecciones, Muestreo
from .forms import GranulometriaForm
import logging
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.ticker import LogLocator, FuncFormatter
from io import BytesIO
import base64
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
import json

logger = logging.getLogger(__name__)

# 1. Agregar
@login_required
def agregar_granulometria(request):
    proyectos_con_muestreo = Proyectos.objects.filter(muestreo__isnull=False).distinct()
    tipo_prospeccion_choices = []
    prospecciones = []
    muestras = []
    area = ""

    if request.method == 'POST':
        logger.info(f"Datos POST recibidos: {request.POST}")
        proyecto_id = request.POST.get('id_proyecto')
        tipo_prospeccion = request.POST.get('tipo_prospeccion')
        id_prospecciones = request.POST.getlist('id_prospeccion[]')
        id_muestras = request.POST.getlist('id_muestra[]')
        n_0075s = request.POST.getlist('n_0075[]')
        n_0110s = request.POST.getlist('n_0110[]')
        n_0250s = request.POST.getlist('n_0250[]')
        n_0420s = request.POST.getlist('n_0420[]')
        n_0840s = request.POST.getlist('n_0840[]')
        n_2000s = request.POST.getlist('n_2000[]')
        n_4760s = request.POST.getlist('n_4760[]')
        n_9520s = request.POST.getlist('n_9520[]')
        n_19000s = request.POST.getlist('n_19000[]')
        n_25400s = request.POST.getlist('n_25400[]')
        n_38100s = request.POST.getlist('n_38100[]')
        n_50800s = request.POST.getlist('n_50800[]')
        n_63500s = request.POST.getlist('n_63500[]')
        n_75000s = request.POST.getlist('n_75000[]')
        areas = request.POST.getlist('area[]')
        errors = []

        if not proyecto_id:
            errors.append("Debe seleccionar un proyecto.")
            return render(request, 'laboratorio/ensayos/granulometria/agregar_granulometria.html', {
                'proyectos': proyectos_con_muestreo,
                'errors': errors
            })

        muestreos = Muestreo.objects.filter(id_proyecto_id=proyecto_id)
        if not muestreos.exists():
            errors.append("No hay muestreos asociados a este proyecto.")
            return render(request, 'laboratorio/ensayos/granulometria/agregar_granulometria.html', {
                'proyectos': proyectos_con_muestreo,
                'errors': errors
            })

        if tipo_prospeccion:
            muestreos = muestreos.filter(tipo_prospeccion=tipo_prospeccion)
            if not muestreos.exists():
                errors.append("No hay muestreos con este tipo de prospección.")

        if id_prospecciones and id_prospecciones[0]:
            muestreos = muestreos.filter(id_prospeccion__id_prospeccion__in=id_prospecciones)
            if not muestreos.exists():
                errors.append("No hay muestreos con las prospecciones seleccionadas.")

        if id_muestras and id_muestras[0]:
            muestreos = muestreos.filter(id_muestra__in=id_muestras)
            if not muestreos.exists():
                errors.append("No hay muestreos con las muestras seleccionadas.")

        if not (len(id_prospecciones) == len(id_muestras) == len(n_0075s) == len(n_0110s) == len(n_0250s) == len(n_0420s) == len(n_0840s) == len(n_2000s) == len(n_4760s) == len(n_9520s) == len(n_19000s) == len(n_25400s) == len(n_38100s) == len(n_50800s) == len(n_63500s) == len(n_75000s) == len(areas)):
            errors.append("El número de prospecciones, muestras y datos granulométricos no coincide.")
            logger.error(f"Longitudes inconsistentes: prospecciones={len(id_prospecciones)}, muestras={len(id_muestras)}")
        else:
            for id_pros, id_muestra, n_0075, n_0110, n_0250, n_0420, n_0840, n_2000, n_4760, n_9520, n_19000, n_25400, n_38100, n_50800, n_63500, n_75000, area_val in zip(id_prospecciones, id_muestras, n_0075s, n_0110s, n_0250s, n_0420s, n_0840s, n_2000s, n_4760s, n_9520s, n_19000s, n_25400s, n_38100s, n_50800s, n_63500s, n_75000s, areas):
                if not id_pros or not id_muestra:
                    errors.append(f"Falta ID de prospección ({id_pros}) o ID de muestra ({id_muestra}).")
                    continue
                
                try:
                    muestreo = muestreos.get(id_prospeccion__id_prospeccion=id_pros, id_muestra=id_muestra)
                    profundidad_desde = float(muestreo.profundidad_desde or 0)  # Si es None, usa 0
                    profundidad_hasta = float(muestreo.profundidad_hasta or 0)  # Si es None, usa 0
                    profundidad_promedio = (profundidad_desde + profundidad_hasta) / 2

                    granulometria_obj = Granulometria(
                        id_proyecto=muestreo.id_proyecto,
                        tipo_prospeccion=muestreo.tipo_prospeccion,
                        id_prospeccion=muestreo.id_prospeccion,
                        id_muestra=muestreo.id_muestra,
                        profundidad_desde=muestreo.profundidad_desde,
                        profundidad_hasta=muestreo.profundidad_hasta,
                        profundidad_promedio=profundidad_promedio,
                        area=area_val,
                        user=request.user,
                        n_0075=float(n_0075) if n_0075 else None,
                        n_0110=float(n_0110) if n_0110 else None,
                        n_0250=float(n_0250) if n_0250 else None,
                        n_0420=float(n_0420) if n_0420 else None,
                        n_0840=float(n_0840) if n_0840 else None,
                        n_2000=float(n_2000) if n_2000 else None,
                        n_4760=float(n_4760) if n_4760 else None,
                        n_9520=float(n_9520) if n_9520 else None,
                        n_19000=float(n_19000) if n_19000 else None,
                        n_25400=float(n_25400) if n_25400 else None,
                        n_38100=float(n_38100) if n_38100 else None,
                        n_50800=float(n_50800) if n_50800 else None,
                        n_63500=float(n_63500) if n_63500 else None,
                        n_75000=float(n_75000) if n_75000 else None
                    )
                    granulometria_obj.save()
                    logger.info(f"Granulometría guardada: ID={granulometria_obj.id}, Muestra={id_muestra}, Profundidad Promedio={profundidad_promedio}")
                except Muestreo.DoesNotExist:
                    errors.append(f"Muestra {id_muestra} con prospección {id_pros} no encontrada.")
                except ValueError as e:
                    errors.append(f"Valor inválido en datos granulométricos: {e}")
                except Exception as e:
                    errors.append(f"Error al guardar muestra {id_muestra}: {e}")

        if errors:
            tipo_prospeccion_choices = Muestreo.objects.filter(id_proyecto_id=proyecto_id).values_list('tipo_prospeccion', flat=True).distinct()
            prospecciones = Muestreo.objects.filter(id_proyecto_id=proyecto_id).values_list('id_prospeccion__id_prospeccion', flat=True).distinct()
            muestras = Muestreo.objects.filter(id_proyecto_id=proyecto_id).values_list('id_muestra', flat=True).distinct()
            return render(request, 'laboratorio/ensayos/granulometria/agregar_granulometria.html', {
                'proyectos': proyectos_con_muestreo,
                'tipo_prospeccion_choices': tipo_prospeccion_choices,
                'prospecciones': prospecciones,
                'muestras': muestras,
                'area': area,
                'errors': errors
            })

        return redirect('listar_granulometria')

    proyecto_id = request.GET.get('id_proyecto')
    tipo_prospeccion = request.GET.get('tipo_prospeccion')
    id_prospeccion = request.GET.get('id_prospeccion')
    id_muestra = request.GET.get('id_muestra')

    muestreos = Muestreo.objects.all()
    if proyecto_id:
        muestreos = muestreos.filter(id_proyecto_id=proyecto_id)
        tipo_prospeccion_choices = muestreos.values_list('tipo_prospeccion', flat=True).distinct()
        prospecciones = muestreos.values_list('id_prospeccion__id_prospeccion', flat=True).distinct()
        muestras = muestreos.values_list('id_muestra', flat=True).distinct()
        area = muestreos.first().area if muestreos.exists() else ""

    if tipo_prospeccion:
        muestreos = muestreos.filter(tipo_prospeccion=tipo_prospeccion)
        prospecciones = muestreos.values_list('id_prospeccion__id_prospeccion', flat=True).distinct()
        muestras = muestreos.values_list('id_muestra', flat=True).distinct()

    if id_prospeccion:
        muestreos = muestreos.filter(id_prospeccion__id_prospeccion=id_prospeccion)
        muestras = muestreos.values_list('id_muestra', flat=True).distinct()
        area = muestreos.first().area if muestreos.exists() else ""

    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if is_ajax:
        data = {
            'tipo_prospeccion_choices': list(tipo_prospeccion_choices),
            'prospecciones': list(prospecciones),
            'muestras': list(muestras),
            'area': area
        }
        if id_muestra:
            muestreo = muestreos.filter(id_muestra=id_muestra).first()
            if muestreo:
                data['profundidad_desde'] = str(muestreo.profundidad_desde) if muestreo.profundidad_desde is not None else '0'
                data['profundidad_hasta'] = str(muestreo.profundidad_hasta) if muestreo.profundidad_hasta is not None else '0'
        return JsonResponse(data)

    return render(request, 'laboratorio/ensayos/granulometria/agregar_granulometria.html', {
        'proyectos': proyectos_con_muestreo,
        'tipo_prospeccion_choices': tipo_prospeccion_choices,
        'prospecciones': prospecciones,
        'muestras': muestras,
        'area': area
    })

# 2. Listar
@login_required
def listar_granulometria(request):
    granulometria_list = Granulometria.objects.all().order_by('id')
    query = request.GET.get('q', '')

    if query:
        granulometria_model = apps.get_model('administrador', 'Granulometria')
        campos = [field.name for field in granulometria_model._meta.get_fields() if isinstance(field, (models.CharField, models.TextField, models.DateField, models.DateTimeField, models.DecimalField, models.FloatField, models.ForeignKey))]
        query_filter = Q()

        try:
            query_numeric = float(query)
            numeric_search = True
        except ValueError:
            numeric_search = False

        for field in campos:
            field_instance = granulometria_model._meta.get_field(field)
            if isinstance(field_instance, (models.DateField, models.DateTimeField)):
                query_filter |= Q(**{f"{field}__year__icontains": query})
                query_filter |= Q(**{f"{field}__month__icontains": query})
                query_filter |= Q(**{f"{field}__day__icontains": query})
            elif isinstance(field_instance, models.ForeignKey):
                related_model = field_instance.related_model
                related_fields = [f"{field}__{related_field.name}" for related_field in related_model._meta.get_fields() if isinstance(related_field, (models.CharField, models.TextField))]
                for related_field in related_fields:
                    query_filter |= Q(**{f"{related_field}__icontains": query})
            elif isinstance(field_instance, (models.DecimalField, models.FloatField)) and numeric_search:
                query_filter |= Q(**{f"{field}": query_numeric})
            else:
                query_filter |= Q(**{f"{field}__icontains": query})

        granulometria_list = granulometria_list.filter(query_filter)

    paginator = Paginator(granulometria_list, 1000)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'laboratorio/ensayos/granulometria/listar_granulometria.html', {
        'page_obj': page_obj,
        'query': query
    })

# 3. Editar
@login_required
def editar_granulometria(request, id):
    granulometria_obj = get_object_or_404(Granulometria, id=id)
    muestreo = Muestreo.objects.filter(id_muestra=granulometria_obj.id_muestra).first()
    proyectos = Proyectos.objects.filter(muestreo__isnull=False).distinct()

    if request.method == 'POST':
        form = GranulometriaForm(request.POST, instance=granulometria_obj)
        if form.is_valid():
            try:
                granulometria_obj = form.save()
                logger.info(f"Granulometría editada: ID {granulometria_obj.id}")
                return redirect('listar_granulometria')
            except Exception as e:
                logger.error(f"Error al editar: {e}")
                form.add_error(None, f"Error al editar: {e}")
    else:
        form = GranulometriaForm(instance=granulometria_obj)

    return render(request, 'laboratorio/ensayos/granulometria/editar_granulometria.html', {
        'form': form,
        'granulometria_obj': granulometria_obj,
        'proyectos': proyectos
    })

# 4. Ver
@login_required
def ver_granulometria(request, id):
    granulometria_obj = get_object_or_404(Granulometria, id=id)

    history_records = []
    for record in granulometria_obj.history.all():
        changes = []
        previous = record.prev_record
        if previous:
            for field in record.history_object._meta.fields:
                old_value = getattr(previous, field.name, None)
                new_value = getattr(record, field.name, None)
                if old_value != new_value:
                    changes.append({
                        'field_name': field.verbose_name,
                        'old_value': old_value if old_value is not None else '-',
                        'new_value': new_value if new_value is not None else '-',
                    })
        history_records.append({
            'date': record.history_date,
            'user': record.history_user,
            'type': record.history_type,
            'reason': record.history_change_reason,
            'changes': changes if changes else None,
        })

    return render(request, 'laboratorio/ensayos/granulometria/ver_granulometria.html', {
        'granulometria': granulometria_obj,
        'history_records': history_records,
    })

# 5. Eliminar
@login_required
def eliminar_granulometria(request, id):
    granulometria_obj = get_object_or_404(Granulometria, id=id)
    if request.method == 'POST':
        granulometria_obj.delete()
        return redirect('listar_granulometria')
    return render(request, 'laboratorio/ensayos/granulometria/eliminar_granulometria.html', {
        'granulometria': granulometria_obj
    })

# 6. Exportar Excel
@login_required
def export_to_excel_granulometria(request):
    query = request.GET.get('q', '').strip()
    headers = ['ID', 'ID Proyecto', 'ID Prospección', 'Tipo Prospección', 'ID Muestra',
               'Profundidad Desde', 'Profundidad Hasta', 'Profundidad Promedio',
               'Área', 'Usuario', 'N° 0.075', 'N° 0.110', 'N° 0.250', 'N° 0.420',
               'N° 0.840', 'N° 2.000', 'N° 4.760', 'N° 9.520', 'N° 19.000',
               'N° 25.400', 'N° 38.100', 'N° 50.800', 'N° 63.500', 'N° 75.000']

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="granulometria_filtrado.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Granulometría"
    ws.append(headers)

    try:
        granulometria_list = Granulometria.objects.select_related(
            'id_proyecto', 'id_prospeccion', 'user'
        ).prefetch_related('id_muestra__muestreo_set').all()

        logger.debug(f"Registros iniciales: {granulometria_list.count()}, query: '{query}'")

        if query:
            filtros = Q()
            try:
                query_numeric = float(query)
                filtros |= (
                    Q(profundidad_desde=query_numeric) |
                    Q(profundidad_hasta=query_numeric) |
                    Q(profundidad_promedio=query_numeric) |
                    Q(n_0075=query_numeric) | Q(n_0110=query_numeric) |
                    Q(n_0250=query_numeric) | Q(n_0420=query_numeric) |
                    Q(n_0840=query_numeric) | Q(n_2000=query_numeric) |
                    Q(n_4760=query_numeric) | Q(n_9520=query_numeric) |
                    Q(n_19000=query_numeric) | Q(n_25400=query_numeric) |
                    Q(n_38100=query_numeric) | Q(n_50800=query_numeric) |
                    Q(n_63500=query_numeric) | Q(n_75000=query_numeric)
                )
            except ValueError:
                pass

            filtros |= (
                Q(id__icontains=query) |
                Q(id_proyecto__id__icontains=query) |
                Q(id_prospeccion__id_prospeccion__icontains=query) |
                Q(tipo_prospeccion__icontains=query) |
                Q(id_muestra__icontains=query) |
                Q(area__icontains=query) |
                Q(user__username__icontains=query)
            )

            granulometria_list = granulometria_list.filter(filtros)

        logger.debug(f"Registros después del filtro: {granulometria_list.count()}")

        for granulometria in granulometria_list:
            try:
                muestreo = Muestreo.objects.filter(id_muestra=granulometria.id_muestra).first()
                row = [
                    str(granulometria.id) if granulometria.id is not None else '',
                    str(granulometria.id_proyecto.id) if granulometria.id_proyecto else '',
                    str(granulometria.id_prospeccion.id_prospeccion) if granulometria.id_prospeccion else '',
                    str(granulometria.tipo_prospeccion or ''),
                    str(granulometria.id_muestra or ''),
                    str(granulometria.profundidad_desde) if granulometria.profundidad_desde is not None else '',
                    str(granulometria.profundidad_hasta) if granulometria.profundidad_hasta is not None else '',
                    str(granulometria.profundidad_promedio) if granulometria.profundidad_promedio is not None else '',
                    str(muestreo.area) if muestreo and hasattr(muestreo, 'area') else '',
                    str(granulometria.user.username) if granulometria.user else 'Sin usuario',
                    str(granulometria.n_0075) if granulometria.n_0075 is not None else '',
                    str(granulometria.n_0110) if granulometria.n_0110 is not None else '',
                    str(granulometria.n_0250) if granulometria.n_0250 is not None else '',
                    str(granulometria.n_0420) if granulometria.n_0420 is not None else '',  # Corregido
                    str(granulometria.n_0840) if granulometria.n_0840 is not None else '',
                    str(granulometria.n_2000) if granulometria.n_2000 is not None else '',
                    str(granulometria.n_4760) if granulometria.n_4760 is not None else '',
                    str(granulometria.n_9520) if granulometria.n_9520 is not None else '',
                    str(granulometria.n_19000) if granulometria.n_19000 is not None else '',
                    str(granulometria.n_25400) if granulometria.n_25400 is not None else '',
                    str(granulometria.n_38100) if granulometria.n_38100 is not None else '',
                    str(granulometria.n_50800) if granulometria.n_50800 is not None else '',
                    str(granulometria.n_63500) if granulometria.n_63500 is not None else '',
                    str(granulometria.n_75000) if granulometria.n_75000 is not None else ''
                ]
                ws.append(row)
            except Exception as e:
                logger.error(f"Error al procesar registro {granulometria.id}: {str(e)}")
                continue

        wb.save(response)
        return response

    except Exception as e:
        logger.error(f"Error en export_to_excel_granulometria: {str(e)}")
        ws.append(["Error al generar el reporte", str(e)])
        wb.save(response)
        return response

# 7. Exportar PDF
@login_required
def export_to_pdf_granulometria(request):
    query = request.GET.get('q', '').strip()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="granulometria_filtrado.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(letter),
                            rightMargin=30, leftMargin=30,
                            topMargin=30, bottomMargin=30)

    elements = []
    styles = getSampleStyleSheet()

    title = Paragraph("Lista de Granulometría",
                     styles['Heading1'].clone('Title', alignment=1, fontSize=12, spaceAfter=10))
    elements.append(title)
    elements.append(Paragraph("<br/>", styles['Normal']))

    headers = ['ID', 'Proyecto', 'Prospección', 'Tipo', 'Muestra',
               'Prof. Ini', 'Prof. Fin', 'Prof. Prom', 'Área', 'Usuario',
               '0.075', '0.110', '0.250', '0.420', '0.840', '2.000',
               '4.760', '9.520', '19.000', '25.400', '38.100', '50.800',
               '63.500', '75.000']

    data = [headers]

    try:
        granulometria_list = Granulometria.objects.select_related(
            'id_proyecto', 'id_prospeccion', 'user'
        ).prefetch_related('id_muestra__muestreo_set').all()

        logger.debug(f"Registros iniciales: {granulometria_list.count()}, query: '{query}'")

        if query:
            filtros = Q()
            try:
                query_numeric = float(query)
                filtros |= (
                    Q(profundidad_desde=query_numeric) |
                    Q(profundidad_hasta=query_numeric) |
                    Q(profundidad_promedio=query_numeric) |
                    Q(n_0075=query_numeric) | Q(n_0110=query_numeric) |
                    Q(n_0250=query_numeric) | Q(n_0420=query_numeric) |
                    Q(n_0840=query_numeric) | Q(n_2000=query_numeric) |
                    Q(n_4760=query_numeric) | Q(n_9520=query_numeric) |
                    Q(n_19000=query_numeric) | Q(n_25400=query_numeric) |
                    Q(n_38100=query_numeric) | Q(n_50800=query_numeric) |
                    Q(n_63500=query_numeric) | Q(n_75000=query_numeric)
                )
            except ValueError:
                pass

            filtros |= (
                Q(id__icontains=query) |
                Q(id_proyecto__id__icontains=query) |
                Q(id_prospeccion__id_prospeccion__icontains=query) |
                Q(tipo_prospeccion__icontains=query) |
                Q(id_muestra__icontains=query) |
                Q(area__icontains=query) |
                Q(user__username__icontains=query)
            )

            granulometria_list = granulometria_list.filter(filtros)

        logger.debug(f"Registros después del filtro: {granulometria_list.count()}")

        for granulometria in granulometria_list:
            try:
                muestreo = Muestreo.objects.filter(id_muestra=granulometria.id_muestra).first()
                row = [
                    str(granulometria.id) if granulometria.id is not None else '',
                    str(granulometria.id_proyecto.id) if granulometria.id_proyecto else '',
                    str(granulometria.id_prospeccion.id_prospeccion) if granulometria.id_prospeccion else '',
                    str(granulometria.tipo_prospeccion or ''),
                    str(granulometria.id_muestra or ''),
                    str(granulometria.profundidad_desde) if granulometria.profundidad_desde is not None else '',
                    str(granulometria.profundidad_hasta) if granulometria.profundidad_hasta is not None else '',
                    str(granulometria.profundidad_promedio) if granulometria.profundidad_promedio is not None else '',
                    str(muestreo.area) if muestreo and hasattr(muestreo, 'area') else '',
                    str(granulometria.user.username) if granulometria.user else 'Sin usuario',
                    str(granulometria.n_0075) if granulometria.n_0075 is not None else '',
                    str(granulometria.n_0110) if granulometria.n_0110 is not None else '',
                    str(granulometria.n_0250) if granulometria.n_0250 is not None else '',
                    str(granulometria.n_0420) if granulometria.n_0420 is not None else '',  # Corregido
                    str(granulometria.n_0840) if granulometria.n_0840 is not None else '',
                    str(granulometria.n_2000) if granulometria.n_2000 is not None else '',
                    str(granulometria.n_4760) if granulometria.n_4760 is not None else '',
                    str(granulometria.n_9520) if granulometria.n_9520 is not None else '',
                    str(granulometria.n_19000) if granulometria.n_19000 is not None else '',
                    str(granulometria.n_25400) if granulometria.n_25400 is not None else '',
                    str(granulometria.n_38100) if granulometria.n_38100 is not None else '',
                    str(granulometria.n_50800) if granulometria.n_50800 is not None else '',
                    str(granulometria.n_63500) if granulometria.n_63500 is not None else '',
                    str(granulometria.n_75000) if granulometria.n_75000 is not None else ''
                ]
                data.append(row)
            except Exception as e:
                logger.error(f"Error al procesar registro {granulometria.id}: {str(e)}")
                continue

        table = Table(data, colWidths=[0.5*inch] + [0.8*inch]*23)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))

        elements.append(table)
        doc.build(elements)
        return response

    except Exception as e:
        logger.error(f"Error en export_to_pdf_granulometria: {str(e)}")
        doc.build([Paragraph(f"Error al generar el reporte: {str(e)}", styles['Normal'])])
        return response

# 8. Gráficos Granulometría
def generar_grafico_granulometria_area(df):
    if 'id_prospeccion' not in df.columns:
        raise KeyError("La columna 'id_prospeccion' no está presente en los datos iniciales.")
    df['id_proyecto'] = df['id_proyecto'].astype(str)
    df['etiqueta'] = df['id_proyecto']
    proyectos_unicos = df['etiqueta'].unique()
    colores = {proyecto: np.random.rand(3,) for proyecto in proyectos_unicos}
    fig, ax = plt.subplots()
    x_values = np.array([0.075, 0.110, 0.250, 0.420, 0.840, 2.000, 4.760, 9.520, 19.000, 25.400, 38.100, 50.800, 63.500, 75.000])
    for proyecto in proyectos_unicos:
        datos_proyecto = df[df['etiqueta'] == proyecto]
        for i in range(len(datos_proyecto)):
            y_values = datos_proyecto.iloc[i, 2:16].values
            sorted_indices = np.argsort(x_values)[::-1]
            x_values_sorted = x_values[sorted_indices]
            y_values_sorted = y_values[sorted_indices]
            ax.plot(x_values_sorted, y_values_sorted, color=colores[proyecto], label=proyecto if i == 0 else "")
    ax.set_xscale('log')
    ax.set_xlim(0.01, 100)
    ax.xaxis.set_major_locator(LogLocator(base=10, numticks=10))
    ax.xaxis.set_minor_locator(LogLocator(base=10, subs='auto', numticks=50))
    def log_format(x, _):
        if x < 1:
            return f"{x:.2f}"
        else:
            return f"{int(x)}"
    ax.xaxis.set_major_formatter(FuncFormatter(log_format))
    ax.invert_xaxis()
    ax.set_xlabel('Tamaño de partícula (mm)')
    ax.set_ylabel('Porcentaje que pasa (%)')
    ax.set_title('Distribución Granulométrica por Área', pad=40)
    handles, labels = ax.get_legend_handles_labels()
    max_legends = 27
    if len(handles) > max_legends:
        handles = handles[:max_legends]
        labels = labels[:max_legends]
    ax.legend(handles, labels, loc='upper right', fontsize=8, ncol=1, columnspacing=0.5,
              handlelength=2, handletextpad=0.5, borderpad=0.5, framealpha=0.5, borderaxespad=0.5)
    ax.axvline(x=76.8, color='black', linestyle='--', linewidth=1, label='3')
    ax.axvline(x=19, color='black', linestyle='--', linewidth=1, label='3/4')
    ax.axvline(x=4.76, color='black', linestyle='--', linewidth=1, label='N°4')
    ax.axvline(x=0.42, color='black', linestyle='--', linewidth=1, label='N°40')
    ax.axvline(x=0.0749, color='black', linestyle='--', linewidth=1, label='N°200')
    ax.axhline(y=50, color='black', linestyle='--', linewidth=1, label='50')
    ax.text(76.8, 106.5, '3', color='black', va='bottom', ha='center', fontsize=9, bbox=dict(facecolor='white', edgecolor='none', alpha=0))
    ax.text(19, 106.5, '3/4', color='black', va='bottom', ha='center', fontsize=9, bbox=dict(facecolor='white', edgecolor='none', alpha=0))
    ax.text(4.76, 106.5, '4', color='black', va='bottom', ha='center', fontsize=9, bbox=dict(facecolor='white', edgecolor='none', alpha=0))
    ax.text(0.42, 106.5, '40', color='black', va='bottom', ha='center', fontsize=9, bbox=dict(facecolor='white', edgecolor='none', alpha=0))
    ax.text(0.0749, 106.5, '200', color='black', va='bottom', ha='center', fontsize=9, bbox=dict(facecolor='white', edgecolor='none', alpha=0))
    ax.text(140, 50, '50', color='black', va='bottom', ha='left', fontsize=9, bbox=dict(facecolor='white', edgecolor='none', alpha=0))
    def add_text_with_box(ax, x, y, text, color, pad=0.5, boxstyle='square'):
        ax.text(x, y, text, color=color, va='bottom', ha='center', fontsize=8,
                bbox=dict(facecolor='white', edgecolor="black", boxstyle=boxstyle, pad=pad))
    add_text_with_box(ax, 19, 112, "    G    R    A    V    A     ", color='black', pad=0.3)
    add_text_with_box(ax, 0.60, 112, '      A         R         E         N         A      ', color='black', pad=0.3)
    add_text_with_box(ax, 0.025, 112, '   F   I   N   O   S   ', color='black', pad=0.3)
    plt.grid(True)
    buffer = BytesIO()
    plt.savefig(buffer, format='png', dpi=100, bbox_inches="tight", pad_inches=0.5)
    buffer.seek(0)
    plt.close()
    image_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
    return image_base64

def generar_grafico_granulometria_prospeccion(df):
    df['id_prospeccion'] = df['id_prospeccion'].astype(str)
    df['id_prospeccion_unico'] = df.groupby('id_prospeccion').cumcount() + 1
    df['etiqueta'] = df['id_prospeccion'] + '-' + df['id_prospeccion_unico'].astype(str)
    prospecciones_unicas = df['etiqueta'].unique()
    colores = {prospeccion: np.random.rand(3,) for prospeccion in prospecciones_unicas}
    fig, ax = plt.subplots()
    x_values = np.array([0.075, 0.110, 0.250, 0.420, 0.840, 2.000, 4.760, 9.520, 19.000, 25.400, 38.100, 50.800, 63.500, 75.000])
    for prospeccion in prospecciones_unicas:
        datos_prospeccion = df[df['etiqueta'] == prospeccion]
        y_values = datos_prospeccion.iloc[0, 2:16].values
        sorted_indices = np.argsort(x_values)[::-1]
        x_values_sorted = x_values[sorted_indices]
        y_values_sorted = y_values[sorted_indices]
        ax.plot(x_values_sorted, y_values_sorted, color=colores[prospeccion], label=prospeccion)
    ax.set_xscale('log')
    ax.set_xlim(0.01, 100)
    ax.xaxis.set_major_locator(LogLocator(base=10, numticks=10))
    ax.xaxis.set_minor_locator(LogLocator(base=10, subs='auto', numticks=50))
    def log_format(x, _):
        if x < 1:
            return f"{x:.2f}"
        else:
            return f"{int(x)}"
    ax.xaxis.set_major_formatter(FuncFormatter(log_format))
    ax.invert_xaxis()
    ax.set_xlabel('Tamaño de partícula (mm)')
    ax.set_ylabel('Porcentaje que pasa (%)')
    ax.set_title('Distribución Granulométrica por Prospección', pad=40)
    handles, labels = ax.get_legend_handles_labels()
    max_legends = 27
    if len(handles) > max_legends:
        handles = handles[:max_legends]
        labels = labels[:max_legends]
    ax.legend(handles, labels, loc='upper right', fontsize=8, ncol=1, columnspacing=0.5,
              handlelength=2, handletextpad=0.5, borderpad=0.5, framealpha=0.5, borderaxespad=0.5)
    ax.axvline(x=76.8, color='black', linestyle='--', linewidth=1, label='3')
    ax.axvline(x=19, color='black', linestyle='--', linewidth=1, label='3/4')
    ax.axvline(x=4.76, color='black', linestyle='--', linewidth=1, label='N°4')
    ax.axvline(x=0.42, color='black', linestyle='--', linewidth=1, label='N°40')
    ax.axvline(x=0.0749, color='black', linestyle='--', linewidth=1, label='N°200')
    ax.axhline(y=50, color='black', linestyle='--', linewidth=1, label='50')
    ax.text(76.8, 106.5, '3', color='black', va='bottom', ha='center', fontsize=9, bbox=dict(facecolor='white', edgecolor='none', alpha=0))
    ax.text(19, 106.5, '3/4', color='black', va='bottom', ha='center', fontsize=9, bbox=dict(facecolor='white', edgecolor='none', alpha=0))
    ax.text(4.76, 106.5, '4', color='black', va='bottom', ha='center', fontsize=9, bbox=dict(facecolor='white', edgecolor='none', alpha=0))
    ax.text(0.42, 106.5, '40', color='black', va='bottom', ha='center', fontsize=9, bbox=dict(facecolor='white', edgecolor='none', alpha=0))
    ax.text(0.0749, 106.5, '200', color='black', va='bottom', ha='center', fontsize=9, bbox=dict(facecolor='white', edgecolor='none', alpha=0))
    ax.text(140, 50, '50', color='black', va='bottom', ha='left', fontsize=9, bbox=dict(facecolor='white', edgecolor='none', alpha=0))
    def add_text_with_box(ax, x, y, text, color, pad=0.5, boxstyle='square'):
        ax.text(x, y, text, color=color, va='bottom', ha='center', fontsize=8,
                bbox=dict(facecolor='white', edgecolor="black", boxstyle=boxstyle, pad=pad))
    add_text_with_box(ax, 19, 112, "    G    R    A    V    A     ", color='black', pad=0.3)
    add_text_with_box(ax, 0.60, 112, '      A         R         E         N         A      ', color='black', pad=0.3)
    add_text_with_box(ax, 0.025, 112, '   F   I   N   O   S   ', color='black', pad=0.3)
    plt.grid(True)
    buffer = BytesIO()
    plt.savefig(buffer, format='png', dpi=100, bbox_inches="tight", pad_inches=0.5)
    buffer.seek(0)
    plt.close()
    image_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
    return image_base64

@login_required
def graficos_granulometria(request):
    id_proyectos = request.GET.getlist('id_proyectos')
    tipos_prospeccion = request.GET.getlist('tipos_prospeccion')
    areas = request.GET.getlist('areas')
    id_prospecciones = request.GET.getlist('id_prospecciones')

    query = Granulometria.objects.all()
    if id_proyectos:
        query = query.filter(id_proyecto__in=id_proyectos)
    if tipos_prospeccion:
        query = query.filter(tipo_prospeccion__in=tipos_prospeccion)
    if areas:
        query = query.filter(area__in=areas)
    if id_prospecciones:
        query = query.filter(id_prospeccion__in=id_prospecciones)

    proyectos = query.values('id_proyecto').distinct()
    tipos_prospeccion_inicial = query.values_list('tipo_prospeccion', flat=True).distinct().exclude(tipo_prospeccion__isnull=True)
    areas_inicial = query.values_list('area', flat=True).distinct().exclude(area__isnull=True).exclude(area='')
    id_prospecciones_inicial = query.values_list('id_prospeccion', flat=True).distinct().exclude(id_prospeccion__isnull=True)

    df = pd.DataFrame(list(query.values(
        'id_proyecto', 'id_prospeccion', 'n_0075', 'n_0110', 'n_0250', 'n_0420', 'n_0840',
        'n_2000', 'n_4760', 'n_9520', 'n_19000', 'n_25400', 'n_38100', 'n_50800', 'n_63500', 'n_75000'
    )))

    context = {
        'proyectos': proyectos,
        'tipos_prospeccion_inicial': tipos_prospeccion_inicial,
        'areas_inicial': areas_inicial,
        'id_prospecciones_inicial': id_prospecciones_inicial,
        'selected_id_proyectos': json.dumps(id_proyectos),
        'selected_tipos_prospeccion': json.dumps(tipos_prospeccion),
        'selected_areas': json.dumps(areas),
        'selected_id_prospecciones': json.dumps(id_prospecciones),
    }

    if df.empty:
        context['error'] = "No hay datos disponibles para los filtros seleccionados."
    else:
        context['image_base64_area'] = generar_grafico_granulometria_area(df)
        context['image_base64_prospeccion'] = generar_grafico_granulometria_prospeccion(df)

    return render(request, 'laboratorio/ensayos/granulometria/graficos_granulometria.html', context)

# Funciones auxiliares para Gráficos
def obtener_tipos_prospeccion_granulometria(request):
    id_proyectos = request.GET.get('id_proyectos')
    query = Granulometria.objects.all()
    if id_proyectos:
        query = query.filter(id_proyecto__in=id_proyectos.split(','))
    tipos_list = list(query.values_list('tipo_prospeccion', flat=True).distinct().exclude(tipo_prospeccion__isnull=True))
    return JsonResponse({'options': tipos_list}, safe=False)

def obtener_id_prospecciones_granulometria(request):
    id_proyectos = request.GET.get('id_proyectos')
    tipos_prospeccion = request.GET.get('tipos_prospeccion')
    areas = request.GET.get('areas')
    query = Granulometria.objects.all()
    if id_proyectos:
        query = query.filter(id_proyecto__in=id_proyectos.split(','))
    if tipos_prospeccion:
        query = query.filter(tipo_prospeccion__in=tipos_prospeccion.split(','))
    if areas:
        query = query.filter(area__in=areas.split(','))
    id_prospecciones_list = list(query.values_list('id_prospeccion', flat=True).distinct().exclude(id_prospeccion__isnull=True))
    return JsonResponse({'options': id_prospecciones_list}, safe=False)

def obtener_area_granulometria(request):
    id_prospecciones = request.GET.get('id_prospecciones')
    if id_prospecciones:
        query = Granulometria.objects.filter(id_prospeccion__in=id_prospecciones.split(','))
        areas_list = list(query.values_list('area', flat=True).distinct().exclude(area__isnull=True).exclude(area=''))
        return JsonResponse({'options': areas_list})
    return JsonResponse({'options': []})

#####CBR############################################################
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.db.models import Q
from django.core.paginator import Paginator
from django.apps import apps
from administrador.models import Cbr, Proyectos, Prospecciones, Muestreo
from .forms import CbrForm
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
import logging
import pandas as pd
import matplotlib.pyplot as plt
from io import BytesIO
import base64
import random
import json

logger = logging.getLogger(__name__)

# 1. Agregar CBR
@login_required
def agregar_cbr(request):
    proyectos_con_muestreo = Proyectos.objects.filter(muestreo__isnull=False).distinct()
    tipo_prospeccion_choices = []
    prospecciones = []
    muestras = []
    area = ""

    if request.method == 'POST':
        logger.info(f"Datos POST recibidos: {request.POST}")
        proyecto_id = request.POST.get('id_proyecto')
        tipo_prospeccion = request.POST.get('tipo_prospeccion')
        id_prospecciones = request.POST.getlist('id_prospeccion[]')
        id_muestras = request.POST.getlist('id_muestra[]')
        densidades_seca_ai = request.POST.getlist('densidad_seca_ai[]')
        densidades_seca_di = request.POST.getlist('densidad_seca_di[]')
        humedades_ai = request.POST.getlist('humedad_ai[]')
        humedades_di = request.POST.getlist('humedad_di[]')
        cbr_01s = request.POST.getlist('cbr_01[]')
        cbr_02s = request.POST.getlist('cbr_02[]')
        observaciones = request.POST.getlist('observacion[]')
        areas = request.POST.getlist('area[]')
        errors = []

        if not proyecto_id:
            errors.append("Debe seleccionar un proyecto.")
            return render(request, 'laboratorio/ensayos/cbr/agregar_cbr.html', {
                'proyectos': proyectos_con_muestreo,
                'errors': errors
            })

        muestreos = Muestreo.objects.filter(id_proyecto_id=proyecto_id)
        if not muestreos.exists():
            errors.append("No hay muestreos asociados a este proyecto.")
            return render(request, 'laboratorio/ensayos/cbr/agregar_cbr.html', {
                'proyectos': proyectos_con_muestreo,
                'errors': errors
            })

        if tipo_prospeccion:
            muestreos = muestreos.filter(tipo_prospeccion=tipo_prospeccion)
            if not muestreos.exists():
                errors.append("No hay muestreos con este tipo de prospección.")

        if id_prospecciones and id_prospecciones[0]:
            muestreos = muestreos.filter(id_prospeccion__id_prospeccion__in=id_prospecciones)
            if not muestreos.exists():
                errors.append("No hay muestreos con las prospecciones seleccionadas.")

        if id_muestras and id_muestras[0]:
            muestreos = muestreos.filter(id_muestra__in=id_muestras)
            if not muestreos.exists():
                errors.append("No hay muestreos con las muestras seleccionadas.")

        if not (len(id_prospecciones) == len(id_muestras) == len(densidades_seca_ai) == len(densidades_seca_di) == len(humedades_ai) == len(humedades_di) == len(cbr_01s) == len(cbr_02s) == len(observaciones) == len(areas)):
            errors.append("El número de campos no coincide.")
            logger.error(f"Longitudes inconsistentes: {len(id_prospecciones)}, {len(id_muestras)}, {len(densidades_seca_ai)}, {len(densidades_seca_di)}, {len(humedades_ai)}, {len(humedades_di)}, {len(cbr_01s)}, {len(cbr_02s)}, {len(observaciones)}, {len(areas)}")
        else:
            for id_pros, id_muestra, ds_ai, ds_di, h_ai, h_di, cbr_01, cbr_02, obs, area_val in zip(id_prospecciones, id_muestras, densidades_seca_ai, densidades_seca_di, humedades_ai, humedades_di, cbr_01s, cbr_02s, observaciones, areas):
                if not id_pros or not id_muestra or not obs:
                    errors.append(f"Falta ID de prospección ({id_pros}), ID de muestra ({id_muestra}) o observación ({obs}).")
                    continue
                
                try:
                    muestreo = muestreos.get(id_prospeccion__id_prospeccion=id_pros, id_muestra=id_muestra)
                    
                    # Cálculo de profundidad_promedio considerando 0 como válido
                    profundidad_desde = float(muestreo.profundidad_desde or 0)  # Si es None, usa 0
                    profundidad_hasta = float(muestreo.profundidad_hasta or 0)  # Si es None, usa 0
                    profundidad_promedio = (profundidad_desde + profundidad_hasta) / 2

                    cbr_obj = Cbr(
                        id_proyecto=muestreo.id_proyecto,
                        tipo_prospeccion=muestreo.tipo_prospeccion,
                        id_prospeccion=muestreo.id_prospeccion,
                        id_muestra=muestreo.id_muestra,
                        profundidad_desde=muestreo.profundidad_desde,
                        profundidad_hasta=muestreo.profundidad_hasta,
                        profundidad_promedio=profundidad_promedio,
                        densidad_seca_ai=float(ds_ai) if ds_ai else None,
                        densidad_seca_di=float(ds_di) if ds_di else None,
                        humedad_ai=float(h_ai) if h_ai else None,
                        humedad_di=float(h_di) if h_di else None,
                        cbr_01=float(cbr_01) if cbr_01 else None,
                        cbr_02=float(cbr_02) if cbr_02 else None,
                        observacion=obs,
                        area=area_val,
                        user=request.user
                    )
                    cbr_obj.save()
                    logger.info(f"CBR guardado: ID={cbr_obj.id}, Muestra={id_muestra}, CBR 0.1={cbr_01}, CBR 0.2={cbr_02}, Profundidad Promedio={profundidad_promedio}")
                except Muestreo.DoesNotExist:
                    errors.append(f"Muestra {id_muestra} con prospección {id_pros} no encontrada.")
                except ValueError as e:
                    errors.append(f"Valor inválido en los campos numéricos: {e}")
                except Exception as e:
                    errors.append(f"Error al guardar muestra {id_muestra}: {e}")

        if errors:
            tipo_prospeccion_choices = Muestreo.objects.filter(id_proyecto_id=proyecto_id).values_list('tipo_prospeccion', flat=True).distinct()
            prospecciones = Muestreo.objects.filter(id_proyecto_id=proyecto_id).values_list('id_prospeccion__id_prospeccion', flat=True).distinct()
            muestras = Muestreo.objects.filter(id_proyecto_id=proyecto_id).values_list('id_muestra', flat=True).distinct()
            return render(request, 'laboratorio/ensayos/cbr/agregar_cbr.html', {
                'proyectos': proyectos_con_muestreo,
                'tipo_prospeccion_choices': tipo_prospeccion_choices,
                'prospecciones': prospecciones,
                'muestras': muestras,
                'area': area,
                'errors': errors
            })

        return redirect('listar_cbr')

    proyecto_id = request.GET.get('id_proyecto')
    tipo_prospeccion = request.GET.get('tipo_prospeccion')
    id_prospeccion = request.GET.get('id_prospeccion')
    id_muestra = request.GET.get('id_muestra')

    muestreos = Muestreo.objects.all()
    if proyecto_id:
        muestreos = muestreos.filter(id_proyecto_id=proyecto_id)
        tipo_prospeccion_choices = muestreos.values_list('tipo_prospeccion', flat=True).distinct()
        prospecciones = muestreos.values_list('id_prospeccion__id_prospeccion', flat=True).distinct()
        muestras = muestreos.values_list('id_muestra', flat=True).distinct()
        area = muestreos.first().area if muestreos.exists() else ""

    if tipo_prospeccion:
        muestreos = muestreos.filter(tipo_prospeccion=tipo_prospeccion)
        prospecciones = muestreos.values_list('id_prospeccion__id_prospeccion', flat=True).distinct()
        muestras = muestreos.values_list('id_muestra', flat=True).distinct()

    if id_prospeccion:
        muestreos = muestreos.filter(id_prospeccion__id_prospeccion=id_prospeccion)
        muestras = muestreos.values_list('id_muestra', flat=True).distinct()
        area = muestreos.first().area if muestreos.exists() else ""

    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if is_ajax:
        data = {
            'tipo_prospeccion_choices': list(tipo_prospeccion_choices),
            'prospecciones': list(prospecciones),
            'muestras': list(muestras),
            'area': area
        }
        if id_muestra:
            muestreo = muestreos.filter(id_muestra=id_muestra).first()
            if muestreo:
                data['profundidad_desde'] = str(muestreo.profundidad_desde) if muestreo.profundidad_desde is not None else '0'
                data['profundidad_hasta'] = str(muestreo.profundidad_hasta) if muestreo.profundidad_hasta is not None else '0'
        return JsonResponse(data)

    return render(request, 'laboratorio/ensayos/cbr/agregar_cbr.html', {
        'proyectos': proyectos_con_muestreo,
        'tipo_prospeccion_choices': tipo_prospeccion_choices,
        'prospecciones': prospecciones,
        'muestras': muestras,
        'area': area
    })

# 2. Listar CBR
@login_required
def listar_cbr(request):
    cbr_list = Cbr.objects.all().order_by('id')
    query = request.GET.get('q', '')

    if query:
        cbr_model = apps.get_model('administrador', 'Cbr')
        campos = [field.name for field in cbr_model._meta.get_fields() if isinstance(field, (models.CharField, models.TextField, models.DecimalField, models.FloatField, models.ForeignKey))]
        query_filter = Q()

        try:
            query_numeric = float(query)
            numeric_search = True
        except ValueError:
            numeric_search = False

        for field in campos:
            field_instance = cbr_model._meta.get_field(field)
            if isinstance(field_instance, models.ForeignKey):
                related_model = field_instance.related_model
                related_fields = [f"{field}__{related_field.name}" for related_field in related_model._meta.get_fields() if isinstance(related_field, (models.CharField, models.TextField))]
                for related_field in related_fields:
                    query_filter |= Q(**{f"{related_field}__icontains": query})
            elif isinstance(field_instance, (models.DecimalField, models.FloatField)) and numeric_search:
                query_filter |= Q(**{f"{field}": query_numeric})
            else:
                query_filter |= Q(**{f"{field}__icontains": query})

        cbr_list = cbr_list.filter(query_filter)

    paginator = Paginator(cbr_list, 1000)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'laboratorio/ensayos/cbr/listar_cbr.html', {
        'page_obj': page_obj,
        'query': query
    })

# 3. Editar CBR
@login_required
def editar_cbr(request, id):
    cbr_obj = get_object_or_404(Cbr, id=id)
    proyectos = Proyectos.objects.filter(muestreo__isnull=False).distinct()

    if request.method == 'POST':
        form = CbrForm(request.POST, instance=cbr_obj)
        if form.is_valid():
            try:
                cbr_obj = form.save()
                logger.info(f"CBR editado: ID {cbr_obj.id}")
                return redirect('listar_cbr')
            except Exception as e:
                logger.error(f"Error al editar: {e}")
                form.add_error(None, f"Error al editar: {e}")
    else:
        form = CbrForm(instance=cbr_obj)

    return render(request, 'laboratorio/ensayos/cbr/editar_cbr.html', {
        'form': form,
        'cbr_obj': cbr_obj,
        'proyectos': proyectos
    })

# 4. Ver CBR
@login_required
def ver_cbr(request, id):
    cbr_obj = get_object_or_404(Cbr, id=id)

    history_records = []
    for record in cbr_obj.history.all():
        changes = []
        previous = record.prev_record
        if previous:
            for field in record.history_object._meta.fields:
                old_value = getattr(previous, field.name, None)
                new_value = getattr(record, field.name, None)
                if old_value != new_value:
                    changes.append({
                        'field_name': field.verbose_name,
                        'old_value': old_value if old_value is not None else '-',
                        'new_value': new_value if new_value is not None else '-',
                    })
        history_records.append({
            'date': record.history_date,
            'user': record.history_user,
            'type': record.history_type,
            'reason': record.history_change_reason,
            'changes': changes if changes else None,
        })

    return render(request, 'laboratorio/ensayos/cbr/ver_cbr.html', {
        'cbr': cbr_obj,
        'history_records': history_records,
    })

# 5. Eliminar CBR
@login_required
def eliminar_cbr(request, id):
    cbr_obj = get_object_or_404(Cbr, id=id)
    if request.method == 'POST':
        cbr_obj.delete()
        return redirect('listar_cbr')
    return render(request, 'laboratorio/ensayos/cbr/eliminar_cbr.html', {
        'cbr': cbr_obj
    })

# 6. Exportar a Excel
@login_required
def export_to_excel_cbr(request):
    query = request.GET.get('q', '')
    headers = ['ID', 'ID Proyecto', 'ID Prospección', 'Tipo Prospección', 'ID Muestra',
               'Profundidad Desde', 'Profundidad Hasta', 'Profundidad Promedio',
               'Densidad Seca AI', 'Densidad Seca DI', 'Humedad AI', 'Humedad DI',
               'CBR 0.1', 'CBR 0.2', 'Observación', 'Área', 'Usuario']

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="cbr_filtrado.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "CBR"
    ws.append(headers)

    cbr_list = Cbr.objects.select_related('id_proyecto', 'id_prospeccion', 'user').all()
    if query:
        filtros = Q()
        try:
            query_numeric = float(query)
            filtros |= (
                Q(profundidad_desde=query_numeric) |
                Q(profundidad_hasta=query_numeric) |
                Q(profundidad_promedio=query_numeric) |
                Q(densidad_seca_ai=query_numeric) |
                Q(densidad_seca_di=query_numeric) |
                Q(humedad_ai=query_numeric) |
                Q(humedad_di=query_numeric) |
                Q(cbr_01=query_numeric) |
                Q(cbr_02=query_numeric)
            )
        except ValueError:
            pass

        filtros |= (
            Q(id__icontains=query) |
            Q(id_proyecto__id__icontains=query) |
            Q(id_prospeccion__id_prospeccion__icontains=query) |
            Q(tipo_prospeccion__icontains=query) |
            Q(id_muestra__icontains=query) |
            Q(observacion__icontains=query) |
            Q(area__icontains=query) |
            Q(user__username__icontains=query)
        )
        cbr_list = cbr_list.filter(filtros)

    for cbr in cbr_list:
        row = [
            str(cbr.id),
            str(cbr.id_proyecto.id) if cbr.id_proyecto else '',
            str(cbr.id_prospeccion.id_prospeccion) if cbr.id_prospeccion else '',
            str(cbr.tipo_prospeccion or ''),
            str(cbr.id_muestra or ''),
            str(cbr.profundidad_desde) if cbr.profundidad_desde else '',
            str(cbr.profundidad_hasta) if cbr.profundidad_hasta else '',
            str(cbr.profundidad_promedio) if cbr.profundidad_promedio else '',
            str(cbr.densidad_seca_ai) if cbr.densidad_seca_ai else '',
            str(cbr.densidad_seca_di) if cbr.densidad_seca_di else '',
            str(cbr.humedad_ai) if cbr.humedad_ai else '',
            str(cbr.humedad_di) if cbr.humedad_di else '',
            str(cbr.cbr_01) if cbr.cbr_01 else '',
            str(cbr.cbr_02) if cbr.cbr_02 else '',
            str(cbr.observacion or ''),
            str(cbr.area or ''),
            str(cbr.user.username) if cbr.user else 'Sin usuario'
        ]
        ws.append(row)

    wb.save(response)
    return response

# 7. Exportar a PDF
@login_required
def export_to_pdf_cbr(request):
    query = request.GET.get('q', '')
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="cbr_filtrado.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(letter), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()

    title = Paragraph("Lista de CBR", styles['Heading1'].clone('Title', alignment=1, fontSize=12, spaceAfter=10))
    elements.append(title)
    elements.append(Paragraph("<br/>", styles['Normal']))

    headers = ['ID', 'Proyecto', 'Prospección', 'Tipo', 'Muestra', 'Prof. Ini', 'Prof. Fin', 'Prof. Prom',
               'Dens. Seca AI', 'Dens. Seca DI', 'Hum. AI', 'Hum. DI', 'CBR 0.1', 'CBR 0.2', 'Obs.', 'Área', 'Usuario']
    data = [headers]

    cbr_list = Cbr.objects.select_related('id_proyecto', 'id_prospeccion', 'user').all()
    if query:
        filtros = Q()
        try:
            query_numeric = float(query)
            filtros |= (
                Q(profundidad_desde=query_numeric) |
                Q(profundidad_hasta=query_numeric) |
                Q(profundidad_promedio=query_numeric) |
                Q(densidad_seca_ai=query_numeric) |
                Q(densidad_seca_di=query_numeric) |
                Q(humedad_ai=query_numeric) |
                Q(humedad_di=query_numeric) |
                Q(cbr_01=query_numeric) |
                Q(cbr_02=query_numeric)
            )
        except ValueError:
            pass

        filtros |= (
            Q(id__icontains=query) |
            Q(id_proyecto__id__icontains=query) |
            Q(id_prospeccion__id_prospeccion__icontains=query) |
            Q(tipo_prospeccion__icontains=query) |
            Q(id_muestra__icontains=query) |
            Q(observacion__icontains=query) |
            Q(area__icontains=query) |
            Q(user__username__icontains=query)
        )
        cbr_list = cbr_list.filter(filtros)

    for cbr in cbr_list:
        row = [
            str(cbr.id),
            str(cbr.id_proyecto.id) if cbr.id_proyecto else '',
            str(cbr.id_prospeccion.id_prospeccion) if cbr.id_prospeccion else '',
            str(cbr.tipo_prospeccion or ''),
            str(cbr.id_muestra or ''),
            str(cbr.profundidad_desde) if cbr.profundidad_desde else '',
            str(cbr.profundidad_hasta) if cbr.profundidad_hasta else '',
            str(cbr.profundidad_promedio) if cbr.profundidad_promedio else '',
            str(cbr.densidad_seca_ai) if cbr.densidad_seca_ai else '',
            str(cbr.densidad_seca_di) if cbr.densidad_seca_di else '',
            str(cbr.humedad_ai) if cbr.humedad_ai else '',
            str(cbr.humedad_di) if cbr.humedad_di else '',
            str(cbr.cbr_01) if cbr.cbr_01 else '',
            str(cbr.cbr_02) if cbr.cbr_02 else '',
            str(cbr.observacion or ''),
            str(cbr.area or ''),
            str(cbr.user.username) if cbr.user else 'Sin usuario'
        ]
        data.append(row)

    table = Table(data, colWidths=[0.5*inch] + [0.8*inch]*16)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(table)
    doc.build(elements)
    return response

# 8. Gráficos CBR
import random
import matplotlib.pyplot as plt
from io import BytesIO
import base64
import pandas as pd

# Función para gráfico agrupado por área
def generar_grafico_cbr_area(df):
    areas_unicas = df['area'].unique()
    colores = {area: (random.random(), random.random(), random.random()) for area in areas_unicas}
    fig, ax = plt.subplots()
    
    for area in areas_unicas:
        datos_area = df[df['area'] == area]
        ax.scatter(
            datos_area['cbr_01'],
            datos_area['profundidad_promedio'],
            color=colores[area],
            s=100,
            label=f"{area} CBR 0.1",
            marker='+'
        )
        ax.scatter(
            datos_area['cbr_02'],
            datos_area['profundidad_promedio'],
            color=colores[area],
            s=100,
            label=f"{area} CBR 0.2",
            marker='o',
            alpha=0.5
        )
    
    ax.xaxis.set_ticks_position("top")
    ax.xaxis.set_label_position("top")
    ax.invert_yaxis()
    ax.set_xlabel("CBR (%)")
    ax.set_ylabel("Profundidad (m)")
    ax.grid(True)
    plt.subplots_adjust(right=0.75)
    num_columns = (len(areas_unicas) * 2 + 24) // 25  # Multiplicamos por 2 por las dos etiquetas por área
    legend = ax.legend(loc='center left', bbox_to_anchor=(1, 0.5), ncol=num_columns)
    plt.title("GRÁFICO AGRUPADO POR ÁREA (DISPERSIÓN)", pad=30)
    
    buffer = BytesIO()
    fig.savefig(buffer, format='png', dpi=100, bbox_inches='tight', bbox_extra_artists=[legend])
    buffer.seek(0)
    plt.close()
    return base64.b64encode(buffer.getvalue()).decode('utf-8')

# Función para gráfico agrupado por prospección
def generar_grafico_cbr_prospeccion(df):
    df['id_prospeccion'] = df['id_prospeccion'].astype(str)
    df['id_prospeccion_unico'] = df.groupby('id_prospeccion').cumcount() + 1
    df['etiqueta'] = df['id_prospeccion'] + '-' + df['id_prospeccion_unico'].astype(str)
    prospecciones_unicas = df['etiqueta'].unique()
    colores = {prospeccion: (random.random(), random.random(), random.random()) for prospeccion in prospecciones_unicas}
    
    fig, ax = plt.subplots()
    
    for prospeccion in prospecciones_unicas:
        datos_prospeccion = df[df['etiqueta'] == prospeccion]
        ax.scatter(
            datos_prospeccion['cbr_01'],
            datos_prospeccion['profundidad_promedio'],
            color=colores[prospeccion],
            s=100,
            label=f"{prospeccion} CBR 0.1",
            marker='+'
        )
        ax.scatter(
            datos_prospeccion['cbr_02'],
            datos_prospeccion['profundidad_promedio'],
            color=colores[prospeccion],
            s=100,
            label=f"{prospeccion} CBR 0.2",
            marker='o',
            alpha=0.5
        )
    
    ax.xaxis.set_ticks_position("top")
    ax.xaxis.set_label_position("top")
    ax.invert_yaxis()
    ax.set_xlabel("CBR (%)")
    ax.set_ylabel("Profundidad (m)")
    ax.grid(True)
    plt.subplots_adjust(right=0.75)
    num_columns = (len(prospecciones_unicas) * 2 + 24) // 25  # Multiplicamos por 2 por las dos etiquetas por prospección
    legend = ax.legend(loc='center left', bbox_to_anchor=(1, 0.5), ncol=num_columns)
    plt.title("GRÁFICO AGRUPADO POR PROSPECCIÓN (DISPERSIÓN)", pad=30)
    
    buffer = BytesIO()
    fig.savefig(buffer, format='png', dpi=100, bbox_inches='tight', bbox_extra_artists=[legend])
    buffer.seek(0)
    plt.close()
    return base64.b64encode(buffer.getvalue()).decode('utf-8')

# Vista principal para gráficos CBR
@login_required
def graficos_cbr(request):
    id_proyectos = request.GET.getlist('id_proyectos')
    tipos_prospeccion = request.GET.getlist('tipos_prospeccion')
    areas = request.GET.getlist('areas')
    id_prospecciones = request.GET.getlist('id_prospecciones')

    query = Cbr.objects.all()
    if id_proyectos:
        query = query.filter(id_proyecto__in=id_proyectos)
    if tipos_prospeccion:
        query = query.filter(tipo_prospeccion__in=tipos_prospeccion)
    if areas:
        query = query.filter(area__in=areas)
    if id_prospecciones:
        query = query.filter(id_prospeccion__in=id_prospecciones)

    proyectos = query.values('id_proyecto').distinct()
    tipos_prospeccion_inicial = query.values_list('tipo_prospeccion', flat=True).distinct().exclude(tipo_prospeccion__isnull=True)
    areas_inicial = query.values_list('area', flat=True).distinct().exclude(area__isnull=True).exclude(area='')
    id_prospecciones_inicial = query.values_list('id_prospeccion', flat=True).distinct().exclude(id_prospeccion__isnull=True)

    df = pd.DataFrame(list(query.values('cbr_01', 'cbr_02', 'profundidad_promedio', 'area', 'id_prospeccion_id')))
    df = df.rename(columns={'id_prospeccion_id': 'id_prospeccion'})

    context = {
        'proyectos': proyectos,
        'tipos_prospeccion_inicial': tipos_prospeccion_inicial,
        'areas_inicial': areas_inicial,
        'id_prospecciones_inicial': id_prospecciones_inicial,
        'selected_id_proyectos': json.dumps(id_proyectos),
        'selected_tipos_prospeccion': json.dumps(tipos_prospeccion),
        'selected_areas': json.dumps(areas),
        'selected_id_prospecciones': json.dumps(id_prospecciones),
    }

    if df.empty:
        context['error'] = "No hay datos disponibles para los filtros seleccionados."
    else:
        context['image_base64_area'] = generar_grafico_cbr_area(df)
        context['image_base64_prospeccion'] = generar_grafico_cbr_prospeccion(df)

    return render(request, 'laboratorio/ensayos/cbr/graficos_cbr.html', context)
# Funciones auxiliares para gráficos
def obtener_tipos_prospeccion_cbr(request):
    id_proyectos = request.GET.get('id_proyectos')
    query = Cbr.objects.all()
    if id_proyectos:
        query = query.filter(id_proyecto__in=id_proyectos.split(','))
    tipos_list = list(query.values_list('tipo_prospeccion', flat=True).distinct().exclude(tipo_prospeccion__isnull=True))
    return JsonResponse({'options': tipos_list}, safe=False)

def obtener_id_prospecciones_cbr(request):
    id_proyectos = request.GET.get('id_proyectos')
    tipos_prospeccion = request.GET.get('tipos_prospeccion')
    areas = request.GET.get('areas')
    query = Cbr.objects.all()
    if id_proyectos:
        query = query.filter(id_proyecto__in=id_proyectos.split(','))
    if tipos_prospeccion:
        query = query.filter(tipo_prospeccion__in=tipos_prospeccion.split(','))
    if areas:
        query = query.filter(area__in=areas.split(','))
    id_prospecciones_list = list(query.values_list('id_prospeccion', flat=True).distinct().exclude(id_prospeccion__isnull=True))
    return JsonResponse({'options': id_prospecciones_list}, safe=False)

#### MUESTREO ##############################################
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from .models import Muestreo, Proyectos, Prospecciones, MuestreoImage
import logging

logger = logging.getLogger(__name__)
# views.py
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from .models import Muestreo, Proyectos, Prospecciones, MuestreoImage
import logging

logger = logging.getLogger(__name__)

@login_required
def agregar_muestreo(request):
    proyectos_con_prospeccion = Proyectos.objects.filter(prospecciones__isnull=False).distinct()

    if request.method == 'POST':
        logger.info(f"Datos POST recibidos: {request.POST}")
        logger.info(f"Archivos recibidos: {request.FILES}")
        
        proyecto_id = request.POST.get('id_proyecto')
        tipo_prospeccion = request.POST.get('tipo_prospeccion')
        id_prospecciones = request.POST.getlist('id_prospeccion[]')
        areas = request.POST.getlist('area[]')
        objetivos = request.POST.getlist('objetivo[]')
        fechas_muestreo = request.POST.getlist('fecha_muestreo[]')
        id_bultos = request.POST.getlist('id_bulto[]')
        tipos_bulto = request.POST.getlist('tipo_bulto[]')
        id_embalajes_muestra = request.POST.getlist('id_embalaje_muestra[]')
        tipos_embalaje_muestra = request.POST.getlist('tipo_embalaje_muestra[]')
        cantidades = request.POST.getlist('cantidad[]')
        pesos_unitarios = request.POST.getlist('peso_unitario[]')
        id_muestras = request.POST.getlist('id_muestra[]')
        profundidades_desde = request.POST.getlist('profundidad_desde[]')
        profundidades_hasta = request.POST.getlist('profundidad_hasta[]')
        estratos = request.POST.getlist('estrato[]')
        tipos = request.POST.getlist('tipo[]')
        fechas_despacho = request.POST.getlist('fecha_despacho[]')
        nombres_despachadores = request.POST.getlist('nombre_despachador[]')
        destinos = request.POST.getlist('destino[]')
        ordenes_transporte = request.POST.getlist('orden_transporte[]')
        observaciones = request.POST.getlist('observacion[]')
        id_laboratorios = request.POST.getlist('id_laboratorio[]')  # Nuevo campo
        images = request.FILES.getlist('image[]')
        errors = []

        if not proyecto_id:
            errors.append("Debe selecting un proyecto.")
        else:
            prospecciones_qs = Prospecciones.objects.filter(id_proyecto_id=proyecto_id)
            if not prospecciones_qs.exists():
                errors.append("No hay prospecciones asociadas a este proyecto.")
            elif tipo_prospeccion:
                prospecciones_qs = prospecciones_qs.filter(tipo_prospeccion=tipo_prospeccion)

            tipo_prospeccion_choices = prospecciones_qs.values_list('tipo_prospeccion', flat=True).distinct()
            prospecciones = prospecciones_qs.values_list('id_prospeccion', flat=True).distinct()

            # Calcular el área inicial basada en la primera prospección seleccionada (si existe)
            area = ""
            if id_prospecciones and id_prospecciones[0]:
                prospeccion = prospecciones_qs.filter(id_prospeccion=id_prospecciones[0]).first()
                area = prospeccion.area if prospeccion else ""

            if not errors and len(id_prospecciones) > 0:
                fields_length = len(id_prospecciones)
                all_lengths = [
                    fields_length, len(areas), len(objetivos), len(fechas_muestreo), len(id_bultos),
                    len(tipos_bulto), len(id_embalajes_muestra), len(tipos_embalaje_muestra),
                    len(cantidades), len(pesos_unitarios), len(id_muestras), len(profundidades_desde),
                    len(profundidades_hasta), len(estratos), len(tipos), len(fechas_despacho),
                    len(nombres_despachadores), len(destinos), len(ordenes_transporte), len(observaciones),
                    len(id_laboratorios)  # Incluimos id_laboratorio
                ]
                if not all(length == fields_length for length in all_lengths):
                    errors.append("El número de campos no coincide.")
                else:
                    for i in range(fields_length):
                        id_pros = id_prospecciones[i]
                        if not id_pros:
                            errors.append(f"Fila {i+1}: Falta ID de prospección.")
                            continue

                        try:
                            prospeccion = prospecciones_qs.get(id_prospeccion=id_pros)
                            peso_total = float(pesos_unitarios[i]) * float(cantidades[i]) if pesos_unitarios[i] and cantidades[i] else None
                            prof_d = float(profundidades_desde[i]) if profundidades_desde[i] else None
                            prof_h = float(profundidades_hasta[i]) if profundidades_hasta[i] else None
                            profundidad_promedio = (prof_d + prof_h) / 2 if prof_d is not None and prof_h is not None else None
                            espesor_estrato = prof_h - prof_d if prof_d is not None and prof_h is not None and prof_h >= prof_d else None

                            muestreo_obj = Muestreo(
                                id_proyecto=prospeccion.id_proyecto,
                                tipo_prospeccion=tipo_prospeccion if tipo_prospeccion else prospeccion.tipo_prospeccion,
                                id_prospeccion=prospeccion,
                                area=areas[i],
                                objetivo=objetivos[i] if objetivos[i] else None,
                                fecha_muestreo=fechas_muestreo[i] if fechas_muestreo[i] else None,
                                id_bulto=id_bultos[i] if id_bultos[i] else None,
                                tipo_bulto=tipos_bulto[i] if tipos_bulto[i] else None,
                                id_embalaje_muestra=id_embalajes_muestra[i] if id_embalajes_muestra[i] else None,
                                tipo_embalaje_muestra=tipos_embalaje_muestra[i] if tipos_embalaje_muestra[i] else None,
                                cantidad=float(cantidades[i]) if cantidades[i] else None,
                                peso_unitario=float(pesos_unitarios[i]) if pesos_unitarios[i] else None,
                                peso_total=peso_total,
                                id_muestra=id_muestras[i],
                                profundidad_desde=prof_d,
                                profundidad_hasta=prof_h,
                                profundidad_promedio=profundidad_promedio,
                                espesor_estrato=espesor_estrato,
                                estrato=estratos[i] if estratos[i] else None,
                                tipo=tipos[i] if tipos[i] else None,
                                fecha_despacho=fechas_despacho[i] if fechas_despacho[i] else None,
                                nombre_despachador=nombres_despachadores[i] if nombres_despachadores[i] else None,
                                destino=destinos[i] if destinos[i] else None,
                                orden_transporte=ordenes_transporte[i] if ordenes_transporte[i] else None,
                                observacion=observaciones[i] if observaciones[i] else None,
                                id_laboratorio=id_laboratorios[i] if id_laboratorios[i] else None,  # Aceptamos el valor del frontend, pero el modelo lo recalcula si está vacío
                                user=request.user
                            )
                            muestreo_obj.save()
                            logger.info(f"Muestreo guardado: ID={muestreo_obj.id}, ID Muestra={muestreo_obj.id_muestra}")

                            for img in images:
                                imagen_obj = MuestreoImage.objects.create(muestreo=muestreo_obj, image=img)
                                muestreo_obj.imagenes.add(imagen_obj)
                            logger.info(f"Guardadas {len(images)} imágenes para muestreo {muestreo_obj.id_muestra}")

                        except Prospecciones.DoesNotExist:
                            errors.append(f"Fila {i+1}: Prospección {id_pros} no encontrada.")
                        except ValueError as e:
                            errors.append(f"Fila {i+1}: Valor inválido: {e}")
                        except Exception as e:
                            errors.append(f"Fila {i+1}: Error al guardar: {e}")
                            logger.error(f"Error al guardar muestreo o imágenes: {e}")

        if errors:
            area = ""
            if id_prospecciones and id_prospecciones[0]:
                prospeccion = Prospecciones.objects.filter(id_prospeccion=id_prospecciones[0]).first()
                area = prospeccion.area if prospeccion else ""

            return render(request, 'terreno/muestreo/agregar_muestreo.html', {
                'proyectos': proyectos_con_prospeccion,
                'tipo_prospeccion_choices': tipo_prospeccion_choices,
                'prospecciones': prospecciones,
                'area': area,
                'errors': errors,
                'objetivo_choices': Muestreo._meta.get_field('objetivo').choices,
                'tipo_bulto_choices': Muestreo._meta.get_field('tipo_bulto').choices,  # Actualizado
                'tipo_embalaje_muestra_choices': Muestreo._meta.get_field('tipo_embalaje_muestra').choices,  # Nuevo
                'estrato_choices': Muestreo._meta.get_field('estrato').choices,
                'tipo_choices': Muestreo._meta.get_field('tipo').choices,
                'form_data': request.POST,
            })

        return redirect('listar_muestreo')

    # GET: Poblar datos iniciales
    proyecto_id = request.GET.get('id_proyecto')
    tipo_prospeccion = request.GET.get('tipo_prospeccion')
    id_prospeccion = request.GET.get('id_prospeccion')
    prospecciones_qs = Prospecciones.objects.all()
    tipo_prospeccion_choices = []
    prospecciones = []
    area = ""

    if proyecto_id:
        prospecciones_qs = prospecciones_qs.filter(id_proyecto_id=proyecto_id)
        tipo_prospeccion_choices = prospecciones_qs.values_list('tipo_prospeccion', flat=True).distinct()
        prospecciones = prospecciones_qs.values_list('id_prospeccion', flat=True).distinct()

    if tipo_prospeccion:
        prospecciones_qs = prospecciones_qs.filter(tipo_prospeccion=tipo_prospeccion)
        prospecciones = prospecciones_qs.values_list('id_prospeccion', flat=True).distinct()

    if id_prospeccion:
        prospeccion = prospecciones_qs.filter(id_prospeccion=id_prospeccion).first()
        area = prospeccion.area if prospeccion else ""

    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if is_ajax:
        return JsonResponse({
            'tipo_prospeccion_choices': list(tipo_prospeccion_choices),
            'prospecciones': list(prospecciones),
            'area': area,
        })

    return render(request, 'terreno/muestreo/agregar_muestreo.html', {
        'proyectos': proyectos_con_prospeccion,
        'tipo_prospeccion_choices': tipo_prospeccion_choices,
        'prospecciones': prospecciones,
        'area': area,
        'objetivo_choices': Muestreo._meta.get_field('objetivo').choices,
        'tipo_bulto_choices': Muestreo._meta.get_field('tipo_bulto').choices,  # Actualizado
        'tipo_embalaje_muestra_choices': Muestreo._meta.get_field('tipo_embalaje_muestra').choices,  # Nuevo
        'estrato_choices': Muestreo._meta.get_field('estrato').choices,
        'tipo_choices': Muestreo._meta.get_field('tipo').choices,
    })


# 2. Listar Muestreo (sin cambios)
@login_required
def listar_muestreo(request):
    muestreos = Muestreo.objects.all().order_by('id')
    query = request.GET.get('q', '')

    if query:
        muestreo_model = apps.get_model('administrador', 'Muestreo')
        campos = [field.name for field in muestreo_model._meta.get_fields() if isinstance(field, (models.CharField, models.TextField, models.DateField, models.DecimalField, models.ForeignKey))]
        query_filter = Q()
        try:
            query_numeric = float(query)
            numeric_search = True
        except ValueError:
            numeric_search = False

        for field in campos:
            field_instance = muestreo_model._meta.get_field(field)
            if isinstance(field_instance, models.DateField):
                query_filter |= Q(**{f"{field}__year__icontains": query})
                query_filter |= Q(**{f"{field}__month__icontains": query})
                query_filter |= Q(**{f"{field}__day__icontains": query})
            elif isinstance(field_instance, models.ForeignKey):
                related_model = field_instance.related_model
                related_fields = [f"{field}__{related_field.name}" for related_field in related_model._meta.get_fields() if isinstance(related_field, (models.CharField, models.TextField))]
                for related_field in related_fields:
                    query_filter |= Q(**{f"{related_field}__icontains": query})
            elif isinstance(field_instance, models.DecimalField) and numeric_search:
                query_filter |= Q(**{f"{field}": query_numeric})
            else:
                query_filter |= Q(**{f"{field}__icontains": query})
        muestreos = muestreos.filter(query_filter)

    paginator = Paginator(muestreos, 1000)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'terreno/muestreo/listar_muestreo.html', {
        'page_obj': page_obj,
        'query': query
    })

# 3. Editar Muestreo (con manejo de imágenes como en Prospecciones)
@login_required
def editar_muestreo(request, id):
    muestreo_obj = get_object_or_404(Muestreo, id=id)
    proyectos = Proyectos.objects.filter(prospecciones__isnull=False).distinct()

    if request.method == 'POST':
        form = MuestreoForm(request.POST, instance=muestreo_obj)
        if form.is_valid():
            try:
                muestreo_obj = form.save(commit=False)
                muestreo_obj.user = request.user
                muestreo_obj.save()
                logger.info(f"Muestreo editado: ID {muestreo_obj.id}")
                return redirect('listar_muestreo')
            except Exception as e:
                logger.error(f"Error al editar: {e}")
                form.add_error(None, f"Error al editar: {e}")
    else:
        form = MuestreoForm(instance=muestreo_obj)

    return render(request, 'terreno/muestreo/editar_muestreo.html', {
        'form': form,
        'muestreo_obj': muestreo_obj,
        'proyectos': proyectos
    })

# Funciones AJAX para agregar y eliminar imágenes (replicadas de Prospecciones)
@login_required
def agregar_imagen_muestreo(request):
    muestreo_id = request.POST.get('muestreo_id')
    try:
        muestreo = Muestreo.objects.get(id=muestreo_id)
        if 'image' not in request.FILES:
            return JsonResponse({'success': False, 'error': 'No se proporcionó ninguna imagen'})

        imagen = MuestreoImage(image=request.FILES['image'])
        imagen.muestreo = muestreo
        imagen.save()
        muestreo.imagenes.add(imagen)

        logger.info(f"Usuario {request.user} agregó una imagen a muestreo {muestreo.id_muestra}")
        return JsonResponse({
            'success': True,
            'image_id': imagen.id,
            'image_url': imagen.image.url
        })
    except Muestreo.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'El muestreo no existe'})
    except Exception as e:
        logger.error(f"Error al agregar imagen: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def eliminar_imagen_muestreo(request):
    if request.method == 'POST':
        image_id = request.POST.get('image_id')
        image = get_object_or_404(MuestreoImage, id=image_id)
        muestreo = image.muestreo
        muestreo.imagenes.remove(image)
        image.delete()

        logger.info(f"Usuario {request.user} eliminó una imagen de muestreo {muestreo.id_muestra}")
        return JsonResponse({'success': True})
    return JsonResponse({'success': False})

# 4. Ver Muestreo (con imágenes como en Prospecciones)
@login_required
def ver_muestreo(request, id):
    muestreo_obj = get_object_or_404(Muestreo, id=id)

    history_records = []
    for record in muestreo_obj.history.all():
        changes = []
        previous = record.prev_record
        if previous:
            for field in record.history_object._meta.fields:
                old_value = getattr(previous, field.name, None)
                new_value = getattr(record, field.name, None)
                if old_value != new_value:
                    changes.append({
                        'field_name': field.verbose_name,
                        'old_value': old_value if old_value is not None else '-',
                        'new_value': new_value if new_value is not None else '-',
                    })
        history_records.append({
            'date': record.history_date,
            'user': record.history_user,
            'type': record.history_type,
            'reason': record.history_change_reason,
            'changes': changes if changes else None,
        })

    return render(request, 'terreno/muestreo/ver_muestreo.html', {
        'muestreo': muestreo_obj,
        'history_records': history_records,
    })

    
# 5. Eliminar
@login_required
def eliminar_muestreo(request, id):
    muestreo_obj = get_object_or_404(Muestreo, id=id)
    if request.method == 'POST':
        muestreo_obj.delete()
        return redirect('listar_muestreo')
    return render(request, 'terreno/muestreo/eliminar_muestreo.html', {
        'muestreo': muestreo_obj
    })

# 6. Exportar Excel
@login_required
def export_to_excel_muestreo(request):
    query = request.GET.get('q', '').strip()
    headers = ['ID', 'ID Embalaje', 'ID Proyecto', 'Tipo Prospección', 'ID Prospección', 'Área', 'Objetivo', 'Fecha Muestreo', 'Tipo Embalaje', 'Cantidad', 'Peso Unitario', 'Peso Total', 'ID Muestra', 'Profundidad Desde', 'Profundidad Hasta', 'Estrato', 'Tipo', 'Fecha Despacho', 'Nombre Despachador', 'Destino', 'Orden Transporte', 'Observación', 'Usuario']

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="muestreo_filtrado.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Muestreo"
    ws.append(headers)

    muestreo_list = Muestreo.objects.select_related('id_proyecto', 'id_prospeccion', 'user').all()

    if query:
        filtros = Q()
        try:
            query_numeric = float(query)
            filtros |= (
                Q(cantidad=query_numeric) |
                Q(peso_unitario=query_numeric) |
                Q(peso_total=query_numeric) |
                Q(profundidad_desde=query_numeric) |
                Q(profundidad_hasta=query_numeric)
            )
        except ValueError:
            pass

        filtros |= (
            Q(id_embalaje__icontains=query) |
            Q(id_proyecto__id__icontains=query) |
            Q(tipo_prospeccion__icontains=query) |
            Q(id_prospeccion__id_prospeccion__icontains=query) |
            Q(area__icontains=query) |
            Q(objetivo__icontains=query) |
            Q(id_muestra__icontains=query) |
            Q(estrato__icontains=query) |
            Q(tipo__icontains=query) |
            Q(nombre_despachador__icontains=query) |
            Q(destino__icontains=query) |
            Q(orden_transporte__icontains=query) |
            Q(observacion__icontains=query) |
            Q(user__username__icontains=query)
        )
        muestreo_list = muestreo_list.filter(filtros)

    for muestreo in muestreo_list:
        row = [
            str(muestreo.id),
            str(muestreo.id_embalaje),
            str(muestreo.id_proyecto.id) if muestreo.id_proyecto else '',
            str(muestreo.tipo_prospeccion or ''),
            str(muestreo.id_prospeccion.id_prospeccion) if muestreo.id_prospeccion else '',
            str(muestreo.area),
            str(muestreo.objetivo or ''),
            str(muestreo.fecha_muestreo or ''),
            str(muestreo.tipo_embalaje or ''),
            str(muestreo.cantidad or ''),
            str(muestreo.peso_unitario or ''),
            str(muestreo.peso_total or ''),
            str(muestreo.id_muestra),
            str(muestreo.profundidad_desde or ''),
            str(muestreo.profundidad_hasta or ''),
            str(muestreo.estrato or ''),
            str(muestreo.tipo or ''),
            str(muestreo.fecha_despacho or ''),
            str(muestreo.nombre_despachador or ''),
            str(muestreo.destino or ''),
            str(muestreo.orden_transporte or ''),
            str(muestreo.observacion or ''),
            str(muestreo.user.username) if muestreo.user else 'Sin usuario'
        ]
        ws.append(row)

    wb.save(response)
    return response

# 7. Exportar PDF
@login_required
def export_to_pdf_muestreo(request):
    query = request.GET.get('q', '').strip()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="muestreo_filtrado.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(letter), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()

    title = Paragraph("Lista de Muestreo", styles['Heading1'].clone('Title', alignment=1, fontSize=12, spaceAfter=10))
    elements.append(title)
    elements.append(Paragraph("<br/>", styles['Normal']))

    headers = ['ID', 'ID Embalaje', 'Proyecto', 'Tipo', 'Prospección', 'Área', 'Objetivo', 'Fecha Muestreo', 'Tipo Embalaje', 'Cantidad', 'Peso Unit.', 'Peso Total', 'ID Muestra', 'Prof. Desde', 'Prof. Hasta', 'Estrato', 'Tipo', 'Fecha Despacho', 'Despachador', 'Destino', 'Orden Transp.', 'Observación', 'Usuario']
    data = [headers]

    muestreo_list = Muestreo.objects.select_related('id_proyecto', 'id_prospeccion', 'user').all()

    if query:
        filtros = Q()
        try:
            query_numeric = float(query)
            filtros |= (
                Q(cantidad=query_numeric) |
                Q(peso_unitario=query_numeric) |
                Q(peso_total=query_numeric) |
                Q(profundidad_desde=query_numeric) |
                Q(profundidad_hasta=query_numeric)
            )
        except ValueError:
            pass

        filtros |= (
            Q(id_embalaje__icontains=query) |
            Q(id_proyecto__id__icontains=query) |
            Q(tipo_prospeccion__icontains=query) |
            Q(id_prospeccion__id_prospeccion__icontains=query) |
            Q(area__icontains=query) |
            Q(objetivo__icontains=query) |
            Q(id_muestra__icontains=query) |
            Q(estrato__icontains=query) |
            Q(tipo__icontains=query) |
            Q(nombre_despachador__icontains=query) |
            Q(destino__icontains=query) |
            Q(orden_transporte__icontains=query) |
            Q(observacion__icontains=query) |
            Q(user__username__icontains=query)
        )
        muestreo_list = muestreo_list.filter(filtros)

    for muestreo in muestreo_list:
        row = [
            str(muestreo.id),
            str(muestreo.id_embalaje),
            str(muestreo.id_proyecto.id) if muestreo.id_proyecto else '',
            str(muestreo.tipo_prospeccion or ''),
            str(muestreo.id_prospeccion.id_prospeccion) if muestreo.id_prospeccion else '',
            str(muestreo.area),
            str(muestreo.objetivo or ''),
            str(muestreo.fecha_muestreo or ''),
            str(muestreo.tipo_embalaje or ''),
            str(muestreo.cantidad or ''),
            str(muestreo.peso_unitario or ''),
            str(muestreo.peso_total or ''),
            str(muestreo.id_muestra),
            str(muestreo.profundidad_desde or ''),
            str(muestreo.profundidad_hasta or ''),
            str(muestreo.estrato or ''),
            str(muestreo.tipo or ''),
            str(muestreo.fecha_despacho or ''),
            str(muestreo.nombre_despachador or ''),
            str(muestreo.destino or ''),
            str(muestreo.orden_transporte or ''),
            str(muestreo.observacion or ''),
            str(muestreo.user.username) if muestreo.user else 'Sin usuario'
        ]
        data.append(row)

    table = Table(data, colWidths=[0.5*inch] * len(headers))
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(table)
    doc.build(elements)
    return response

# Gráficos Muestreo
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from io import BytesIO
import base64
import json
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from .models import Muestreo

@login_required
def graficos_muestreo(request):
    # Obtener filtros desde la solicitud GET
    id_proyectos = request.GET.getlist('id_proyecto')
    tipos_prospeccion = request.GET.getlist('tipo_prospeccion')
    areas = request.GET.getlist('area')
    id_prospecciones = request.GET.getlist('id_prospeccion')

    # Consulta inicial
    query = Muestreo.objects.all()

    # Aplicar filtros si existen
    if id_proyectos:
        query = query.filter(id_proyecto__in=id_proyectos)
    if tipos_prospeccion:
        query = query.filter(tipo_prospeccion__in=tipos_prospeccion)
    if areas:
        query = query.filter(area__in=areas)
    if id_prospecciones:
        query = query.filter(id_prospeccion__in=id_prospecciones)

    # Obtener opciones iniciales para los selectores
    proyectos = query.values('id_proyecto').distinct()
    tipos_prospeccion_inicial = (query.values_list('tipo_prospeccion', flat=True)
                                .distinct()
                                .exclude(tipo_prospeccion__isnull=True)
                                .exclude(tipo_prospeccion=''))
    areas_inicial = (query.values_list('area', flat=True)
                     .distinct()
                     .exclude(area__isnull=True)
                     .exclude(area=''))
    id_prospecciones_inicial = (query.values_list('id_prospeccion', flat=True)
                                .distinct()
                                .exclude(id_prospeccion__isnull=True))

    # Convertir los datos filtrados a un DataFrame
    df = pd.DataFrame.from_records(query.values('id_proyecto', 'id_prospeccion', 'objetivo', 
                                                'profundidad_promedio', 'espesor_estrato', 
                                                'cantidad', 'area'))

    # Preparar el contexto para la plantilla
    context = {
        'proyectos': proyectos,
        'tipos_prospeccion_inicial': tipos_prospeccion_inicial,
        'areas_inicial': areas_inicial,
        'id_prospecciones_inicial': id_prospecciones_inicial,
        'selected_id_proyectos': json.dumps(id_proyectos),
        'selected_tipos_prospeccion': json.dumps(tipos_prospeccion),
        'selected_areas': json.dumps(areas),
        'selected_id_prospecciones': json.dumps(id_prospecciones),
    }

    # Verificar si hay datos para graficar
    if df.empty:
        context['error'] = "No hay datos para graficar."
        return render(request, 'terreno/muestreo/graficos_muestreo.html', context)

    # Generar los gráficos y añadirlos al contexto
    image_base64_distribucion = generar_grafico_distribucion_cantidades(df)
    image_base64_profundidad_espesor = generar_grafico_profundidad_espesor(df)
    context.update({
        'image_base64_distribucion': image_base64_distribucion,
        'image_base64_profundidad_espesor': image_base64_profundidad_espesor,
    })

    # Renderizar la plantilla con el contexto
    return render(request, 'terreno/muestreo/graficos_muestreo.html', context)

# Función para generar el gráfico de distribución porcentual (anillo)
def generar_grafico_distribucion_cantidades(df):
    if df.empty:
        return None

    # Agrupar por objetivo y sumar cantidades
    cantidad_por_objetivo = df.groupby('objetivo')['cantidad'].sum()
    total = cantidad_por_objetivo.sum()
    porcentajes = (cantidad_por_objetivo / total * 100).round(1)

    # Crear gráfico de anillo
    fig, ax = plt.subplots()
    wedges, texts, autotexts = ax.pie(porcentajes, labels=porcentajes.index, autopct='%1.1f%%', 
                                      startangle=90, wedgeprops=dict(width=0.3, edgecolor='white'))
    ax.grid(True)
    plt.subplots_adjust(right=0.75)
    num_columns = (len(porcentajes) + 24) // 25
    legend = ax.legend(loc='center left', bbox_to_anchor=(1.05, 0.5), ncol=num_columns)
    ax.set_title("Distribución Porcentual de Cantidades por Objetivo")

    # Guardar en buffer y codificar en base64
    buffer = BytesIO()
    plt.savefig(buffer, format='png', dpi=100, bbox_inches='tight', bbox_extra_artists=[legend])
    buffer.seek(0)
    plt.close()
    return base64.b64encode(buffer.getvalue()).decode('utf-8')

# Función para generar el gráfico de dispersión
def generar_grafico_profundidad_espesor(df):
    if df.empty:
        return None

    # Crear gráfico de dispersión
    fig, ax = plt.subplots()
    ax.scatter(df['profundidad_promedio'], df['espesor_estrato'], c='blue', alpha=0.5)
    ax.set_xlabel('Profundidad Promedio')
    ax.set_ylabel('Espesor de Estrato')
    ax.set_title('Relación entre Profundidad Promedio y Espesor de Estrato')
    ax.grid(True)

    # Guardar en buffer y codificar en base64
    buffer = BytesIO()
    plt.savefig(buffer, format='png', dpi=100, bbox_inches='tight')
    buffer.seek(0)
    plt.close()
    return base64.b64encode(buffer.getvalue()).decode('utf-8')


############ PROGRAMA ##############################################

@login_required
def agregar_programa(request):
    proyectos = Proyectos.objects.all()
    if request.method == 'POST':
        post_data = request.POST.copy()
        proyecto_id = post_data.get('id_proyecto')
        if proyecto_id:
            try:
                proyecto = Proyectos.objects.get(id=proyecto_id)
                post_data['id_proyecto'] = proyecto
            except Proyectos.DoesNotExist:
                post_data['id_proyecto'] = None
        
        form = ProgramaForm(post_data)
        if form.is_valid():
            try:
                programa = form.save(commit=False)
                programa.user = request.user
                programa.save()
                print(f"Programa creado: ID {programa.id}")  # Depuración
                return redirect('listar_programa')
            except Exception as e:
                print(f"Error al guardar programa: {e}")  # Depuración
                form.add_error(None, f"Error al guardar: {e}")
        else:
            print("Formulario no válido:", form.errors)  # Depuración
    else:
        form = ProgramaForm()
    return render(request, 'laboratorio/programa/agregar_programa.html', {'form': form, 'proyectos': proyectos})

# Listar programas
@login_required
def listar_programa(request):
    programas = Programa.objects.all().order_by('id')
    query = request.GET.get('q', '')
    if query:
        Programa_model = apps.get_model('administrador', 'Programa')
        fields = [field.name for field in Programa_model._meta.get_fields() if isinstance(field, (CharField, TextField, DateField, DateTimeField, ForeignKey))]
        query_filter = Q()
        for field in fields:
            if isinstance(Programa_model._meta.get_field(field), (DateField, DateTimeField)):
                query_filter |= Q(**{f"{field}__year__icontains": query})
                query_filter |= Q(**{f"{field}__month__icontains": query})
                query_filter |= Q(**{f"{field}__day__icontains": query})
            elif isinstance(Programa_model._meta.get_field(field), ForeignKey):
                related_model = Programa_model._meta.get_field(field).related_model
                related_fields = [f"{field}__{related_field.name}" for related_field in related_model._meta.get_fields() if isinstance(related_field, (CharField, TextField))]
                for related_field in related_fields:
                    query_filter |= Q(**{f"{related_field}__icontains": query})
            else:
                query_filter |= Q(**{f"{field}__icontains": query})
        programas = programas.filter(query_filter)

    paginator = Paginator(programas, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'laboratorio/programa/programa_table.html', {'page_obj': page_obj})
    return render(request, 'laboratorio/programa/listar_programa.html', {'page_obj': page_obj, 'query': query})

# Exportar a Excel
@login_required
def export_to_excel_programa(request):
    query = request.GET.get('q', '')
    headers = [
        'ID', 'ID Proyecto', 'Tipo Prospección', 'ID Prospección', 'Área', 'Objetivo', 'Cantidad', 
        'Fecha Ingreso Lab', 'ID Ingreso', 'Fecha Envío', 'Asignar Muestra', 'Fecha Informe', 
        'Cantidad Recibida', 'N° Informe', 'N° EP', 'Usuario'
    ]

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="programas_filtrados.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Programas Filtrados"
    ws.append(headers)

    programas = Programa.objects.all().order_by('id')
    if query:
        Programa_model = apps.get_model('administrador', 'Programa')
        fields = [field.name for field in Programa_model._meta.get_fields() if isinstance(field, (CharField, TextField, DateField, DateTimeField, ForeignKey))]
        query_filter = Q()
        for field in fields:
            if isinstance(Programa_model._meta.get_field(field), (DateField, DateTimeField)):
                query_filter |= Q(**{f"{field}__year__icontains": query})
                query_filter |= Q(**{f"{field}__month__icontains": query})
                query_filter |= Q(**{f"{field}__day__icontains": query})
            elif isinstance(Programa_model._meta.get_field(field), ForeignKey):
                related_model = Programa_model._meta.get_field(field).related_model
                related_fields = [f"{field}__{related_field.name}" for related_field in related_model._meta.get_fields() if isinstance(related_field, (CharField, TextField))]
                for related_field in related_fields:
                    query_filter |= Q(**{f"{related_field}__icontains": query})
            else:
                query_filter |= Q(**{f"{field}__icontains": query})
        programas = programas.filter(query_filter)

    for programa in programas:
        row = [
            programa.id,
            str(programa.id_proyecto.id) if programa.id_proyecto else 'N/A',
            programa.tipo_prospeccion or 'N/A',
            str(programa.id_prospeccion.id) if programa.id_prospeccion else 'N/A',
            programa.area or 'N/A',
            programa.objetivo or 'N/A',
            programa.cantidad or 'N/A',
            programa.fecha_ingreso_lab.strftime('%d/%m/%Y') if programa.fecha_ingreso_lab else 'N/A',
            programa.id_ingreso or 'N/A',
            programa.fecha_envio_programa.strftime('%d/%m/%Y') if programa.fecha_envio_programa else 'N/A',
            programa.asignar_muestra or 'N/A',
            programa.fecha_informe.strftime('%d/%m/%Y') if programa.fecha_informe else 'N/A',
            programa.cantidad_recibida or 'N/A',
            programa.n_informe or 'N/A',
            programa.n_ep or 'N/A',
            programa.user.username if programa.user else 'Sin usuario'
        ]
        ws.append(row)

    wb.save(response)
    return response

# Exportar a PDF
@login_required
def export_to_pdf_programa(request):
    query = request.GET.get('q', '')
    headers = [
        'ID', 'ID Proyecto', 'Tipo Prospección', 'ID Prospección', 'Área', 'Objetivo', 'Cantidad', 
        'Fecha Ingreso Lab', 'ID Ingreso', 'Fecha Envío', 'Asignar Muestra', 'Fecha Informe', 
        'Cantidad Recibida', 'N° Informe', 'N° EP', 'Usuario'
    ]

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="programas_filtrados.pdf"'

    p = canvas.Canvas(response)
    p.drawString(100, 750, "Reporte de Programas Filtrados")
    p.drawString(100, 730, " | ".join(headers))

    programas = Programa.objects.all().order_by('id')
    if query:
        Programa_model = apps.get_model('administrador', 'Programa')
        fields = [field.name for field in Programa_model._meta.get_fields() if isinstance(field, (CharField, TextField, DateField, DateTimeField, ForeignKey))]
        query_filter = Q()
        for field in fields:
            if isinstance(Programa_model._meta.get_field(field), (DateField, DateTimeField)):
                query_filter |= Q(**{f"{field}__year__icontains": query})
                query_filter |= Q(**{f"{field}__month__icontains": query})
                query_filter |= Q(**{f"{field}__day__icontains": query})
            elif isinstance(Programa_model._meta.get_field(field), ForeignKey):
                related_model = Programa_model._meta.get_field(field).related_model
                related_fields = [f"{field}__{related_field.name}" for related_field in related_model._meta.get_fields() if isinstance(related_field, (CharField, TextField))]
                for related_field in related_fields:
                    query_filter |= Q(**{f"{related_field}__icontains": query})
            else:
                query_filter |= Q(**{f"{field}__icontains": query})
        programas = programas.filter(query_filter)

    y = 710
    for programa in programas:
        row = [
            str(programa.id),
            str(programa.id_proyecto.id) if programa.id_proyecto else 'N/A',
            programa.tipo_prospeccion or 'N/A',
            str(programa.id_prospeccion.id) if programa.id_prospeccion else 'N/A',
            programa.area or 'N/A',
            programa.objetivo or 'N/A',
            str(programa.cantidad) if programa.cantidad else 'N/A',
            programa.fecha_ingreso_lab.strftime('%d/%m/%Y') if programa.fecha_ingreso_lab else 'N/A',
            programa.id_ingreso or 'N/A',
            programa.fecha_envio_programa.strftime('%d/%m/%Y') if programa.fecha_envio_programa else 'N/A',
            programa.asignar_muestra or 'N/A',
            programa.fecha_informe.strftime('%d/%m/%Y') if programa.fecha_informe else 'N/A',
            str(programa.cantidad_recibida) if programa.cantidad_recibida else 'N/A',
            programa.n_informe or 'N/A',
            programa.n_ep or 'N/A',
            programa.user.username if programa.user else 'Sin usuario'
        ]
        text = " | ".join(row)
        p.drawString(100, y, text[:500])  # Limitar longitud para evitar desbordamiento
        y -= 20
        if y < 50:
            p.showPage()
            y = 750

    p.showPage()
    p.save()
    return response

# Ver programa
@login_required
def ver_programa(request, id):
    programa = get_object_or_404(Programa, id=id)
    return render(request, 'laboratorio/programa/ver_programa.html', {'programa': programa})

# Editar programa
@login_required
def editar_programa(request, id):
    programa = get_object_or_404(Programa, id=id)
    proyectos = Proyectos.objects.all()
    if request.method == 'POST':
        form = ProgramaForm(request.POST, instance=programa)
        if form.is_valid():
            form.save()
            return redirect('listar_programa')
    else:
        form = ProgramaForm(instance=programa)
    return render(request, 'laboratorio/programa/editar_programa.html', {
        'form': form,
        'programa': programa,
        'proyectos': proyectos
    })

# Eliminar programa
@login_required
def eliminar_programa(request, id):
    programa = get_object_or_404(Programa, id=id)
    if request.method == 'POST':
        programa.delete()
        return redirect('listar_programa')
    return render(request, 'laboratorio/programa/eliminar_programa.html', {'programa': programa})

# Estatus programa (sin cambios significativos, ya usa id)
@login_required
def estatus_programa(request):
    id_proyecto = request.GET.get('id_proyecto')
    tipo_prospeccion = request.GET.get('tipo_prospeccion')
    objetivo = request.GET.get('objetivo')
    programas = Programa.objects.all()
    if id_proyecto:
        programas = programas.filter(id_proyecto=id_proyecto)
    if tipo_prospeccion:
        programas = programas.filter(tipo_prospeccion=tipo_prospeccion)
    if objetivo:
        programas = programas.filter(objetivo=objetivo)

    df = pd.DataFrame.from_records(programas.values('objetivo', 'cantidad', 'cantidad_recibida'))
    objetivo_counts = df.groupby('objetivo')['cantidad'].sum().sort_index()
    avance_counts = df.groupby('objetivo')['cantidad_recibida'].sum().sort_index()

    if len(objetivo_counts) != len(avance_counts):
        avance_counts = avance_counts.reindex(objetivo_counts.index, fill_value=0)

    objetivo_counts = objetivo_counts.astype(float)
    avance_counts = avance_counts.astype(float)
    total_count = float(objetivo_counts.sum())
    total_avance = float(avance_counts.sum())

    # Gráfico de barras horizontales
    plt.figure(figsize=(12, len(objetivo_counts) * 0.5))
    indices = np.arange(len(objetivo_counts))
    bar_width = 0.35
    plt.barh(indices, objetivo_counts, bar_width, label='Total', color='lightgray')
    plt.barh(indices, avance_counts, bar_width, label='Avance', color='blue')
    plt.xlabel('Conteo')
    plt.ylabel('Objetivo')
    plt.title('Estatus de Programas')
    plt.yticks(indices, objetivo_counts.index)
    plt.gca().xaxis.set_major_locator(MaxNLocator(integer=True))
    
    # Mover la leyenda fuera del gráfico
    plt.legend(loc='upper left', bbox_to_anchor=(1, 1))
    
    # Etiquetas de valores fuera de las barras
    max_value = objetivo_counts.max()
    for i, (total, avance) in enumerate(zip(objetivo_counts, avance_counts)):
        plt.text(max_value + 0.5, i, f'{avance}', va='center', color='blue')
    
    plt.tight_layout()

    buffer_barras = BytesIO()
    plt.savefig(buffer_barras, format='png')
    buffer_barras.seek(0)
    image_base64_barras = base64.b64encode(buffer_barras.getvalue()).decode('utf-8')
    plt.close()

    # Gráfico de anillo (sin cambios)
    plt.figure(figsize=(4, 4))
    labels = ['Avance', 'Pendiente']
    sizes = [total_avance, total_count - total_avance]
    colors = ['blue', 'lightgray']
    plt.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=140, colors=colors)
    centre_circle = plt.Circle((0, 0), 0.70, fc='white')
    fig = plt.gcf()
    fig.gca().add_artist(centre_circle)
    plt.title('Porcentaje Total de Avance')
    plt.tight_layout()

    buffer_anillo = BytesIO()
    plt.savefig(buffer_anillo, format='png')
    buffer_anillo.seek(0)
    image_base64_anillo = base64.b64encode(buffer_anillo.getvalue()).decode('utf-8')
    plt.close()

    proyectos = Proyectos.objects.all()
    return render(request, 'laboratorio/programa/estatus_programa.html', {
        'image_base64_barras': image_base64_barras,
        'image_base64_anillo': image_base64_anillo,
        'proyectos': proyectos
    })


############ AUDITORIAS ##############################################

# Auditoria prospecciones
def historial(request):
    historiales = Prospecciones.history.all()
    cambios = []
    for historial in historiales:
        if historial.prev_record:
            diff = historial.diff_against(historial.prev_record)
            for change in diff.changes:
                cambios.append({
                    'fecha': historial.history_date,
                    'usuario': historial.history_user,  # Cambiar nomina por usuario
                    'campo': change.field,
                    'valor_anterior': change.old,
                    'valor_nuevo': change.new,
                    'resumen': f"En la fecha {historial.history_date} se modificó el campo {change.field} de {change.old} a {change.new}"
                })
    return render(request, 'historial.html', {'cambios': cambios})







#######NOMINA##########################################################################################
# Configurar logging

logger = logging.getLogger(__name__)

# Agregar nómina
@login_required
def agregar_nomina(request):
    form = NominaForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            try:
                nomina = form.save(commit=False)
                nomina.user = request.user
                nomina.primer_dia = nomina.primer_dia or nomina.fecha_ingreso or date.today()
                nomina.save()
                logger.info(f"Nómina creada: ID {nomina.id}, Turno: {nomina.turno}")
                nomina.generar_roster_para_nomina()
                return redirect('listar_nomina')
            except Exception as e:
                logger.error(f"Error al guardar nómina: {e}")
                form.add_error(None, f"Error al guardar: {e}")
        else:
            logger.error(f"Formulario no válido: {form.errors}")
            print('Formulario no válido', form.errors)
    return render(request, 'nomina/agregar_nomina.html', {'form': form, 'errors': form.errors})

# Listar nómina
@login_required
def listar_nomina(request):
    nomina_list = Nomina.objects.select_related('id_proyecto').all().order_by('id')
    consulta = request.GET.get('q', '')
    if consulta:
        campos = [field.name for field in Nomina._meta.get_fields() 
                  if isinstance(field, (CharField, TextField, DateField, DateTimeField, ForeignKey))]
        filtro_consulta = Q()
        for campo in campos:
            if isinstance(Nomina._meta.get_field(campo), (DateField, DateTimeField)):
                filtro_consulta |= Q(**{f"{campo}__year__icontains": consulta})
                filtro_consulta |= Q(**{f"{campo}__month__icontains": consulta})
                filtro_consulta |= Q(**{f"{campo}__day__icontains": consulta})
            elif isinstance(Nomina._meta.get_field(campo), ForeignKey):
                related_model = Nomina._meta.get_field(campo).related_model
                related_fields = [f"{campo}__{related_field.name}" for related_field in related_model._meta.get_fields() 
                                  if isinstance(related_field, (CharField, TextField))]
                for related_field in related_fields:
                    filtro_consulta |= Q(**{f"{related_field}__icontains": consulta})
            else:
                filtro_consulta |= Q(**{f"{campo}__icontains": consulta})
        nomina_list = nomina_list.filter(filtro_consulta)

    paginator = Paginator(nomina_list, 1000)  # 10 registros por página
    page_number = request.GET.get('page')
    nomina_list = paginator.get_page(page_number)
    return render(request, 'nomina/listar_nomina.html', {'nomina_list': nomina_list, 'consulta': consulta})

from django.http import HttpResponse
from openpyxl import Workbook
def export_to_excel_nomina(request):
    query = request.GET.get('q', '')
    headers_str = request.GET.get('headers', '')
    headers = headers_str.split(',') if headers_str else [
        'ID', 'Proyecto', 'Empresa', 'Fecha Ingreso', 'Nombre', 'Apellido', 'RUT', 'Email',
        'Teléfono', 'Cargo', 'Título', 'Turno', 'Horas Semanales', 'Primer Día'
    ]

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="nomina_filtrada.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Nómina Filtrada"
    ws.append(headers)

    nominas = Nomina.objects.select_related('id_proyecto').all().order_by('id')
    if query:
        campos = [field.name for field in Nomina._meta.get_fields() 
                  if isinstance(field, (CharField, TextField, DateField, DateTimeField, ForeignKey))]
        filtro_consulta = Q()
        for campo in campos:
            if isinstance(Nomina._meta.get_field(campo), (DateField, DateTimeField)):
                filtro_consulta |= Q(**{f"{campo}__year__icontains": query})
                filtro_consulta |= Q(**{f"{campo}__month__icontains": query})
                filtro_consulta |= Q(**{f"{campo}__day__icontains": query})
            elif isinstance(Nomina._meta.get_field(campo), ForeignKey):
                related_model = Nomina._meta.get_field(campo).related_model
                related_fields = [f"{campo}__{related_field.name}" for related_field in related_model._meta.get_fields() 
                                  if isinstance(related_field, (CharField, TextField))]
                for related_field in related_fields:
                    filtro_consulta |= Q(**{f"{related_field}__icontains": query})
            else:
                filtro_consulta |= Q(**{f"{campo}__icontains": query})
        nominas = nominas.filter(filtro_consulta)

    for nomina in nominas:
        row = []
        for header in headers:
            if header == 'ID':
                value = nomina.id
            elif header == 'Proyecto':
                value = str(nomina.id_proyecto) if nomina.id_proyecto else 'N/A'
            elif header == 'Empresa':
                value = nomina.empresa
            elif header == 'Fecha Ingreso':
                value = nomina.fecha_ingreso.strftime('%d/%m/%Y') if nomina.fecha_ingreso else 'N/A'
            elif header == 'Nombre':
                value = nomina.nombre
            elif header == 'Apellido':
                value = nomina.apellido
            elif header == 'RUT':
                value = nomina.rut
            elif header == 'Email':
                value = nomina.email
            elif header == 'Teléfono':
                value = nomina.telefono
            elif header == 'Cargo':
                value = nomina.cargo
            elif header == 'Título':
                value = nomina.titulo
            elif header == 'Turno':
                value = nomina.turno if nomina.turno else 'N/A'
            elif header == 'Horas Semanales':
                value = float(nomina.horas_semanales) if nomina.horas_semanales is not None else 'N/A'
            elif header == 'Primer Día':
                value = nomina.primer_dia.strftime('%d/%m/%Y') if nomina.primer_dia else 'N/A'
            else:
                value = 'N/A'
            row.append(value)
        ws.append(row)

    wb.save(response)
    return response



@login_required
def export_to_pdf_nomina(request):
    query = request.GET.get('q', '')
    headers_str = request.GET.get('headers', '')
    headers = headers_str.split(',') if headers_str else [
        'ID', 'Proyecto', 'Empresa', 'Fecha Ingreso', 'Nombre', 'Apellido', 'RUT', 'Email',
        'Teléfono', 'Cargo', 'Título', 'Turno', 'Horas Semanales', 'Primer Día'
    ]

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="nomina_filtrada.pdf"'

    p = canvas.Canvas(response)
    p.drawString(100, 750, "Reporte de Nómina Filtrada")
    p.drawString(100, 730, " | ".join(headers))

    nominas = Nomina.objects.select_related('id_proyecto').all().order_by('id')
    if query:
        campos = [field.name for field in Nomina._meta.get_fields() 
                  if isinstance(field, (CharField, TextField, DateField, DateTimeField, ForeignKey))]
        filtro_consulta = Q()
        for campo in campos:
            if isinstance(Nomina._meta.get_field(campo), (DateField, DateTimeField)):
                filtro_consulta |= Q(**{f"{campo}__year__icontains": query})
                filtro_consulta |= Q(**{f"{campo}__month__icontains": query})
                filtro_consulta |= Q(**{f"{campo}__day__icontains": query})
            elif isinstance(Nomina._meta.get_field(campo), ForeignKey):
                related_model = Nomina._meta.get_field(campo).related_model
                related_fields = [f"{campo}__{related_field.name}" for related_field in related_model._meta.get_fields() 
                                  if isinstance(related_field, (CharField, TextField))]
                for related_field in related_fields:
                    filtro_consulta |= Q(**{f"{related_field}__icontains": query})
            else:
                filtro_consulta |= Q(**{f"{campo}__icontains": query})
        nominas = nominas.filter(filtro_consulta)

    y = 710
    for nomina in nominas:
        row = []
        for header in headers:
            if header == 'ID':
                value = str(nomina.id)
            elif header == 'Proyecto':
                value = str(nomina.id_proyecto) if nomina.id_proyecto else 'N/A'
            elif header == 'Empresa':
                value = nomina.empresa
            elif header == 'Fecha Ingreso':
                value = nomina.fecha_ingreso.strftime('%d/%m/%Y') if nomina.fecha_ingreso else 'N/A'
            elif header == 'Nombre':
                value = nomina.nombre
            elif header == 'Apellido':
                value = nomina.apellido
            elif header == 'RUT':
                value = nomina.rut
            elif header == 'Email':
                value = nomina.email
            elif header == 'Teléfono':
                value = nomina.telefono
            elif header == 'Cargo':
                value = nomina.cargo
            elif header == 'Título':
                value = nomina.titulo
            elif header == 'Turno':
                value = nomina.turno if nomina.turno else 'N/A'
            elif header == 'Horas Semanales':
                value = str(nomina.horas_semanales) if nomina.horas_semanales is not None else 'N/A'
            elif header == 'Primer Día':
                value = nomina.primer_dia.strftime('%d/%m/%Y') if nomina.primer_dia else 'N/A'
            else:
                value = 'N/A'
            row.append(value)
        text = " | ".join(row)
        p.drawString(100, y, text[:500])  # Limitar longitud para evitar desbordamiento
        y -= 20
        if y < 50:
            p.showPage()
            y = 750

    p.showPage()
    p.save()
    return response

# Editar nómina
@login_required
def editar_nomina(request, id):
    nomina = get_object_or_404(Nomina, id=id, user=request.user)
    proyectos = Proyectos.objects.all()
    if request.method == 'POST':
        form = NominaForm(request.POST, instance=nomina)
        if form.is_valid():
            try:
                nomina = form.save()
                logger.info(f"Nómina editada: ID {nomina.id}, Turno: {nomina.turno}")
                nomina.generar_roster_para_nomina()
                return redirect('listar_nomina')
            except Exception as e:
                logger.error(f"Error al editar nómina {nomina.id}: {e}")
                form.add_error(None, f"Error al editar: {e}")
        else:
            logger.error(f"Formulario no válido al editar nómina {nomina.id}: {form.errors}")
    else:
        form = NominaForm(instance=nomina)
    return render(request, 'nomina/editar_nomina.html', {'form': form, 'nomina': nomina, 'proyectos': proyectos})

# Ver nómina
@login_required
def ver_nomina(request, id):
    nomina = get_object_or_404(Nomina, id=id)
    return render(request, 'nomina/ver_nomina.html', {'nomina': nomina})

# Eliminar nómina
@login_required
def eliminar_nomina(request, id):
    nomina = get_object_or_404(Nomina, id=id, user=request.user)
    if request.method == 'POST':
        try:
            nomina.delete()
            logger.info(f"Nómina eliminada: ID {id}")
            return redirect('listar_nomina')
        except Exception as e:
            logger.error(f"Error al eliminar nómina {id}: {e}")
            return HttpResponse(f"Error al eliminar: {e}", status=500)
    return render(request, 'nomina/eliminar_nomina.html', {'nomina': nomina})




###Listar Roster
@login_required
def listar_roster(request):
    # Obtener consulta de búsqueda si existe
    consulta = request.GET.get('q', '')
    
    roster_list = Roster.objects.select_related('nomina__id_proyecto').all()
    nominas = Nomina.objects.select_related('id_proyecto').all()
    
    # Aplicar filtro dinámico si hay consulta
    if consulta:
        # Crear filtro dinámico para Nomina
        campos_nomina = [field.name for field in Nomina._meta.get_fields() 
                        if isinstance(field, (CharField, TextField, DateField, DateTimeField, ForeignKey))]
        filtro_nomina = Q()
        for campo in campos_nomina:
            if isinstance(Nomina._meta.get_field(campo), (DateField, DateTimeField)):
                filtro_nomina |= Q(**{f"{campo}__year__icontains": consulta})
                filtro_nomina |= Q(**{f"{campo}__month__icontains": consulta})
                filtro_nomina |= Q(**{f"{campo}__day__icontains": consulta})
            elif isinstance(Nomina._meta.get_field(campo), ForeignKey):
                related_model = Nomina._meta.get_field(campo).related_model
                related_fields = [f"{campo}__{related_field.name}" for related_field in related_model._meta.get_fields() 
                                if isinstance(related_field, (CharField, TextField))]
                for related_field in related_fields:
                    filtro_nomina |= Q(**{f"{related_field}__icontains": consulta})
            else:
                filtro_nomina |= Q(**{f"{campo}__icontains": consulta})
        
        # Filtrar las nóminas primero
        nominas = nominas.filter(filtro_nomina)
        
        # Crear filtro dinámico para Roster
        campos_roster = [field.name for field in Roster._meta.get_fields() 
                        if isinstance(field, (CharField, TextField, DateField, DateTimeField, ForeignKey))]
        filtro_roster = Q()
        for campo in campos_roster:
            if isinstance(Roster._meta.get_field(campo), (DateField, DateTimeField)):
                filtro_roster |= Q(**{f"{campo}__year__icontains": consulta})
                filtro_roster |= Q(**{f"{campo}__month__icontains": consulta})
                filtro_roster |= Q(**{f"{campo}__day__icontains": consulta})
            elif isinstance(Roster._meta.get_field(campo), ForeignKey):
                filtro_roster |= Q(**{f"{campo}__id__icontains": consulta})
            else:
                filtro_roster |= Q(**{f"{campo}__icontains": consulta})
        
        # Combinar filtros: buscar en Roster y también en las nóminas filtradas
        roster_list = roster_list.filter(filtro_roster | Q(nomina__in=nominas))

    fechas_unicas = roster_list.values_list('fecha', flat=True).distinct().order_by('fecha')
    dias_semana = {0: 'L', 1: 'M', 2: 'X', 3: 'J', 4: 'V', 5: 'S', 6: 'D'}
    dias_del_año = [
        f"{dias_semana[fecha.weekday()]}-{fecha.strftime('%d-%m-%y')}"
        for fecha in fechas_unicas
    ]
    fechas_unicas_str = [fecha.strftime('%Y-%m-%d') for fecha in fechas_unicas]
    
    sumatoria_dias = [
        roster_list.filter(fecha=fecha).aggregate(Sum('horas_asignadas'))['horas_asignadas__sum'] or 0
        for fecha in fechas_unicas
    ]
    
    for nomina in nominas:
        horas_por_dia = []
        roster_nomina = roster_list.filter(nomina=nomina)
        for fecha in fechas_unicas:
            registro = roster_nomina.filter(fecha=fecha).first()
            horas_por_dia.append(float(registro.horas_asignadas) if registro else 0.0)
        nomina.horas_por_dia_lista = horas_por_dia

    # Obtener meses únicos desde los registros de Roster
    meses_unicos = roster_list.values_list('fecha__year', 'fecha__month').distinct().order_by('fecha__year', 'fecha__month')
    resumen_meses = []

    for año, mes in meses_unicos:
        horas_legales = Decimal('0.0')
        horas_totales = Decimal('0.0')
        horas_efectivas = Decimal('0.0')

        for nomina in nominas:
            jornada_teorica = nomina.get_jornada_teorica()
            if jornada_teorica:
                roster_nomina_mes = roster_list.filter(nomina=nomina, fecha__month=mes, fecha__year=año)
                for registro in roster_nomina_mes:
                    horas_asignadas = Decimal(registro.horas_asignadas)
                    if nomina.turno == '5x2 ordinaria':
                        weekday = registro.fecha.weekday()
                        if weekday < 4:  # Lunes a Jueves
                            legal = Decimal('8.5')
                            total = Decimal('9.0')
                        elif weekday == 4:  # Viernes
                            legal = Decimal('6.0')
                            total = Decimal('6.0')
                        else:
                            legal = total = Decimal('0.0')
                    elif nomina.turno == '4x3 excepcional':
                        if horas_asignadas > 0:
                            legal = Decimal('10.0')
                            total = Decimal('11.0')
                        else:
                            legal = total = Decimal('0.0')
                    else:
                        if horas_asignadas > 0:
                            legal = Decimal(jornada_teorica.horas_legales_diarias)
                            total = Decimal(jornada_teorica.horas_totales_diarias)
                        else:
                            legal = total = Decimal('0.0')
                    horas_legales += legal
                    horas_totales += total
                    horas_efectivas += horas_asignadas

        resumen_meses.append({
            'mes': datetime(año, mes, 1),
            'horas_legales': float(horas_legales),
            'horas_efectivas': float(horas_efectivas),
            'horas_totales': float(horas_totales)
        })

    context = {
        'nominas': nominas,
        'dias_del_año': dias_del_año,
        'fechas_unicas': fechas_unicas_str,
        'sumatoria_dias': sumatoria_dias,
        'resumen_meses': resumen_meses,
        'consulta': consulta,
    }
    return render(request, 'nomina/listar_roster.html', context)


# Actualizar horas en el roster
@csrf_exempt
@login_required
def actualizar_hora(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            updated = []
            totals = {}
            for item in data:
                nomina = Nomina.objects.get(id=item['id'], user=request.user)
                roster = Roster.objects.get(nomina=nomina, fecha=item['dia'])
                horas = Decimal(item['valor'])
                if horas < 0 or horas > 24:
                    raise ValueError("Las horas deben estar entre 0 y 24")
                roster.horas_asignadas = horas
                roster.save()
                logger.info(f"Horas actualizadas para roster de nómina {nomina.id}, fecha {item['dia']}: {horas}")
                
                total_horas = Roster.objects.filter(nomina=nomina).aggregate(Sum('horas_asignadas'))['horas_asignadas__sum'] or 0
                nomina.total_horas = total_horas
                nomina.save()
                
                updated.append({
                    'id': nomina.id,
                    'total_trabajador': float(total_horas)
                })
                
                sumatoria_dia = Roster.objects.filter(fecha=item['dia']).aggregate(Sum('horas_asignadas'))['horas_asignadas__sum'] or 0
                totals[item['dia']] = float(sumatoria_dia)
            
            return JsonResponse({
                'success': True,
                'updated': updated,
                'totals': totals
            })
        except (Nomina.DoesNotExist, Roster.DoesNotExist):
            logger.error(f"Registro no encontrado para actualizar hora: {request.body}")
            return JsonResponse({'success': False, 'error': 'Registro no encontrado'}, status=404)
        except ValueError as e:
            logger.error(f"Error de validación al actualizar hora: {e}")
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
        except Exception as e:
            logger.error(f"Error inesperado al actualizar hora: {e}")
            return JsonResponse({'success': False, 'error': f"Error inesperado: {e}"}, status=500)
    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)




#exportar excel
logger = logging.getLogger(__name__)

@login_required
def export_to_excel_roster(request):
    try:
        consulta = request.GET.get('q', '')
        wb = Workbook()
        ws = wb.active
        ws.title = "Roster"

        # Filtrar datos como en listar_roster
        roster_list = Roster.objects.select_related('nomina__id_proyecto').all()
        nominas = Nomina.objects.select_related('id_proyecto').all()
        
        if consulta:
            campos_nomina = [field.name for field in Nomina._meta.get_fields() 
                            if isinstance(field, (CharField, TextField, DateField, DateTimeField, ForeignKey))]
            filtro_nomina = Q()
            for campo in campos_nomina:
                if isinstance(Nomina._meta.get_field(campo), (DateField, DateTimeField)):
                    filtro_nomina |= Q(**{f"{campo}__year__icontains": consulta})
                    filtro_nomina |= Q(**{f"{campo}__month__icontains": consulta})
                    filtro_nomina |= Q(**{f"{campo}__day__icontains": consulta})
                elif isinstance(Nomina._meta.get_field(campo), ForeignKey):
                    related_model = Nomina._meta.get_field(campo).related_model
                    related_fields = [f"{campo}__{related_field.name}" for related_field in related_model._meta.get_fields() 
                                      if isinstance(related_field, (CharField, TextField))]
                    for related_field in related_fields:
                        filtro_nomina |= Q(**{f"{related_field}__icontains": consulta})
                else:
                    filtro_nomina |= Q(**{f"{campo}__icontains": consulta})
            nominas = nominas.filter(filtro_nomina)
            roster_list = roster_list.filter(Q(nomina__in=nominas))

        # Obtener fechas únicas y calcular horas como en listar_roster
        fechas_unicas = roster_list.values_list('fecha', flat=True).distinct().order_by('fecha')
        dias_semana = ['L', 'M', 'X', 'J', 'V', 'S', 'D']
        dias_del_año = [f"{dias_semana[fecha.weekday()]}-{fecha.strftime('%d-%m-%y')}" for fecha in fechas_unicas]
        headers = ["Proyecto", "Nombre", "RUT", "Cargo", "Turno"] + dias_del_año + ["Total Trabajador"]
        ws.append(headers)

        # Calcular horas por día y totales
        for nomina in nominas:
            roster_nomina = roster_list.filter(nomina=nomina)
            roster_dict = {roster.fecha: float(roster.horas_asignadas) for roster in roster_nomina}
            horas_por_dia = [roster_dict.get(fecha, 0.0) for fecha in fechas_unicas]
            nomina.horas_por_dia_lista = horas_por_dia  # Asignar para consistencia

        totales_por_dia = [0.0] * len(fechas_unicas)
        for nomina in nominas:
            row = [
                str(nomina.id_proyecto) if nomina.id_proyecto else "Sin Proyecto",
                f"{nomina.nombre} {nomina.apellido}",
                nomina.rut,
                nomina.cargo,
                nomina.turno if nomina.turno else "N/A"
            ]
            total_trabajador = 0.0
            for i, horas in enumerate(nomina.horas_por_dia_lista):
                row.append(horas)
                total_trabajador += horas
                totales_por_dia[i] += horas
            row.append(total_trabajador)
            ws.append(row)

        total_row = ["Total por Día", "", "", "", ""] + totales_por_dia + [""]
        ws.append(total_row)

        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col if cell.value)
            column = col[0].column_letter
            ws.column_dimensions[column].width = min(max_length + 2, 15)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="roster.xlsx"'
        wb.save(response)
        return response
    except Exception as e:
        logger.error(f"Error al generar archivo Excel: {str(e)}")
        return HttpResponse(f"Error al generar el archivo Excel: {str(e)}", status=500)
    



#Exportar pdf
logger = logging.getLogger(__name__)

@login_required
def export_to_pdf_roster(request):
    try:
        consulta = request.GET.get('q', '')
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="roster.pdf"'
        doc = SimpleDocTemplate(response, pagesize=landscape(letter))
        elements = []

        # Filtrar datos como en listar_roster
        roster_list = Roster.objects.select_related('nomina__id_proyecto').all()
        nominas = Nomina.objects.select_related('id_proyecto').all()
        
        if consulta:
            campos_nomina = [field.name for field in Nomina._meta.get_fields() 
                            if isinstance(field, (CharField, TextField, DateField, DateTimeField, ForeignKey))]
            filtro_nomina = Q()
            for campo in campos_nomina:
                if isinstance(Nomina._meta.get_field(campo), (DateField, DateTimeField)):
                    filtro_nomina |= Q(**{f"{campo}__year__icontains": consulta})
                    filtro_nomina |= Q(**{f"{campo}__month__icontains": consulta})
                    filtro_nomina |= Q(**{f"{campo}__day__icontains": consulta})
                elif isinstance(Nomina._meta.get_field(campo), ForeignKey):
                    related_model = Nomina._meta.get_field(campo).related_model
                    related_fields = [f"{campo}__{related_field.name}" for related_field in related_model._meta.get_fields() 
                                      if isinstance(related_field, (CharField, TextField))]
                    for related_field in related_fields:
                        filtro_nomina |= Q(**{f"{related_field}__icontains": consulta})
                else:
                    filtro_nomina |= Q(**{f"{campo}__icontains": consulta})
            nominas = nominas.filter(filtro_nomina)
            roster_list = roster_list.filter(Q(nomina__in=nominas))

        # Obtener fechas únicas y calcular horas como en listar_roster
        fechas_unicas = roster_list.values_list('fecha', flat=True).distinct().order_by('fecha')
        dias_semana = ['L', 'M', 'X', 'J', 'V', 'S', 'D']
        dias_del_año = [f"{dias_semana[fecha.weekday()]}-{fecha.strftime('%d-%m-%y')}" for fecha in fechas_unicas]
        headers = ["Proyecto", "Nombre", "RUT", "Cargo", "Turno"] + dias_del_año + ["Total Trabajador"]
        data = [headers]

        # Calcular horas por día y totales
        for nomina in nominas:
            roster_nomina = roster_list.filter(nomina=nomina)
            roster_dict = {roster.fecha: float(roster.horas_asignadas) for roster in roster_nomina}
            horas_por_dia = [roster_dict.get(fecha, 0.0) for fecha in fechas_unicas]
            nomina.horas_por_dia_lista = horas_por_dia  # Asignar para consistencia

        totales_por_dia = [0.0] * len(fechas_unicas)
        for nomina in nominas:
            row = [
                str(nomina.id_proyecto) if nomina.id_proyecto else "Sin Proyecto",
                f"{nomina.nombre} {nomina.apellido}",
                nomina.rut,
                nomina.cargo,
                nomina.turno if nomina.turno else "N/A"
            ]
            total_trabajador = 0.0
            for i, horas in enumerate(nomina.horas_por_dia_lista):
                row.append(str(horas))
                total_trabajador += horas
                totales_por_dia[i] += horas
            row.append(str(total_trabajador))
            data.append(row)

        total_row = ["Total por Día", "", "", "", ""] + [str(total) for total in totales_por_dia] + [""]
        data.append(total_row)

        # Crear la tabla
        table = Table(data, colWidths=[80, 120, 80, 60, 80] + [40] * (len(fechas_unicas) + 1))
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
        ]))
        elements.append(table)

        doc.build(elements)
        return response
    except Exception as e:
        logger.error(f"Error al generar archivo PDF: {str(e)}")
        return HttpResponse(f"Error al generar el archivo PDF: {str(e)}", status=500)
    
    
    
    
from django.shortcuts import render

def docs_view(request):
    return render(request, 'docs.html')  # Renderiza un archivo HTML llamado "docs.html"



#####
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
from io import BytesIO
import base64
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from administrador.models import densidad_insitu, Proyectos, Prospecciones
from .forms import DensidadInsituForm
import json
import random
from django.core.paginator import Paginator
from django.db.models import Q
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

# Funciones de gráficos por área
def generar_grafico_densidad_area(df):
    areas_unicas = df['area'].unique()
    colores = {area: (random.random(), random.random(), random.random()) for area in areas_unicas}
    fig, ax = plt.subplots()
    
    for area in areas_unicas:
        datos_area = df[df['area'] == area]
        ax.scatter(
            datos_area['densidad_natural_del_suelo'],
            datos_area['profundidad_promedio'],
            color=colores[area],
            s=100,
            label=area,
            marker='s'
        )
    
    ax.xaxis.set_ticks_position("top")
    ax.xaxis.set_label_position("top")
    ax.invert_yaxis()
    ax.set_xlabel("Densidad Natural del Suelo (g/cm³)")
    ax.set_ylabel("Profundidad (m)")
    ax.grid(True)
    plt.subplots_adjust(right=0.75)
    num_columns = (len(areas_unicas) + 24) // 25
    legend = ax.legend(loc='center left', bbox_to_anchor=(1, 0.5), ncol=num_columns)
    plt.title("GRÁFICO AGRUPADO POR ÁREA (DISPERSIÓN)", pad=30)
    
    buffer = BytesIO()
    fig.savefig(buffer, format='png', dpi=100, bbox_inches='tight', bbox_extra_artists=[legend])
    buffer.seek(0)
    plt.close()
    image_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
    return image_base64

# Funciones de gráficos por prospección
def generar_grafico_densidad_prospeccion(df):
    df['id_prospeccion'] = df['id_prospeccion'].astype(str)
    df['id_prospeccion_unico'] = df.groupby('id_prospeccion').cumcount() + 1
    df['etiqueta'] = df['id_prospeccion'] + '-' + df['id_prospeccion_unico'].astype(str)
    prospecciones_unicas = df['etiqueta'].unique()
    colores = {prospeccion: (random.random(), random.random(), random.random()) for prospeccion in prospecciones_unicas}
    
    fig, ax = plt.subplots()
    
    for prospeccion in prospecciones_unicas:
        datos_prospeccion = df[df['etiqueta'] == prospeccion]
        ax.scatter(
            datos_prospeccion['densidad_natural_del_suelo'],
            datos_prospeccion['profundidad_promedio'],
            color=colores[prospeccion],
            s=100,
            label=prospeccion,
            marker='s'
        )
    
    ax.xaxis.set_ticks_position("top")
    ax.xaxis.set_label_position("top")
    ax.invert_yaxis()
    ax.set_xlabel("Densidad Natural del Suelo (g/cm³)")
    ax.set_ylabel("Profundidad (m)")
    ax.grid(True)
    plt.subplots_adjust(right=0.75)
    num_columns = (len(prospecciones_unicas) + 24) // 25
    legend = ax.legend(loc='center left', bbox_to_anchor=(1, 0.5), ncol=num_columns)
    plt.title("GRÁFICO AGRUPADO POR PROSPECCIÓN (DISPERSIÓN)", pad=30)
    
    buffer = BytesIO()
    fig.savefig(buffer, format='png', dpi=100, bbox_inches='tight', bbox_extra_artists=[legend])
    buffer.seek(0)
    plt.close()
    image_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
    return image_base64

# Vista principal para gráficos de Densidad Insitu
@login_required
def graficos_densidad_insitu(request):
    id_proyectos = request.GET.getlist('id_proyectos')
    tipos_prospeccion = request.GET.getlist('tipos_prospeccion')
    areas = request.GET.getlist('areas')
    id_prospecciones = request.GET.getlist('id_prospecciones')

    query = densidad_insitu.objects.all()
    if id_proyectos:
        query = query.filter(id_proyecto__in=id_proyectos)
    if tipos_prospeccion:
        query = query.filter(tipo_prospeccion__in=tipos_prospeccion)
    if areas:
        query = query.filter(area__in=areas)
    if id_prospecciones:
        query = query.filter(id_prospeccion__in=id_prospecciones)

    proyectos = query.values('id_proyecto').distinct()
    tipos_prospeccion_inicial = query.values_list('tipo_prospeccion', flat=True).distinct().exclude(tipo_prospeccion__isnull=True)
    areas_inicial = query.values_list('area', flat=True).distinct().exclude(area__isnull=True).exclude(area='')
    id_prospecciones_inicial = query.values_list('id_prospeccion', flat=True).distinct().exclude(id_prospeccion__isnull=True)

    df = pd.DataFrame(list(query.values('densidad_natural_del_suelo', 'profundidad_promedio', 'area', 'id_prospeccion_id')))
    df = df.rename(columns={'id_prospeccion_id': 'id_prospeccion'})

    context = {
        'proyectos': proyectos,
        'tipos_prospeccion_inicial': tipos_prospeccion_inicial,
        'areas_inicial': areas_inicial,
        'id_prospecciones_inicial': id_prospecciones_inicial,
        'selected_id_proyectos': json.dumps(id_proyectos),
        'selected_tipos_prospeccion': json.dumps(tipos_prospeccion),
        'selected_areas': json.dumps(areas),
        'selected_id_prospecciones': json.dumps(id_prospecciones),
    }

    if df.empty:
        context['error'] = "No hay datos disponibles para los filtros seleccionados."
    else:
        context['image_base64_area'] = generar_grafico_densidad_area(df)
        context['image_base64_prospeccion'] = generar_grafico_densidad_prospeccion(df)

    return render(request, 'terreno/ensayos_terreno/densidad_insitu/graficos_densidad_insitu.html', context)

# Funciones auxiliares para Densidad Insitu
def obtener_tipos_prospeccion_densidad_insitu(request):
    id_proyectos = request.GET.get('id_proyectos')
    query = densidad_insitu.objects.all()
    if id_proyectos:
        query = query.filter(id_proyecto__in=id_proyectos.split(','))
    tipos_list = list(query.values_list('tipo_prospeccion', flat=True).distinct().exclude(tipo_prospeccion__isnull=True))
    return JsonResponse({'options': tipos_list}, safe=False)

def obtener_id_prospecciones_densidad_insitu(request):
    id_proyectos = request.GET.get('id_proyectos')
    tipos_prospeccion = request.GET.get('tipos_prospeccion')
    areas = request.GET.get('areas')
    query = densidad_insitu.objects.all()
    if id_proyectos:
        query = query.filter(id_proyecto__in=id_proyectos.split(','))
    if tipos_prospeccion:
        query = query.filter(tipo_prospeccion__in=tipos_prospeccion.split(','))
    if areas:
        query = query.filter(area__in=areas.split(','))
    id_prospecciones_list = list(query.values_list('id_prospeccion', flat=True).distinct().exclude(id_prospeccion__isnull=True))
    return JsonResponse({'options': id_prospecciones_list}, safe=False)

def obtener_area_densidad_insitu(request):
    id_prospecciones = request.GET.get('id_prospecciones')
    if id_prospecciones:
        query = densidad_insitu.objects.filter(id_prospeccion__in=id_prospecciones.split(','))
        areas_list = list(query.values_list('area', flat=True).distinct().exclude(area__isnull=True).exclude(area=''))
        return JsonResponse({'options': areas_list})
    return JsonResponse({'options': []})

# Vista para agregar densidad insitu
from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.core.exceptions import ObjectDoesNotExist
from .models import Muestreo  # Ajusta según tu app

def agregar_densidad_insitu(request):
    """
    Vista para agregar densidad in situ, obteniendo todos los datos desde Muestreo.
    Maneja GET para llenar selectores y POST para guardar datos.
    """
    if request.method == 'GET':
        # Parámetros de consulta
        load_proyectos = request.GET.get('load_proyectos')
        id_proyecto = request.GET.get('id_proyecto')
        tipo_prospeccion = request.GET.get('tipo_prospeccion')
        id_prospeccion = request.GET.get('id_prospeccion')
        id_muestra = request.GET.get('id_muestra')

        if load_proyectos:
            # Cargar proyectos únicos desde Muestreo
            proyectos = Muestreo.objects.values_list('id_proyecto', flat=True).distinct()
            return JsonResponse({'proyectos': list(proyectos)})

        if id_proyecto and not tipo_prospeccion:
            # Cargar tipos de prospección para el proyecto
            tipos = Muestreo.objects.filter(id_proyecto=id_proyecto).values_list('tipo_prospeccion', flat=True).distinct()
            return JsonResponse({'tipo_prospeccion_choices': list(tipos)})

        if id_proyecto and tipo_prospeccion and not id_prospeccion:
            # Cargar prospecciones para el proyecto y tipo
            prospecciones = Muestreo.objects.filter(
                id_proyecto=id_proyecto,
                tipo_prospeccion=tipo_prospeccion
            ).values_list('id_prospeccion', flat=True).distinct()
            return JsonResponse({'prospecciones': list(prospecciones)})

        if id_proyecto and tipo_prospeccion and id_prospeccion and not id_muestra:
            # Cargar muestras para la prospección
            muestras = Muestreo.objects.filter(
                id_proyecto=id_proyecto,
                tipo_prospeccion=tipo_prospeccion,
                id_prospeccion=id_prospeccion
            ).values_list('id_muestra', flat=True).distinct()
            return JsonResponse({'muestras': list(muestras)})

        if id_proyecto and tipo_prospeccion and id_prospeccion and id_muestra:
            # Obtener datos de la muestra específica
            try:
                muestra = Muestreo.objects.get(
                    id_proyecto=id_proyecto,
                    tipo_prospeccion=tipo_prospeccion,
                    id_prospeccion=id_prospeccion,
                    id_muestra=id_muestra
                )
                return JsonResponse({
                    'area': muestra.area or '',
                    'profundidad_desde': muestra.profundidad_desde or '',
                    'profundidad_hasta': muestra.profundidad_hasta or '',
                    'profundidad_promedio': muestra.profundidad_promedio or ''
                })
            except ObjectDoesNotExist:
                return JsonResponse({'error': 'Muestra no encontrada'}, status=404)

    elif request.method == 'POST':
        errors = []
        try:
            # Recolectar datos del formulario
            id_proyecto = request.POST.get('id_proyecto')
            tipo_prospeccion = request.POST.get('tipo_prospeccion')
            data = {
                'id_prospeccion': request.POST.getlist('id_prospeccion[]'),
                'id_muestra': request.POST.getlist('id_muestra[]'),
                'area': request.POST.getlist('area[]'),
                'profundidad_desde': request.POST.getlist('profundidad_desde[]'),
                'profundidad_hasta': request.POST.getlist('profundidad_hasta[]'),
                'profundidad_promedio': request.POST.getlist('profundidad_promedio[]'),
                'profundidad_ensayo': request.POST.getlist('profundidad_ensayo[]'),
                'cota': request.POST.getlist('cota[]'),
                'profundidad_nivel_freatico': request.POST.getlist('profundidad_nivel_freatico[]'),
                'condicion_ambiental': request.POST.getlist('condicion_ambiental[]'),
                'peso_materia_humedo': request.POST.getlist('peso_materia_humedo[]'),
                'masa_arena_inicial': request.POST.getlist('masa_arena_inicial_en_cono_superior[]'),
                'masa_arena_remanente': request.POST.getlist('masa_arena_remanente_en_cono_superior[]'),
                'masa_arena_cono_inferior': request.POST.getlist('masa_arena_en_cono_inferior[]'),
                'masa_arena_excavacion': request.POST.getlist('masa_arena_excavacion[]'),
                'densidad_aparente_arena': request.POST.getlist('densidad_aparente_arena[]'),
                'volumen_perforacion': request.POST.getlist('volumen_perforacion[]'),
                'densidad_natural_suelo': request.POST.getlist('densidad_natural_del_suelo[]'),
                'peso_suelo_humedo': request.POST.getlist('peso_suelo_humedo[]'),
                'peso_suelo_seco': request.POST.getlist('peso_suelo_seco[]'),
                'peso_agua': request.POST.getlist('peso_agua[]'),
                'humedad': request.POST.getlist('humedad[]'),
                'densidad_seca_suelo': request.POST.getlist('densidad_seca_del_suelo[]'),
                'observacion': request.POST.getlist('observacion[]')
            }

            # Validar y procesar cada fila
            for i in range(len(data['id_muestra'])):
                if not data['id_muestra'][i] or not data['observacion'][i]:
                    errors.append(f"Fila {i+1}: ID Muestra y Observación son obligatorios")
                    continue

                # Guardar datos (descomenta y ajusta según tu modelo DensidadInsitu)
                """
                DensidadInsitu.objects.create(
                    id_proyecto_id=id_proyecto,
                    tipo_prospeccion=tipo_prospeccion,
                    id_prospeccion=data['id_prospeccion'][i],
                    id_muestra=data['id_muestra'][i],
                    area=data['area'][i],
                    profundidad_desde=data['profundidad_desde'][i],
                    profundidad_hasta=data['profundidad_hasta'][i],
                    profundidad_promedio=data['profundidad_promedio'][i],
                    profundidad_ensayo=data['profundidad_ensayo'][i],
                    cota=data['cota'][i],
                    profundidad_nivel_freatico=data['profundidad_nivel_freatico'][i],
                    condicion_ambiental=data['condicion_ambiental'][i],
                    peso_materia_humedo=data['peso_materia_humedo'][i],
                    masa_arena_inicial_en_cono_superior=data['masa_arena_inicial'][i],
                    masa_arena_remanente_en_cono_superior=data['masa_arena_remanente'][i],
                    masa_arena_en_cono_inferior=data['masa_arena_cono_inferior'][i],
                    masa_arena_excavacion=data['masa_arena_excavacion'][i],
                    densidad_aparente_arena=data['densidad_aparente_arena'][i],
                    volumen_perforacion=data['volumen_perforacion'][i],
                    densidad_natural_del_suelo=data['densidad_natural_suelo'][i],
                    peso_suelo_humedo=data['peso_suelo_humedo'][i],
                    peso_suelo_seco=data['peso_suelo_seco'][i],
                    peso_agua=data['peso_agua'][i],
                    humedad=data['humedad'][i],
                    densidad_seca_del_suelo=data['densidad_seca_suelo'][i],
                    observacion=data['observacion'][i]
                )
                """

            if errors:
                return render(request, 'terreno/ensayos_terreno/densidad_insitu/agregar_densidad_insitu.html', {
                    'errors': errors
                })

            # Redirigir tras éxito (ajusta la URL según tu proyecto)
            return redirect('agregar_densidad_insitu')

        except Exception as e:
            return render(request, 'terreno/ensayos_terreno/densidad_insitu/agregar_densidad_insitu.html', {
                'errors': [f"Error al procesar los datos: {str(e)}"]
            })

    return render(request, 'terreno/ensayos_terreno/densidad_insitu/agregar_densidad_insitu.html', {})


# Vista para listar densidad insitu
@login_required
def listar_densidad_insitu(request):
    query = request.GET.get('q', '')
    densidades = densidad_insitu.objects.all()
    
    if query:
        densidades = densidades.filter(
            Q(id_proyecto__id__icontains=query) |
            Q(tipo_prospeccion__icontains=query) |
            Q(id_prospeccion__id_prospeccion__icontains=query) |
            Q(id_muestra__icontains=query) |
            Q(area__icontains=query)
        )
    
    paginator = Paginator(densidades, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'terreno/ensayos_terreno/densidad_insitu/listar_densidad_insitu.html', {
        'page_obj': page_obj,
        'query': query
    })

# Vista para ver detalles de densidad insitu
@login_required
def ver_densidad_insitu(request, id):
    densidad = get_object_or_404(densidad_insitu, id=id)
    history_records = densidad.history.all()
    return render(request, 'terreno/ensayos_terreno/densidad_insitu/ver_densidad_insitu.html', {
        'densidad': densidad,
        'history_records': history_records
    })

# Vista para editar densidad insitu
@login_required
def editar_densidad_insitu(request, id):
    densidad = get_object_or_404(densidad_insitu, id=id)
    if request.method == 'POST':
        form = DensidadInsituForm(request.POST, instance=densidad)
        if form.is_valid():
            form.save()
            return redirect('listar_densidad_insitu')
    else:
        form = DensidadInsituForm(instance=densidad)
    return render(request, 'terreno/ensayos_terreno/densidad_insitu/editar_densidad_insitu.html', {
        'form': form,
        'densidad': densidad
    })

# Vista para eliminar densidad insitu
@login_required
def eliminar_densidad_insitu(request, id):
    densidad = get_object_or_404(densidad_insitu, id=id)
    if request.method == 'POST':
        densidad.delete()
        return redirect('listar_densidad_insitu')
    return render(request, 'terreno/ensayos_terreno/densidad_insitu/eliminar_densidad_insitu.html', {
        'densidad': densidad
    })

# Vista para exportar a Excel
@login_required
def export_to_excel_densidad_insitu(request):
    query = request.GET.get('q', '')
    densidades = densidad_insitu.objects.all()
    
    if query:
        densidades = densidades.filter(
            Q(id_proyecto__id__icontains=query) |
            Q(tipo_prospeccion__icontains=query) |
            Q(id_prospeccion__id_prospeccion__icontains=query) |
            Q(id_muestra__icontains=query) |
            Q(area__icontains=query)
        )
    
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Densidad Insitu'
    
    headers = [
        'ID Proyecto', 'Tipo Prospección', 'ID Prospección', 'ID Muestra', 'Profundidad Desde',
        'Profundidad Hasta', 'Profundidad Promedio', 'Profundidad Ensayo', 'Cota',
        'Profundidad Nivel Freático', 'Condición Ambiental', 'Peso Materia Húmedo',
        'Masa Arena Inicial', 'Masa Arena Remanente', 'Masa Arena Cono Inferior',
        'Masa Arena Excavación', 'Densidad Aparente Arena', 'Volumen Perforación',
        'Densidad Natural del Suelo', 'Peso Suelo Húmedo', 'Peso Suelo Seco',
        'Peso Agua', 'Humedad', 'Densidad Seca del Suelo', 'Área', 'Observación', 'Usuario'
    ]
    worksheet.append(headers)
    
    for densidad in densidades:
        worksheet.append([
            str(densidad.id_proyecto.id) if densidad.id_proyecto else '-',
            densidad.tipo_prospeccion or '-',
            str(densidad.id_prospeccion.id_prospeccion) if densidad.id_prospeccion else '-',
            densidad.id_muestra or '-',
            densidad.profundidad_desde or '-',
            densidad.profundidad_hasta or '-',
            densidad.profundidad_promedio or '-',
            densidad.profundidad_ensayo or '-',
            densidad.cota or '-',
            densidad.profundidad_nivel_freatico or '-',
            densidad.condicion_ambiental or '-',
            densidad.peso_materia_humedo or '-',
            densidad.masa_arena_inicial_en_cono_superior or '-',
            densidad.masa_arena_remanente_en_cono_superior or '-',
            densidad.masa_arena_en_cono_inferior or '-',
            densidad.masa_arena_excavacion or '-',
            densidad.densidad_aparente_arena or '-',
            densidad.volumen_perforacion or '-',
            densidad.densidad_natural_del_suelo or '-',
            densidad.peso_suelo_humedo or '-',
            densidad.peso_suelo_seco or '-',
            densidad.peso_agua or '-',
            densidad.humedad or '-',
            densidad.densidad_seca_del_suelo or '-',
            densidad.area or '-',
            densidad.observacion or '-',
            densidad.user.username if densidad.user else 'Sin usuario'
        ])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=densidad_insitu.xlsx'
    workbook.save(response)
    return response

# Vista para exportar a PDF
@login_required
def export_to_pdf_densidad_insitu(request):
    query = request.GET.get('q', '')
    densidades = densidad_insitu.objects.all()
    
    if query:
        densidades = densidades.filter(
            Q(id_proyecto__id__icontains=query) |
            Q(tipo_prospeccion__icontains=query) |
            Q(id_prospeccion__id_prospeccion__icontains=query) |
            Q(id_muestra__icontains=query) |
            Q(area__icontains=query)
        )
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="densidad_insitu.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    headers = [
        'ID Proyecto', 'Tipo Prospección', 'ID Prospección', 'ID Muestra', 'Profundidad Desde',
        'Profundidad Hasta', 'Profundidad Promedio', 'Profundidad Ensayo', 'Cota',
        'Profundidad Nivel Freático', 'Condición Ambiental', 'Peso Materia Húmedo',
        'Masa Arena Inicial', 'Masa Arena Remanente', 'Masa Arena Cono Inferior',
        'Masa Arena Excavación', 'Densidad Aparente Arena', 'Volumen Perforación',
        'Densidad Natural del Suelo', 'Peso Suelo Húmedo', 'Peso Suelo Seco',
        'Peso Agua', 'Humedad', 'Densidad Seca del Suelo', 'Área', 'Observación', 'Usuario'
    ]
    
    data = [headers]
    for densidad in densidades:
        data.append([
            str(densidad.id_proyecto.id) if densidad.id_proyecto else '-',
            densidad.tipo_prospeccion or '-',
            str(densidad.id_prospeccion.id_prospeccion) if densidad.id_prospeccion else '-',
            densidad.id_muestra or '-',
            str(densidad.profundidad_desde) if densidad.profundidad_desde else '-',
            str(densidad.profundidad_hasta) if densidad.profundidad_hasta else '-',
            str(densidad.profundidad_promedio) if densidad.profundidad_promedio else '-',
            str(densidad.profundidad_ensayo) if densidad.profundidad_ensayo else '-',
            str(densidad.cota) if densidad.cota else '-',
            str(densidad.profundidad_nivel_freatico) if densidad.profundidad_nivel_freatico else '-',
            densidad.condicion_ambiental or '-',
            str(densidad.peso_materia_humedo) if densidad.peso_materia_humedo else '-',
            str(densidad.masa_arena_inicial_en_cono_superior) if densidad.masa_arena_inicial_en_cono_superior else '-',
            str(densidad.masa_arena_remanente_en_cono_superior) if densidad.masa_arena_remanente_en_cono_superior else '-',
            str(densidad.masa_arena_en_cono_inferior) if densidad.masa_arena_en_cono_inferior else '-',
            str(densidad.masa_arena_excavacion) if densidad.masa_arena_excavacion else '-',
            str(densidad.densidad_aparente_arena) if densidad.densidad_aparente_arena else '-',
            str(densidad.volumen_perforacion) if densidad.volumen_perforacion else '-',
            str(densidad.densidad_natural_del_suelo) if densidad.densidad_natural_del_suelo else '-',
            str(densidad.peso_suelo_humedo) if densidad.peso_suelo_humedo else '-',
            str(densidad.peso_suelo_seco) if densidad.peso_suelo_seco else '-',
            str(densidad.peso_agua) if densidad.peso_agua else '-',
            str(densidad.humedad) if densidad.humedad else '-',
            str(densidad.densidad_seca_del_suelo) if densidad.densidad_seca_del_suelo else '-',
            densidad.area or '-',
            densidad.observacion or '-',
            densidad.user.username if densidad.user else 'Sin usuario'
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(table)
    doc.build(elements)
    return response